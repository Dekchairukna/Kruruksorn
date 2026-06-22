import os
import re
import uuid
import zipfile
import json
import urllib.request
import urllib.error
from io import BytesIO
from datetime import datetime, date, time, timedelta
import calendar as py_calendar
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, send_from_directory, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.exceptions import BadRequest, BadRequestKeyError
from openpyxl import load_workbook, Workbook
from types import SimpleNamespace
from zoneinfo import ZoneInfo
from sqlalchemy import inspect as sa_inspect

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, 'instance')
# ไฟล์อัปโหลดต้องไม่ผูกกับโฟลเดอร์โค้ดอย่างเดียว เพราะบน Railway เมื่อ redeploy ไฟล์ใน container อาจหายได้
# ถ้ามี Railway Volume ให้ตั้งตัวแปร PERSISTENT_UPLOAD_DIR=/data/uploads หรือ /app/uploads ตามที่ mount ไว้
UPLOAD_DIR = os.environ.get('PERSISTENT_UPLOAD_DIR') or os.environ.get('UPLOAD_DIR') or os.path.join(BASE_DIR, 'uploads')
os.makedirs(INSTANCE_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'krurakson-dev')

@app.errorhandler(BadRequestKeyError)
def handle_bad_request_key_error(e):
    """กันหน้า Bad Request เวลากดปุ่ม/ฟอร์มบางตัวส่งข้อมูลไม่ครบ เช่น field หายหลังปรับ modal/form"""
    try:
        missing = getattr(e, 'args', [''])[0]
    except Exception:
        missing = ''
    flash(f'ระบบได้รับข้อมูลจากฟอร์มไม่ครบ กรุณาลองใหม่อีกครั้ง {"(ขาดช่อง: " + str(missing) + ")" if missing else ""}', 'danger')
    return redirect(request.referrer or url_for('index'))

@app.errorhandler(BadRequest)
def handle_bad_request(e):
    """แสดงข้อความในระบบแทนหน้า Bad Request สีขาวของ Flask"""
    flash('คำขอไม่สมบูรณ์หรือฟอร์มส่งข้อมูลไม่ครบ กรุณาย้อนกลับแล้วลองใหม่อีกครั้ง', 'danger')
    return redirect(request.referrer or url_for('index'))


def get_database_uri():
    """เลือกฐานข้อมูลให้ปลอดภัยสำหรับ Railway PostgreSQL

    - Production/Railway ต้องใช้ PostgreSQL เท่านั้น เพื่อกันเว็บเผลอไปเปิด SQLite ใหม่
    - รองรับหลายชื่อตัวแปรของ Railway/Postgres
    - ถ้าค่า DATABASE_URL ถูกใส่เป็น reference ผิด เช่น ${{Postgres.DATABASE_URL}}
      ระบบจะแจ้ง error ชัดเจนแทนการปล่อยให้ SQLAlchemy ขึ้นข้อความอ่านยาก
    """
    def clean_url(value):
        if not value:
            return None
        value = str(value).strip().strip('"').strip("'")
        # กันกรณีเผลอใส่ทั้งบรรทัดในช่อง VALUE เช่น DATABASE_URL=postgresql://...
        if value.startswith('DATABASE_URL='):
            value = value.split('=', 1)[1].strip().strip('"').strip("'")
        if value.startswith('POSTGRES_URL='):
            value = value.split('=', 1)[1].strip().strip('"').strip("'")
        if value.startswith('POSTGRESQL_URL='):
            value = value.split('=', 1)[1].strip().strip('"').strip("'")
        if value.startswith('postgres://'):
            value = value.replace('postgres://', 'postgresql://', 1)
        return value

    # Railway/Postgres อาจใช้ชื่อ variable ต่างกันตามวิธี attach/reference
    candidate_names = [
        'DATABASE_URL',
        'DATABASE_PRIVATE_URL',
        'DATABASE_PUBLIC_URL',
        'POSTGRES_URL',
        'POSTGRESQL_URL',
        'PGDATABASE_URL',
    ]
    for name in candidate_names:
        database_url = clean_url(os.environ.get(name))
        if not database_url:
            continue
        if database_url.startswith('${{') or database_url.endswith('}}'):
            raise RuntimeError(
                f'{name} ยังเป็นค่า Reference ที่ Railway ไม่ได้ resolve: {database_url} | '
                'ให้ลบตัวแปรนี้ แล้วกด Add Reference จากช่อง VALUE เลือก PostgreSQL -> DATABASE_URL '
                'หรือ copy ค่า postgresql://... จาก service PostgreSQL มาใส่โดยตรง'
            )
        if database_url.startswith('postgresql://'):
            return database_url
        raise RuntimeError(
            f'{name} ไม่ใช่ PostgreSQL URL ที่ถูกต้อง: {database_url[:80]} | '
            'ค่าที่ถูกต้องต้องขึ้นต้นด้วย postgresql:// หรือ postgres://'
        )

    # ถ้ามีตัวแปร PG* แยกกัน ให้ประกอบ connection string ให้เอง
    pg_host = os.environ.get('PGHOST') or os.environ.get('POSTGRES_HOST')
    pg_user = os.environ.get('PGUSER') or os.environ.get('POSTGRES_USER')
    pg_password = os.environ.get('PGPASSWORD') or os.environ.get('POSTGRES_PASSWORD')
    pg_database = os.environ.get('PGDATABASE') or os.environ.get('POSTGRES_DB') or os.environ.get('POSTGRES_DATABASE')
    pg_port = os.environ.get('PGPORT') or os.environ.get('POSTGRES_PORT') or '5432'
    if pg_host and pg_user and pg_password and pg_database:
        from urllib.parse import quote_plus
        return f"postgresql://{quote_plus(pg_user)}:{quote_plus(pg_password)}@{pg_host}:{pg_port}/{pg_database}"

    running_on_railway = bool(
        os.environ.get('RAILWAY_ENVIRONMENT')
        or os.environ.get('RAILWAY_PROJECT_ID')
        or os.environ.get('RAILWAY_SERVICE_ID')
    )
    if running_on_railway:
        visible_db_vars = ', '.join([k for k in os.environ.keys() if 'DATABASE' in k or 'POSTGRES' in k or k.startswith('PG')]) or 'ไม่มี'
        raise RuntimeError(
            'ไม่พบ DATABASE_URL/PostgreSQL variables ใน Railway web service: '
            f'ตัวแปรเกี่ยวกับฐานข้อมูลที่เห็นตอนนี้ = {visible_db_vars}. '
            'ให้ผูก PostgreSQL DATABASE_URL ให้ web ก่อน deploy เพื่อป้องกันระบบเปิดฐานข้อมูล SQLite ใหม่'
        )

    return f"sqlite:///{os.path.join(INSTANCE_DIR, 'krurakson.db')}"

app.config['SQLALCHEMY_DATABASE_URI'] = get_database_uri()
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB รองรับวิดีโอบทเรียน
app.config['OPENAI_MODEL'] = os.environ.get('OPENAI_MODEL', 'gpt-4.1-mini')
app.config['GEMINI_MODEL'] = os.environ.get('GEMINI_MODEL', 'gemini-2.5-flash')
# AI API Key ไม่บันทึกลงฐานข้อมูล อ่านจาก Environment Variable เท่านั้น

# ลดปัญหาเด้งออกจากระบบบ่อย: ให้ session อยู่ได้นานขึ้นและ refresh ทุกครั้งที่ใช้งาน
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=int(os.environ.get('SESSION_DAYS', '30')))
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=int(os.environ.get('REMEMBER_DAYS', '30')))
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['REMEMBER_COOKIE_SAMESITE'] = 'Lax'
# เปิด secure cookie เฉพาะตอนกำหนดเอง เพื่อไม่ทำให้ localhost ใช้ไม่ได้
if os.environ.get('COOKIE_SECURE') == '1':
    app.config['SESSION_COOKIE_SECURE'] = True
    app.config['REMEMBER_COOKIE_SECURE'] = True
ALLOWED_WORKSHEET_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'ppt', 'pptx', 'mp4', 'webm', 'mov'}
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp', 'gif'}

def allowed_worksheet_file(filename):
    return filename and '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_WORKSHEET_EXTENSIONS

def save_uploaded_file(file_obj, subdir='worksheets'):
    if not file_obj or not file_obj.filename:
        return '', ''
    if not allowed_worksheet_file(file_obj.filename):
        raise ValueError('รองรับเฉพาะไฟล์ PDF, PNG, JPG, Word, PowerPoint, MP4, WebM, MOV')
    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], subdir)
    os.makedirs(target_dir, exist_ok=True)
    original = file_obj.filename
    safe = secure_filename(original) or 'upload'
    stamp = datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
    filename = f"{stamp}_{safe}"
    file_obj.save(os.path.join(target_dir, filename))
    return f"{subdir}/{filename}", original


def safe_remove_upload_file(file_path):
    """ลบไฟล์จริงจากโฟลเดอร์ uploads แบบปลอดภัย ไม่ให้ path หลุดออกนอก uploads"""
    if not file_path:
        return False
    try:
        upload_root = os.path.abspath(app.config['UPLOAD_FOLDER'])
        abs_path = os.path.abspath(os.path.join(upload_root, file_path))
        if not abs_path.startswith(upload_root):
            return False
        if os.path.exists(abs_path) and os.path.isfile(abs_path):
            os.remove(abs_path)
            return True
    except Exception:
        pass
    return False


def upload_file_exists(file_path):
    """ตรวจว่าไฟล์แนบยังอยู่จริงหรือไม่ ใช้กันภาพ/ไฟล์หายแต่ข้อมูลค้างในฐานข้อมูล"""
    if not file_path:
        return False
    try:
        upload_root = os.path.abspath(app.config['UPLOAD_FOLDER'])
        abs_path = os.path.abspath(os.path.join(upload_root, file_path))
        if not abs_path.startswith(upload_root):
            return False
        return os.path.exists(abs_path) and os.path.isfile(abs_path)
    except Exception:
        return False

@app.template_filter('upload_file_exists')
def upload_file_exists_filter(file_path):
    return upload_file_exists(file_path)



def infer_subject_level(subject_name=''):
    """เดาระดับชั้นจากชื่อวิชาหรือรหัสวิชา เช่น ว21102 = ม.1, ว22102 = ม.2"""
    text = str(subject_name or '')
    m = re.search(r'ม\.\s*([1-6])', text)
    if m:
        return f"ม.{m.group(1)}"
    m = re.search(r'ว\s*2([1-6])', text)
    if m:
        return f"ม.{m.group(1)}"
    return 'ชั้นเรียน'


def lesson_period_no(lesson):
    """หาลำดับคาบของบทเรียนภายในรายวิชา โดยเรียงตามหน่วยและบทเรียน"""
    try:
        subject_id = lesson.unit.subject_id
        lessons = (
            Lesson.query.join(Unit, Lesson.unit_id == Unit.id)
            .filter(Unit.subject_id == subject_id)
            .order_by(Unit.id.asc(), Lesson.id.asc())
            .all()
        )
        for i, item in enumerate(lessons, start=1):
            if item.id == lesson.id:
                return i
    except Exception:
        pass
    return lesson.id if lesson else ''


def clean_download_text(text):
    text = re.sub(r'[\\/:*?"<>|]+', '-', str(text or '').strip())
    text = re.sub(r'\s+', ' ', text).strip()
    return text[:160] or 'ไฟล์'


def worksheet_display_name(worksheet, with_ext=False):
    """ชื่อไฟล์/ชื่อแสดงผลแบบอ่านง่าย ไม่โชว์ชื่อไฟล์จริงที่ระบบเก็บ"""
    try:
        subject = worksheet.lesson.unit.subject.name
        level = infer_subject_level(subject)
        period = lesson_period_no(worksheet.lesson)
        title = worksheet.title or 'ใบงาน'
        name = f"ใบงานครูรัก {subject} {level} คาบ {period} - {title}"
    except Exception:
        name = 'ใบงานครูรัก'
    name = clean_download_text(name)
    if with_ext and worksheet and worksheet.file_path and '.' in worksheet.file_path:
        ext = worksheet.file_path.rsplit('.', 1)[1].lower()
        name = f"{name}.{ext}"
    return name


@app.template_filter('worksheet_display_name')
def worksheet_display_name_filter(worksheet):
    return worksheet_display_name(worksheet, with_ext=False)


def allowed_image_file(filename):
    return filename and '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


def save_school_logo(file_obj):
    if not file_obj or not file_obj.filename:
        return ''
    if not allowed_image_file(file_obj.filename):
        raise ValueError('ตราโรงเรียนรองรับเฉพาะไฟล์รูปภาพ PNG, JPG, JPEG, WEBP, GIF')
    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'school')
    os.makedirs(target_dir, exist_ok=True)
    safe = secure_filename(file_obj.filename) or 'school-logo.png'
    stamp = datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
    filename = f"{stamp}_{safe}"
    file_obj.save(os.path.join(target_dir, filename))
    return f"school/{filename}"


def detect_lesson_file_type(file_path):
    ext = (file_path or '').rsplit('.', 1)[-1].lower() if '.' in (file_path or '') else ''
    if ext in {'png', 'jpg', 'jpeg'}:
        return 'image'
    if ext == 'pdf':
        return 'pdf'
    if ext in {'doc', 'docx'}:
        return 'document'
    if ext in {'ppt', 'pptx'}:
        return 'presentation'
    if ext in {'mp4', 'webm', 'mov'}:
        return 'video'
    return 'file'


def youtube_embed_url(url):
    """แปลงลิงก์ YouTube ให้ฝังในหน้าบทเรียนได้"""
    if not url:
        return ''
    u = url.strip()
    if 'youtube.com/embed/' in u:
        return u
    if 'youtu.be/' in u:
        vid = u.split('youtu.be/', 1)[1].split('?', 1)[0].split('&', 1)[0].strip('/')
        return f'https://www.youtube.com/embed/{vid}' if vid else ''
    if 'youtube.com/watch' in u and 'v=' in u:
        vid = u.split('v=', 1)[1].split('&', 1)[0]
        return f'https://www.youtube.com/embed/{vid}' if vid else ''
    if 'youtube.com/shorts/' in u:
        vid = u.split('youtube.com/shorts/', 1)[1].split('?', 1)[0].split('&', 1)[0].strip('/')
        return f'https://www.youtube.com/embed/{vid}' if vid else ''
    return ''

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'


ATTENDANCE_STATUSES = ['มา', 'สาย', 'ขาด', 'โดดเรียน', 'ลาป่วย', 'ไปกิจกรรม']
ATTENDANCE_STATUS_ALIASES = {
    'มาสาย': 'สาย',
    'ลา': 'ลาป่วย',
    'กิจกรรม': 'ไปกิจกรรม',
    'present': 'มา',
    'late': 'สาย',
    'absent': 'ขาด',
    'skip': 'โดดเรียน',
    'skipped': 'โดดเรียน',
    'sick': 'ลาป่วย',
    'activity': 'ไปกิจกรรม',
}

# Railway / Render / VPS หลายเจ้ารันเวลาเครื่องเป็น UTC ทำให้แดชบอร์ดคาบเรียนคลาดเคลื่อน
# จึงล็อกเวลาที่ใช้คำนวณคาบเรียนเป็นเวลาไทยเสมอ เว้นแต่กำหนด APP_TIMEZONE เอง
APP_TIMEZONE = os.environ.get('APP_TIMEZONE', 'Asia/Bangkok')
LOCAL_TZ = ZoneInfo(APP_TIMEZONE)

def local_now():
    return datetime.now(LOCAL_TZ)

def local_today():
    return local_now().date()


THAI_WEEKDAYS = ['จันทร์', 'อังคาร', 'พุธ', 'พฤหัสบดี', 'ศุกร์', 'เสาร์', 'อาทิตย์']
THAI_MONTHS = ['', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
ATTENDANCE_SYMBOLS = {
    'มา': '/',
    'สาย': 'ส',
    'ขาด': 'ข',
    'ลาป่วย': 'ลป',
    'โดดเรียน': 'ด',
    'ไปกิจกรรม': 'ก',
}

def thai_date_label(d):
    if not d:
        return ''
    return f"วัน{THAI_WEEKDAYS[d.weekday()]}ที่ {d.day} {THAI_MONTHS[d.month]} {d.year + 543}"

def thai_date_short(d):
    if not d:
        return ''
    return f"{THAI_WEEKDAYS[d.weekday()]} {d.day}/{d.month}/{str(d.year + 543)[-2:]}"

def normalize_attendance_status(value):
    value = (value or 'มา').strip()
    return ATTENDANCE_STATUS_ALIASES.get(value, value if value in ATTENDANCE_STATUSES else 'มา')

def split_thai_full_name(full_name):
    """แยกคำนำหน้า/ชื่อ/นามสกุลแบบง่าย เพื่อช่วยเติมจากเครื่องอ่านบัตรประชาชน"""
    txt = (full_name or '').strip()
    prefixes = ['เด็กชาย', 'เด็กหญิง', 'นาย', 'นางสาว', 'นาง', 'ด.ช.', 'ด.ญ.']
    prefix = ''
    for pf in prefixes:
        if txt.startswith(pf):
            prefix = pf
            txt = txt[len(pf):].strip()
            break
    parts = txt.split()
    first = parts[0] if parts else ''
    last = ' '.join(parts[1:]) if len(parts) > 1 else ''
    return prefix, first, last

def fill_student_personal_fields(user, form):
    """บันทึกข้อมูลพื้นฐานนักเรียนจากฟอร์ม โดยไม่กระทบประวัติเช็กชื่อ/คะแนน"""
    for field in ['citizen_id','birth_date','student_no','prefix','first_name','last_name','gender','nationality','ethnicity','religion','blood_type','phone','address','guardian_name','guardian_phone']:
        if field in form:
            setattr(user, field, (form.get(field) or '').strip())
    # ถ้าแยกชื่อไว้ ให้ประกอบ full_name อัตโนมัติ แต่ยังยอมให้กรอก full_name เองได้
    if (form.get('first_name') or form.get('last_name')) and not form.get('full_name'):
        user.full_name = f"{form.get('prefix','').strip()}{form.get('first_name','').strip()} {form.get('last_name','').strip()}".strip()

def create_or_update_student_from_form(form, default_room_id=None):
    """เพิ่มนักเรียนแบบเร็ว ใช้ได้จากหน้าห้องและหน้าเช็กชื่อ"""
    full_name = (form.get('full_name') or '').strip()
    citizen_id = (form.get('citizen_id') or '').strip()
    username = (form.get('username') or form.get('student_no') or citizen_id or '').strip()
    if not full_name and (form.get('first_name') or form.get('last_name')):
        full_name = f"{form.get('prefix','').strip()}{form.get('first_name','').strip()} {form.get('last_name','').strip()}".strip()
    if not username:
        username = f"student{int(datetime.utcnow().timestamp())}"
    if not full_name:
        raise ValueError('กรุณากรอกชื่อ-สกุลนักเรียน')
    user = None
    if citizen_id:
        user = User.query.filter_by(citizen_id=citizen_id).first()
    if not user:
        user = User.query.filter_by(username=username).first()
    if user and user.role != 'student':
        raise ValueError('username/เลขบัตรนี้มีอยู่แล้ว แต่ไม่ใช่บัญชีนักเรียน')
    if not user:
        base = username[:100]
        username_try = base
        n = 2
        while User.query.filter_by(username=username_try).first():
            username_try = f"{base[:90]}{n}"
            n += 1
        user = User(username=username_try, full_name=full_name, role='student', must_change_password=True)
        user.set_password(form.get('password') or (citizen_id[-6:] if citizen_id and len(citizen_id) >= 6 else '1234'))
        db.session.add(user); db.session.flush()
    else:
        user.full_name = full_name
        user.is_active = True
    fill_student_personal_fields(user, form)
    room_id = form.get('classroom_id', type=int) or default_room_id
    if room_id:
        if not ClassroomStudent.query.filter_by(classroom_id=room_id, student_id=user.id).first():
            db.session.add(ClassroomStudent(classroom_id=room_id, student_id=user.id))
    return user

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='student')  # admin teacher student
    citizen_id = db.Column(db.String(20))
    birth_date = db.Column(db.String(20))
    must_change_password = db.Column(db.Boolean, default=False)
    position = db.Column(db.String(80), default='ครู')
    is_active = db.Column(db.Boolean, default=True)
    # ข้อมูลพื้นฐานนักเรียน/บุคคล เพิ่มแบบไม่ลบข้อมูลเดิม รองรับการอ่าน/กรอกข้อมูลจากบัตรประชาชน
    student_no = db.Column(db.String(30), default='')
    prefix = db.Column(db.String(30), default='')
    first_name = db.Column(db.String(120), default='')
    last_name = db.Column(db.String(120), default='')
    gender = db.Column(db.String(20), default='')
    nationality = db.Column(db.String(80), default='')
    ethnicity = db.Column(db.String(80), default='')
    religion = db.Column(db.String(80), default='')
    blood_type = db.Column(db.String(10), default='')
    phone = db.Column(db.String(50), default='')
    address = db.Column(db.Text, default='')
    guardian_name = db.Column(db.String(255), default='')
    guardian_phone = db.Column(db.String(50), default='')

    def set_password(self, password):
        self.password_hash = generate_password_hash(str(password))

    def check_password(self, password):
        return check_password_hash(self.password_hash, str(password))


class SchoolSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    school_name = db.Column(db.String(255), default='')
    logo_path = db.Column(db.String(500), default='')
    director_name = db.Column(db.String(255), default='')
    deputy_director_name = db.Column(db.String(255), default='')
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


TEACHER_POSITION_CHOICES = ['ครูผู้ช่วย', 'ครู', 'ครู คศ.1', 'ครู คศ.2', 'ครู คศ.3', 'ครูช่วยสอน', 'ครูอัตราจ้าง', 'ผอ.', 'รอง ผอ.', 'ผู้อำนวยการโรงเรียน', 'รองผู้อำนวยการโรงเรียน']


def get_school_setting():
    setting = SchoolSetting.query.first()
    if not setting:
        setting = SchoolSetting(school_name='')
        db.session.add(setting)
        db.session.commit()
    return setting


@app.context_processor
def inject_school_setting():
    try:
        return {'school_setting': get_school_setting(), 'teacher_position_choices': TEACHER_POSITION_CHOICES}
    except Exception:
        return {'school_setting': None, 'teacher_position_choices': TEACHER_POSITION_CHOICES}


class Semester(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    academic_year = db.Column(db.String(20), nullable=False, default='2569')
    term_no = db.Column(db.String(10), nullable=False, default='1')
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    is_active = db.Column(db.Boolean, default=False)

class CalendarImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    semester_id = db.Column(db.Integer, db.ForeignKey('semester.id'))
    file_path = db.Column(db.String(500), nullable=False)
    original_name = db.Column(db.String(255), default='')
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    semester = db.relationship('Semester')

class Classroom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    teacher = db.relationship('User')

class TeacherClassroom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    teacher = db.relationship('User')
    classroom = db.relationship('Classroom')

class ClassroomStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    classroom = db.relationship('Classroom')
    student = db.relationship('User')

class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    credit = db.Column(db.Float, default=1.0)
    total_periods = db.Column(db.Integer, default=40)
    is_active = db.Column(db.Boolean, default=True)
    teacher = db.relationship('User')

class TeacherSubject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    teacher = db.relationship('User')
    subject = db.relationship('Subject')

class SubjectClassroom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    subject = db.relationship('Subject')
    classroom = db.relationship('Classroom')

class Unit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    indicators = db.Column(db.Text, default='')
    required_periods = db.Column(db.Integer, default=1)
    subject = db.relationship('Subject')

class Lesson(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    unit_id = db.Column(db.Integer, db.ForeignKey('unit.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    objective = db.Column(db.Text, default='')
    content = db.Column(db.Text, default='')
    media_url = db.Column(db.String(500), default='')
    required_minutes = db.Column(db.Integer, default=50)
    unit = db.relationship('Unit')


class LessonFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=False)
    file_path = db.Column(db.String(500), default='')
    original_file_name = db.Column(db.String(255), default='')
    file_type = db.Column(db.String(50), default='file')  # image, pdf, document, presentation, file
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    lesson = db.relationship('Lesson')

class Worksheet(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    worksheet_type = db.Column(db.String(50), default='academic') # academic, petanque_score
    description = db.Column(db.Text, default='')
    file_path = db.Column(db.String(500), default='')
    original_file_name = db.Column(db.String(255), default='')
    lesson = db.relationship('Lesson')

class WorksheetQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    worksheet_id = db.Column(db.Integer, db.ForeignKey('worksheet.id'), nullable=False)
    number = db.Column(db.Integer, default=1)
    question_text = db.Column(db.Text, nullable=False)
    answer_type = db.Column(db.String(50), default='text')
    max_score = db.Column(db.Float, default=1)
    worksheet = db.relationship('Worksheet')

class Quiz(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    pass_percent = db.Column(db.Integer, default=60)
    lesson = db.relationship('Lesson')

class QuizQuestion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    quiz_id = db.Column(db.Integer, db.ForeignKey('quiz.id'), nullable=False)
    question_text = db.Column(db.Text, nullable=False)
    question_type = db.Column(db.String(30), default='choice')
    choices = db.Column(db.Text, default='')  # one choice per line
    correct_answer = db.Column(db.Text, default='')
    score = db.Column(db.Float, default=1)
    quiz = db.relationship('Quiz')

class Assignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    due_date = db.Column(db.Date)
    assignment_type = db.Column(db.String(30), default='special')  # special=งานสั่งพิเศษ, lesson=บทเรียนประจำคาบ
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    teacher = db.relationship('User')
    teacher = db.relationship('User')
    subject = db.relationship('Subject')
    classroom = db.relationship('Classroom')
    lesson = db.relationship('Lesson')

class AssignmentStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('assignment.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    status = db.Column(db.String(50), default='งานใหม่')
    lesson_viewed = db.Column(db.Boolean, default=False)
    worksheet_submitted = db.Column(db.Boolean, default=False)
    quiz_submitted = db.Column(db.Boolean, default=False)
    quiz_score = db.Column(db.Float, default=0)
    total_score = db.Column(db.Float, default=0)
    submitted_at = db.Column(db.DateTime)
    assignment = db.relationship('Assignment')
    student = db.relationship('User')

class WorksheetAnswer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('assignment.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('worksheet_question.id'), nullable=False)
    answer_text = db.Column(db.Text, default='')
    file_path = db.Column(db.String(500), default='')
    original_file_name = db.Column(db.String(255), default='')
    score = db.Column(db.Float, default=0)
    question = db.relationship('WorksheetQuestion')

class QuizAnswer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('assignment.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('quiz_question.id'), nullable=False)
    answer_text = db.Column(db.Text, default='')
    is_correct = db.Column(db.Boolean, default=False)
    score = db.Column(db.Float, default=0)

class TeachingSchedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    weekday = db.Column(db.Integer, nullable=False) # 0=Mon
    period_no = db.Column(db.Integer, nullable=False)
    start_time = db.Column(db.String(5), nullable=False)
    end_time = db.Column(db.String(5), nullable=False)
    room_name = db.Column(db.String(100), default='')
    topic = db.Column(db.String(255), default='')
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'))
    subject = db.relationship('Subject')
    classroom = db.relationship('Classroom')
    lesson = db.relationship('Lesson')




class PeriodLessonLog(db.Model):
    """บันทึกว่าคาบเรียนวันใดใช้บทเรียนใด และเปิดให้นักเรียนเห็นหรือยัง"""
    id = db.Column(db.Integer, primary_key=True)
    schedule_id = db.Column(db.Integer, db.ForeignKey('teaching_schedule.id'), nullable=False)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=False)
    taught_date = db.Column(db.Date, nullable=False)
    is_published = db.Column(db.Boolean, default=True)
    taught_at = db.Column(db.DateTime, default=datetime.utcnow)
    taught_by_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    note = db.Column(db.Text, default='')
    schedule = db.relationship('TeachingSchedule')
    lesson = db.relationship('Lesson')
    taught_by = db.relationship('User')


class CalendarEvent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    event_date = db.Column(db.Date, nullable=False)
    title = db.Column(db.String(255), nullable=False)
    event_type = db.Column(db.String(50), default='กิจกรรมโรงเรียน')
    note = db.Column(db.Text, default='')
    teacher = db.relationship('User')

class ClassroomActivity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    activity_type = db.Column(db.String(30), default='special')  # assembly, special
    target_scope = db.Column(db.String(30), default='classroom')  # classroom, all, scout_m1_m3
    event_date = db.Column(db.Date, nullable=True)  # ว่าง = ใช้ได้ทุกวันที่เปิดเรียน
    note = db.Column(db.Text, default='')
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    classroom = db.relationship('Classroom')

class ActivityAttendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    activity_id = db.Column(db.Integer, db.ForeignKey('classroom_activity.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='มา')
    activity = db.relationship('ClassroomActivity')
    classroom = db.relationship('Classroom')
    student = db.relationship('User')

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    # period_no ทำให้วันเดียวกันนับเป็นคนละคาบได้ เช่น คาบ 3 และ 4 เป็น 2 ชั่วโมงเรียน
    period_no = db.Column(db.Integer, nullable=True)
    schedule_id = db.Column(db.Integer, db.ForeignKey('teaching_schedule.id'), nullable=True)
    checked_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    substitute_for_teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    status = db.Column(db.String(20), default='มา') # มา สาย ขาด โดดเรียน ลาป่วย ไปกิจกรรม
    subject = db.relationship('Subject')
    classroom = db.relationship('Classroom')
    student = db.relationship('User', foreign_keys=[student_id])
    schedule = db.relationship('TeachingSchedule')
    checked_by = db.relationship('User', foreign_keys=[checked_by_id])
    substitute_for_teacher = db.relationship('User', foreign_keys=[substitute_for_teacher_id])

class GradeSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    worksheet_weight = db.Column(db.Float, default=30)
    quiz_weight = db.Column(db.Float, default=20)
    attendance_weight = db.Column(db.Float, default=10)
    classwork_weight = db.Column(db.Float, default=10)
    midterm_weight = db.Column(db.Float, default=20)
    final_weight = db.Column(db.Float, default=20)
    subject = db.relationship('Subject')

class ManualScore(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    midterm = db.Column(db.Float, default=0)
    final = db.Column(db.Float, default=0)
    behavior = db.Column(db.Float, default=0)

class ClassworkScoreItem(db.Model):
    """หัวข้อคะแนนรายคาบ เช่น แบบฝึกในคาบ/ทักษะปฏิบัติ/ตรวจงานระหว่างเรียน

    ผูกกับรายวิชา ห้อง วันที่ และคาบเรียน เพื่อให้ครูกรอกคะแนนในหน้าเช็กชื่อ
    แล้วนำคะแนนรวมไปคำนวณในสมุดคะแนนและตัดเกรดได้
    """
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    schedule_id = db.Column(db.Integer, db.ForeignKey('teaching_schedule.id'), nullable=True)
    lesson_id = db.Column(db.Integer, db.ForeignKey('lesson.id'), nullable=True)
    date = db.Column(db.Date, nullable=False)
    period_no = db.Column(db.Integer, nullable=True)
    title = db.Column(db.String(255), nullable=False)
    max_score = db.Column(db.Float, default=10)
    score_type = db.Column(db.String(50), default='classwork')
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    subject = db.relationship('Subject')
    classroom = db.relationship('Classroom')
    schedule = db.relationship('TeachingSchedule')
    lesson = db.relationship('Lesson')
    created_by = db.relationship('User')

class ClassworkScore(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('classwork_score_item.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    score = db.Column(db.Float, default=0)
    note = db.Column(db.String(255), default='')
    item = db.relationship('ClassworkScoreItem')
    student = db.relationship('User')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                flash('ไม่มีสิทธิ์เข้าถึงหน้านี้', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return wrapper
    return deco

def teacher_subject_ids():
    if current_user.role == 'admin':
        return [x.id for x in Subject.query.filter_by(is_active=True).all()]
    # ใช้ทั้งตารางมอบหมายครู + เจ้าของวิชา + ตารางสอนจริง
    # เพื่อกันกรณีนำเข้าตารางสอนแล้ว teacher_subject ยังไม่ถูกสร้าง/ซิงก์ในฐานข้อมูลบนเซิร์ฟเวอร์
    ids = {x.subject_id for x in TeacherSubject.query.filter_by(teacher_id=current_user.id).all()}
    ids.update(x.id for x in Subject.query.filter_by(teacher_id=current_user.id, is_active=True).all())
    ids.update(x.subject_id for x in TeachingSchedule.query.filter_by(teacher_id=current_user.id).all())
    return list(ids)

def teacher_classroom_ids():
    if current_user.role == 'admin':
        return [x.id for x in Classroom.query.filter_by(is_active=True).all()]
    # ใช้ทั้งตารางมอบหมายห้อง + ครูประจำชั้น + ตารางสอนจริง
    ids = {x.classroom_id for x in TeacherClassroom.query.filter_by(teacher_id=current_user.id).all()}
    ids.update(x.id for x in Classroom.query.filter_by(teacher_id=current_user.id, is_active=True).all())
    ids.update(x.classroom_id for x in TeachingSchedule.query.filter_by(teacher_id=current_user.id).all())
    return list(ids)

def teacher_filter(model):
    if current_user.role == 'admin':
        return model.query
    if model.__name__ == 'Subject':
        return model.query.filter(db.or_(model.is_active == True, model.is_active.is_(None)), model.id.in_(teacher_subject_ids() or [-1]))
    if model.__name__ == 'Classroom':
        return model.query.filter(db.or_(model.is_active == True, model.is_active.is_(None)), model.id.in_(teacher_classroom_ids() or [-1]))
    return model.query.filter_by(teacher_id=current_user.id)


def active_users_query():
    """คืน query ผู้ใช้ที่ยังใช้งานอยู่เท่านั้น ใช้กับ PostgreSQL/SQLite ได้"""
    return User.query.filter(db.or_(User.is_active == True, User.is_active.is_(None)))


def active_teachers_query():
    """คืน query ครูที่ยังใช้งานอยู่ สำหรับ dropdown ครู/ตารางสอน/รายวิชา"""
    return active_users_query().filter(User.role == 'teacher')


def active_students_query():
    """คืน query นักเรียนที่ยังใช้งานอยู่ สำหรับนับ/เลือกนักเรียน"""
    return active_users_query().filter(User.role == 'student')


def owns_subject(subject):
    return current_user.role == 'admin' or subject.teacher_id == current_user.id or TeacherSubject.query.filter_by(teacher_id=current_user.id, subject_id=subject.id).first() is not None

def owns_classroom(room):
    return current_user.role == 'admin' or room.teacher_id == current_user.id or TeacherClassroom.query.filter_by(teacher_id=current_user.id, classroom_id=room.id).first() is not None

def classroom_activity_teachers(room_id):
    links = (TeacherClassroom.query
             .filter_by(classroom_id=room_id)
             .join(User, TeacherClassroom.teacher_id == User.id)
             .order_by(User.full_name)
             .all())
    return [link.teacher for link in links if link.teacher]

def build_classroom_teacher_map(room_ids=None):
    q = TeacherClassroom.query.join(User, TeacherClassroom.teacher_id == User.id)
    if room_ids is not None:
        q = q.filter(TeacherClassroom.classroom_id.in_(room_ids or [-1]))
    out = {}
    for link in q.order_by(User.full_name).all():
        out.setdefault(link.classroom_id, []).append(link.teacher)
    return out

def sync_activity_teachers_for_classroom(classroom_id, teacher_ids, keep_teacher_id=None):
    ids = {int(x) for x in teacher_ids if str(x).strip()}
    if keep_teacher_id:
        ids.add(int(keep_teacher_id))

    old_links = TeacherClassroom.query.filter_by(classroom_id=classroom_id).all()
    for link in old_links:
        if link.teacher_id not in ids:
            db.session.delete(link)

    for teacher_id in ids:
        if not TeacherClassroom.query.filter_by(teacher_id=teacher_id, classroom_id=classroom_id).first():
            db.session.add(TeacherClassroom(teacher_id=teacher_id, classroom_id=classroom_id))

def owns_unit(unit):
    return owns_subject(unit.subject)

def owns_lesson(lesson):
    return owns_unit(lesson.unit)

def deny_redirect(endpoint='index'):
    flash('ไม่มีสิทธิ์จัดการข้อมูลนี้', 'danger')
    return redirect(url_for(endpoint))

def get_current_schedule():
    if not current_user.is_authenticated or current_user.role != 'teacher':
        return None, None
    now = local_now()
    weekday = now.weekday()
    schedules = TeachingSchedule.query.filter_by(teacher_id=current_user.id, weekday=weekday).order_by(TeachingSchedule.start_time).all()
    cur = None
    nxt = None
    now_str = now.strftime('%H:%M')
    for s in schedules:
        if s.start_time <= now_str <= s.end_time:
            cur = s
        if s.start_time > now_str and not nxt:
            nxt = s
    return cur, nxt


def current_period_info(target_date=None, teacher_id=None):
    """คืนคาบปัจจุบัน/คาบถัดไปตามเวลาไทย เพื่อใช้แจ้งเตือนหน้าแดชบอร์ด"""
    now = local_now()
    target_date = target_date or now.date()
    now_str = now.strftime('%H:%M')
    weekday = target_date.weekday()
    q = TeachingSchedule.query.filter_by(weekday=weekday)
    if teacher_id:
        q = q.filter_by(teacher_id=teacher_id)
    rows = q.order_by(TeachingSchedule.start_time.asc(), TeachingSchedule.period_no.asc()).all()
    current_rows = [r for r in rows if r.start_time <= now_str <= r.end_time]
    next_row = next((r for r in rows if r.start_time > now_str), None)
    if current_rows:
        first = current_rows[0]
        return {
            'status': 'current', 'period_no': first.period_no, 'start_time': first.start_time,
            'end_time': first.end_time, 'rows': current_rows, 'next': next_row,
            'now_text': now.strftime('%H:%M'), 'date': target_date,
        }
    if next_row:
        return {
            'status': 'before_next', 'period_no': next_row.period_no, 'start_time': next_row.start_time,
            'end_time': next_row.end_time, 'rows': [], 'next': next_row,
            'now_text': now.strftime('%H:%M'), 'date': target_date,
        }
    return {'status': 'finished', 'period_no': None, 'rows': [], 'next': None, 'now_text': now.strftime('%H:%M'), 'date': target_date}



def subject_classrooms_for_user(subject_id=None):
    """คืนค่าห้องที่ผูกกับรายวิชา และกรองตามสิทธิ์ครู"""
    q = SubjectClassroom.query
    if subject_id:
        q = q.filter_by(subject_id=subject_id)
    links = q.all()
    rows = []
    allowed_rooms = set(teacher_classroom_ids()) if current_user.is_authenticated and current_user.role != 'admin' else None
    for link in links:
        if allowed_rooms is None or link.classroom_id in allowed_rooms:
            rows.append(link)
    return rows

def get_grade_setting(subject_id):
    setting = GradeSetting.query.filter_by(subject_id=subject_id).first()
    if not setting:
        setting = GradeSetting(subject_id=subject_id)
        db.session.add(setting)
        db.session.commit()
    return setting

def get_manual_score(subject_id, student_id):
    row = ManualScore.query.filter_by(subject_id=subject_id, student_id=student_id).first()
    if not row:
        row = ManualScore(subject_id=subject_id, student_id=student_id)
        db.session.add(row)
        db.session.flush()
    return row

def calculate_classwork_percent(subject_id, classroom_id, student_id):
    items = ClassworkScoreItem.query.filter_by(subject_id=subject_id, classroom_id=classroom_id).all()
    if not items:
        return 100, 0, 0
    item_ids = [x.id for x in items]
    scores = {x.item_id: x for x in ClassworkScore.query.filter(
        ClassworkScore.item_id.in_(item_ids),
        ClassworkScore.student_id == student_id
    ).all()}
    max_total = sum(max(0, float(x.max_score or 0)) for x in items)
    if max_total <= 0:
        return 100, 0, 0
    raw_total = 0
    for item in items:
        row = scores.get(item.id)
        raw_total += min(max(float(row.score or 0), 0), float(item.max_score or 0)) if row else 0
    return (raw_total / max_total) * 100, raw_total, max_total

def calculate_grade_row(subject, room, student):
    setting = get_grade_setting(subject.id)
    statuses = AssignmentStatus.query.join(Assignment).filter(
        Assignment.subject_id==subject.id,
        Assignment.classroom_id==room.id,
        AssignmentStatus.student_id==student.id
    ).all()
    quiz_avg = round(sum(x.quiz_score or 0 for x in statuses)/len(statuses),2) if statuses else 0
    complete = sum(1 for x in statuses if x.status=='เรียนจบ')
    completion_percent = (complete / max(1, len(statuses))) * 100
    atts = Attendance.query.filter_by(subject_id=subject.id, classroom_id=room.id, student_id=student.id).all()
    normalized_atts = [normalize_attendance_status(a.status) for a in atts]
    present = sum(1 for st in normalized_atts if st == 'มา')
    late = sum(1 for st in normalized_atts if st == 'สาย')
    absent = sum(1 for st in normalized_atts if st == 'ขาด')
    skipped = sum(1 for st in normalized_atts if st == 'โดดเรียน')
    leave = sum(1 for st in normalized_atts if st == 'ลาป่วย')
    activity = sum(1 for st in normalized_atts if st == 'ไปกิจกรรม')
    total_att = len(atts)
    bad_units = absent + skipped + leave + (late * 0.5)
    attendance_percent = max(0, 100 - (bad_units * 5)) if total_att else 100
    manual = get_manual_score(subject.id, student.id)
    classwork_percent, classwork_raw, classwork_max = calculate_classwork_percent(subject.id, room.id, student.id)
    classwork_weight = float(getattr(setting, 'classwork_weight', 0) or 0)
    worksheet_score = completion_percent * (setting.worksheet_weight / 100)
    quiz_score = quiz_avg * (setting.quiz_weight / 100)
    attendance_score = attendance_percent * (setting.attendance_weight / 100)
    classwork_score = classwork_percent * (classwork_weight / 100)
    midterm_score = min(100, manual.midterm or 0) * (setting.midterm_weight / 100)
    final_score = min(100, manual.final or 0) * (setting.final_weight / 100)
    used_weight = setting.worksheet_weight + setting.quiz_weight + setting.attendance_weight + classwork_weight + setting.midterm_weight + setting.final_weight
    behavior_score = min(100, manual.behavior or 0) * (max(0, 100 - used_weight) / 100)
    total = min(100, worksheet_score + quiz_score + attendance_score + classwork_score + midterm_score + final_score + behavior_score)
    return {
        'student': student, 'quiz_avg': quiz_avg, 'complete': complete, 'total_assignments': len(statuses),
        'present': present, 'absent': absent, 'leave': leave, 'late': late, 'activity': activity,
        'attendance_percent': round(attendance_percent, 2), 'manual': manual, 'skipped': skipped,
        'classwork_percent': round(classwork_percent, 2), 'classwork_raw': round(classwork_raw, 2), 'classwork_max': round(classwork_max, 2),
        'worksheet_score': round(worksheet_score,2), 'quiz_score': round(quiz_score,2), 'attendance_score': round(attendance_score,2),
        'classwork_score': round(classwork_score,2), 'midterm_score': round(midterm_score,2), 'final_score': round(final_score,2), 'behavior_score': round(behavior_score,2),
        'total': round(total,2), 'grade': grade_from_score(total)
    }

def grade_from_score(score):
    if score >= 80: return '4'
    if score >= 75: return '3.5'
    if score >= 70: return '3'
    if score >= 65: return '2.5'
    if score >= 60: return '2'
    if score >= 55: return '1.5'
    if score >= 50: return '1'
    return '0'

def update_assignment_complete(status):
    assignment = status.assignment
    quiz = Quiz.query.filter_by(lesson_id=assignment.lesson_id).first()
    pass_ok = True
    if quiz:
        pass_ok = status.quiz_submitted and status.quiz_score >= quiz.pass_percent
    if status.lesson_viewed and status.worksheet_submitted and pass_ok:
        status.status = 'เรียนจบ'
    elif status.worksheet_submitted or status.quiz_submitted or status.lesson_viewed:
        status.status = 'กำลังทำ'


def is_school_blocked_day(target_date, teacher_id=None):
    q = CalendarEvent.query.filter_by(event_date=target_date)
    q = q.filter(CalendarEvent.event_type.in_(['วันหยุดราชการ','วันหยุดโรงเรียน']))
    if teacher_id:
        q = q.filter((CalendarEvent.teacher_id==teacher_id) | (CalendarEvent.teacher_id==None))
    return q.first() is not None or target_date.weekday() >= 5

def ordered_lessons_for_subject(subject_id):
    return Lesson.query.join(Unit).filter(Unit.subject_id==subject_id).order_by(Unit.id.asc(), Lesson.id.asc()).all()

def auto_lesson_for_schedule(schedule_row, target_date=None):
    """ถ้าครูผูกบทเรียนไว้ ใช้บทเรียนนั้นก่อน; ถ้าไม่ผูก ให้เรียงบทเรียนตามวันที่มีเรียนจริงในภาคเรียน"""
    if schedule_row.lesson_id:
        return schedule_row.lesson
    lessons = ordered_lessons_for_subject(schedule_row.subject_id)
    if not lessons:
        return None
    sem = get_active_semester()
    target_date = target_date or date.today()
    start = sem.start_date if sem else target_date
    end = min(target_date, sem.end_date if sem else target_date)
    if end < start:
        return lessons[0]
    cur = start
    count = 0
    while cur <= end:
        if cur.weekday() == schedule_row.weekday and not is_school_blocked_day(cur, schedule_row.teacher_id):
            count += 1
        cur += timedelta(days=1)
    idx = max(0, count - 1)
    return lessons[idx] if idx < len(lessons) else lessons[-1]


def get_period_lesson_log(schedule_row, target_date=None):
    """คืนค่าบันทึกคาบเรียนของตาราง/วันที่ ถ้ามี"""
    if not schedule_row:
        return None
    target_date = target_date or local_today()
    return PeriodLessonLog.query.filter_by(schedule_id=schedule_row.id, taught_date=target_date).first()


def lesson_for_schedule(schedule_row, target_date=None, for_student=False):
    """เลือกบทเรียนสำหรับคาบ
    - ถ้าครูกดเปิด/สอนแล้ว จะใช้บทเรียนนั้น
    - นักเรียนจะเห็นเฉพาะบทเรียนที่ครู published แล้ว
    - ครู/แอดมินยังเห็นบทเรียนอัตโนมัติเพื่อเตรียมกดเปิดได้
    """
    log = get_period_lesson_log(schedule_row, target_date)
    if log and log.lesson_id:
        if for_student and not log.is_published:
            return None
        return log.lesson
    if for_student:
        return None
    return auto_lesson_for_schedule(schedule_row, target_date)


def student_can_view_lesson(lesson, student):
    """นักเรียนเปิดบทเรียนได้เมื่อครูกดเปิดบทเรียนนั้นให้ห้องของนักเรียนแล้ว"""
    if not lesson or not student or getattr(student, 'role', None) != 'student':
        return False
    room_ids = [x.classroom_id for x in ClassroomStudent.query.filter_by(student_id=student.id).all()]
    if not room_ids:
        return False
    return db.session.query(PeriodLessonLog.id).join(TeachingSchedule, PeriodLessonLog.schedule_id == TeachingSchedule.id).filter(
        PeriodLessonLog.lesson_id == lesson.id,
        PeriodLessonLog.is_published == True,
        TeachingSchedule.classroom_id.in_(room_ids),
        TeachingSchedule.subject_id == lesson.unit.subject_id,
    ).first() is not None


def create_lesson_assignment_for_classroom(lesson, schedule_row):
    """สร้าง AssignmentStatus ให้นักเรียนในห้อง เพื่อให้ใบงาน/แบบทดสอบของบทเรียนนั้นเปิดทำได้ทันที"""
    if not lesson or not schedule_row:
        return 0
    ass = Assignment.query.filter_by(
        lesson_id=lesson.id,
        classroom_id=schedule_row.classroom_id,
        assignment_type='lesson'
    ).first()
    if not ass:
        ass = Assignment(
            teacher_id=schedule_row.teacher_id,
            subject_id=schedule_row.subject_id,
            classroom_id=schedule_row.classroom_id,
            lesson_id=lesson.id,
            title=f'บทเรียนประจำคาบ: {lesson.title}',
            assignment_type='lesson'
        )
        db.session.add(ass)
        db.session.flush()
    created = 0
    students = ClassroomStudent.query.filter_by(classroom_id=schedule_row.classroom_id).all()
    for cs in students:
        st = AssignmentStatus.query.filter_by(assignment_id=ass.id, student_id=cs.student_id).first()
        if not st:
            db.session.add(AssignmentStatus(assignment_id=ass.id, student_id=cs.student_id, status='กำลังเรียน', lesson_viewed=False))
            created += 1
    return created


def build_student_learning_plan(student, limit=80):
    room_ids = [x.classroom_id for x in ClassroomStudent.query.filter_by(student_id=student.id).all()]
    schedules = TeachingSchedule.query.filter(TeachingSchedule.classroom_id.in_(room_ids or [-1])).order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all()
    sem = get_active_semester()
    if not sem:
        return []
    rows=[]
    cur = sem.start_date
    while cur <= sem.end_date and len(rows) < limit:
        for srow in [x for x in schedules if x.weekday == cur.weekday()]:
            if is_school_blocked_day(cur, srow.teacher_id):
                continue
            les = lesson_for_schedule(srow, cur, for_student=True)
            rows.append(SimpleNamespace(date=cur, schedule=srow, lesson=les))
        cur += timedelta(days=1)
    return rows

def get_or_create_lesson_status(lesson, student):
    # หา/สร้างงานภายในสำหรับบทเรียนประจำคาบ ไม่ให้ปนกับงานพิเศษ
    room_ids = [x.classroom_id for x in ClassroomStudent.query.filter_by(student_id=student.id).all()]
    subject_id = lesson.unit.subject_id
    sched = TeachingSchedule.query.filter(TeachingSchedule.subject_id==subject_id, TeachingSchedule.classroom_id.in_(room_ids or [-1])).first()
    classroom_id = sched.classroom_id if sched else (room_ids[0] if room_ids else None)
    teacher_id = sched.teacher_id if sched else (lesson.unit.subject.teacher_id or User.query.filter_by(role='admin').first().id)
    if not classroom_id:
        return None
    ass = Assignment.query.filter_by(lesson_id=lesson.id, classroom_id=classroom_id, assignment_type='lesson').first()
    if not ass:
        ass = Assignment(teacher_id=teacher_id, subject_id=subject_id, classroom_id=classroom_id, lesson_id=lesson.id, title=f'บทเรียนประจำคาบ: {lesson.title}', assignment_type='lesson')
        db.session.add(ass); db.session.flush()
    st = AssignmentStatus.query.filter_by(assignment_id=ass.id, student_id=student.id).first()
    if not st:
        st = AssignmentStatus(assignment_id=ass.id, student_id=student.id, status='กำลังเรียน', lesson_viewed=True)
        db.session.add(st); db.session.flush()
    else:
        st.lesson_viewed = True
        update_assignment_complete(st)
    db.session.commit()
    return st

# -----------------------------
# Phase AI Assistant
# -----------------------------
def openai_key_ready():
    return bool((os.environ.get('OPENAI_API_KEY') or '').strip())


def openai_masked_status():
    key = (os.environ.get('OPENAI_API_KEY') or '').strip()
    if not key:
        return 'ยังไม่ได้ตั้งค่า OPENAI_API_KEY'
    if len(key) <= 10:
        return 'ตั้งค่าแล้ว'
    return f"ตั้งค่าแล้ว ({key[:4]}...{key[-4:]})"


def gemini_key_ready():
    return bool((os.environ.get('GEMINI_API_KEY') or os.environ.get('GOOGLE_API_KEY') or '').strip())


def gemini_masked_status():
    key = (os.environ.get('GEMINI_API_KEY') or os.environ.get('GOOGLE_API_KEY') or '').strip()
    if not key:
        return 'ยังไม่ได้ตั้งค่า GEMINI_API_KEY'
    if len(key) <= 10:
        return 'ตั้งค่าแล้ว'
    return f"ตั้งค่าแล้ว ({key[:6]}...{key[-4:]})"


def ai_selected_provider():
    raw = (os.environ.get('AI_PROVIDER') or '').strip().lower()
    if raw in ('gemini', 'google', 'google-gemini'):
        return 'gemini'
    if raw in ('openai', 'gpt'):
        return 'openai'
    if raw in ('offline', 'template'):
        return 'offline'
    # auto mode: ใช้ Gemini ก่อน เพราะมี Free Tier เหมาะกับงานครูเริ่มต้น
    if gemini_key_ready():
        return 'gemini'
    if openai_key_ready():
        return 'openai'
    return 'offline'


def ai_key_ready():
    provider = ai_selected_provider()
    if provider == 'gemini':
        return gemini_key_ready()
    if provider == 'openai':
        return openai_key_ready()
    return True


def ai_status_text():
    provider = ai_selected_provider()
    if provider == 'gemini':
        return 'Gemini: ' + gemini_masked_status()
    if provider == 'openai':
        return 'OpenAI: ' + openai_masked_status()
    return 'Offline Template Mode: ใช้งานได้โดยไม่ใช้ API แต่ไม่ใช่ AI จริง'


def ai_model_text():
    provider = ai_selected_provider()
    if provider == 'gemini':
        return os.environ.get('GEMINI_MODEL') or app.config.get('GEMINI_MODEL') or 'gemini-2.5-flash'
    if provider == 'openai':
        return os.environ.get('OPENAI_MODEL') or app.config.get('OPENAI_MODEL') or 'gpt-4.1-mini'
    return 'offline-template'


def extract_openai_text(data):
    """รองรับรูปแบบคำตอบของ Responses API หลายแบบ เพื่อไม่ให้หน้า AI พังถ้า API เปลี่ยน field ย่อย"""
    if not isinstance(data, dict):
        return ''
    if data.get('output_text'):
        return str(data.get('output_text') or '').strip()
    parts = []
    for item in data.get('output', []) or []:
        for content in item.get('content', []) or []:
            if isinstance(content, dict):
                text = content.get('text') or content.get('value') or ''
                if text:
                    parts.append(str(text))
    if parts:
        return '\n'.join(parts).strip()
    # fallback สำหรับบาง SDK/endpoint ที่อาจคืน choices แบบ Chat Completions
    choices = data.get('choices') or []
    if choices:
        msg = choices[0].get('message') or {}
        if msg.get('content'):
            return str(msg.get('content')).strip()
    return ''


def call_openai_text(prompt, *, max_output_tokens=2600):
    """เรียก OpenAI ผ่าน HTTPS ตรง ๆ ด้วย stdlib เพื่อลด dependency เพิ่มเติมในโปรเจกต์เดิม"""
    api_key = (os.environ.get('OPENAI_API_KEY') or '').strip()
    if not api_key:
        raise RuntimeError('ยังไม่ได้ตั้งค่า OPENAI_API_KEY ใน Environment Variables')
    model = (os.environ.get('OPENAI_MODEL') or app.config.get('OPENAI_MODEL') or 'gpt-4.1-mini').strip()
    payload = {
        'model': model,
        'instructions': (
            'คุณคือ AI Assistant สำหรับครูไทยในระบบ KRURUKSORN ช่วยสรุปการสอน '
            'สร้างใบความรู้ และสร้างใบงานเป็นภาษาไทยแบบพร้อมให้ครูตรวจแก้ก่อนบันทึกจริง '
            'ห้ามสั่งแก้ไขฐานข้อมูล ห้ามอ้างว่าบันทึกให้แล้ว และให้ใช้รูปแบบที่ครูคัดลอกไปใช้ได้ทันที'
        ),
        'input': prompt,
        'max_output_tokens': max_output_tokens,
    }
    req = urllib.request.Request(
        'https://api.openai.com/v1/responses',
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        },
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=70) as resp:
            raw = resp.read().decode('utf-8')
            data = json.loads(raw)
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace') if e.fp else ''
        raise RuntimeError(f'OpenAI API error {e.code}: {body[:600]}')
    except urllib.error.URLError as e:
        raise RuntimeError(f'เชื่อมต่อ OpenAI API ไม่สำเร็จ: {e}')
    text = extract_openai_text(data)
    if not text:
        raise RuntimeError('OpenAI API ตอบกลับมาแล้ว แต่ไม่พบข้อความ output')
    return text


def extract_gemini_text(data):
    if not isinstance(data, dict):
        return ''
    parts = []
    for cand in data.get('candidates', []) or []:
        content = cand.get('content') or {}
        for part in content.get('parts', []) or []:
            text = part.get('text') if isinstance(part, dict) else None
            if text:
                parts.append(str(text))
    return '\n'.join(parts).strip()


def call_gemini_text(prompt, *, max_output_tokens=2600):
    """เรียก Gemini ผ่าน REST API โดยไม่เพิ่ม dependency เพื่อให้ deploy ง่ายบน Railway"""
    api_key = (os.environ.get('GEMINI_API_KEY') or os.environ.get('GOOGLE_API_KEY') or '').strip()
    if not api_key:
        raise RuntimeError('ยังไม่ได้ตั้งค่า GEMINI_API_KEY ใน Environment Variables')
    model = (os.environ.get('GEMINI_MODEL') or app.config.get('GEMINI_MODEL') or 'gemini-2.5-flash').strip()
    endpoint = 'https://generativelanguage.googleapis.com/v1beta/models/' + urllib.parse.quote(model, safe='') + ':generateContent?key=' + urllib.parse.quote(api_key, safe='')
    system_text = (
        'คุณคือ AI Assistant สำหรับครูไทยในระบบ KRURUKSORN ช่วยสรุปการสอน '
        'สร้างใบความรู้ และสร้างใบงานเป็นภาษาไทยแบบพร้อมให้ครูตรวจแก้ก่อนบันทึกจริง '
        'ห้ามสั่งแก้ไขฐานข้อมูล ห้ามอ้างว่าบันทึกให้แล้ว และให้ใช้รูปแบบที่ครูคัดลอกไปใช้ได้ทันที'
    )
    payload = {
        'contents': [
            {
                'role': 'user',
                'parts': [{'text': system_text + '\n\n' + prompt}],
            }
        ],
        'generationConfig': {
            'temperature': 0.45,
            'maxOutputTokens': max_output_tokens,
        },
    }
    req = urllib.request.Request(
        endpoint,
        data=json.dumps(payload).encode('utf-8'),
        headers={'Content-Type': 'application/json'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=70) as resp:
            raw = resp.read().decode('utf-8')
            data = json.loads(raw)
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8', errors='replace') if e.fp else ''
        raise RuntimeError(f'Gemini API error {e.code}: {body[:700]}')
    except urllib.error.URLError as e:
        raise RuntimeError(f'เชื่อมต่อ Gemini API ไม่สำเร็จ: {e}')
    text = extract_gemini_text(data)
    if not text:
        raise RuntimeError('Gemini API ตอบกลับมาแล้ว แต่ไม่พบข้อความ output')
    return text


def offline_template_text(prompt):
    """โหมดสำรอง: ไม่ใช่ AI จริง แต่ช่วยให้หน้าทำงานได้แม้ยังไม่มี API"""
    return (
        'โหมดแม่แบบออฟไลน์ (ยังไม่ได้ใช้ AI จริง)\n\n'
        'ระบบเตรียมข้อมูลสำหรับครูไว้แล้ว กรุณานำ Prompt ด้านล่างไปปรับใช้หรือตั้งค่า GEMINI_API_KEY เพื่อให้ AI สร้างคำตอบจริง\n\n'
        '--- Prompt ที่ระบบเตรียมไว้ ---\n'
        + prompt
    )


def call_ai_text(prompt, *, max_output_tokens=2600):
    provider = ai_selected_provider()
    if provider == 'gemini':
        return call_gemini_text(prompt, max_output_tokens=max_output_tokens)
    if provider == 'openai':
        return call_openai_text(prompt, max_output_tokens=max_output_tokens)
    return offline_template_text(prompt)



def ai_subjects_for_user():
    if not current_user.is_authenticated:
        return []
    q = Subject.query.filter(db.or_(Subject.is_active == True, Subject.is_active.is_(None)))
    if current_user.role == 'admin':
        return q.order_by(Subject.name.asc()).all()
    ids = teacher_subject_ids() or [-1]
    return q.filter(Subject.id.in_(ids)).order_by(Subject.name.asc()).all()


def ai_classrooms_for_user():
    if not current_user.is_authenticated:
        return []
    q = Classroom.query.filter(db.or_(Classroom.is_active == True, Classroom.is_active.is_(None)))
    if current_user.role == 'admin':
        return q.order_by(Classroom.name.asc()).all()
    ids = teacher_classroom_ids() or [-1]
    return q.filter(Classroom.id.in_(ids)).order_by(Classroom.name.asc()).all()


def ai_lessons_for_user(subject_id=None):
    q = Lesson.query.join(Unit).join(Subject).filter(db.or_(Subject.is_active == True, Subject.is_active.is_(None)))
    if subject_id:
        q = q.filter(Subject.id == int(subject_id))
    if current_user.role != 'admin':
        q = q.filter(Subject.id.in_(teacher_subject_ids() or [-1]))
    return q.order_by(Subject.name.asc(), Unit.id.asc(), Lesson.id.asc()).limit(300).all()


def ai_get_selected_context(subject_id=None, lesson_id=None, classroom_id=None):
    subject = Subject.query.get(int(subject_id)) if subject_id else None
    lesson = Lesson.query.get(int(lesson_id)) if lesson_id else None
    classroom = Classroom.query.get(int(classroom_id)) if classroom_id else None

    if subject and not owns_subject(subject):
        raise PermissionError('ไม่มีสิทธิ์ใช้รายวิชานี้กับ AI')
    if lesson and not owns_lesson(lesson):
        raise PermissionError('ไม่มีสิทธิ์ใช้บทเรียนนี้กับ AI')
    if classroom and current_user.role != 'admin' and not owns_classroom(classroom):
        raise PermissionError('ไม่มีสิทธิ์ใช้ห้องเรียนนี้กับ AI')

    lines = []
    if subject:
        lines.append(f'รายวิชา: {subject.name} | หน่วยกิต {subject.credit} | จำนวนคาบ {subject.total_periods}')
        units = Unit.query.filter_by(subject_id=subject.id).order_by(Unit.id.asc()).limit(12).all()
        if units:
            lines.append('หน่วยการเรียนรู้ในรายวิชา:')
            for u in units:
                ind = (u.indicators or '').strip()
                lines.append(f'- {u.title} ({u.required_periods} คาบ)' + (f' | ตัวชี้วัด: {ind[:160]}' if ind else ''))
    if lesson:
        lines.append(f'บทเรียนที่เลือก: {lesson.title}')
        if lesson.objective:
            lines.append(f'จุดประสงค์: {lesson.objective[:800]}')
        if lesson.content:
            lines.append(f'เนื้อหาเดิม: {lesson.content[:1400]}')
    if classroom:
        lines.append(f'ห้องเรียน/กลุ่มผู้เรียน: {classroom.name}')
    return '\n'.join(lines).strip(), subject, lesson, classroom


def ai_recent_teaching_context(subject_id=None, classroom_id=None, limit=18):
    q = PeriodLessonLog.query.join(TeachingSchedule, PeriodLessonLog.schedule_id == TeachingSchedule.id)
    if current_user.role != 'admin':
        q = q.filter(db.or_(PeriodLessonLog.taught_by_id == current_user.id, TeachingSchedule.teacher_id == current_user.id))
    if subject_id:
        q = q.filter(TeachingSchedule.subject_id == int(subject_id))
    if classroom_id:
        q = q.filter(TeachingSchedule.classroom_id == int(classroom_id))
    rows = q.order_by(PeriodLessonLog.taught_date.desc(), PeriodLessonLog.taught_at.desc()).limit(limit).all()
    if not rows:
        return 'ยังไม่พบประวัติบันทึกการสอนในระบบตามเงื่อนไขที่เลือก'
    lines = []
    for row in rows:
        schedule = row.schedule
        lesson = row.lesson
        subject_name = schedule.subject.name if schedule and schedule.subject else '-'
        room_name = schedule.classroom.name if schedule and schedule.classroom else '-'
        period_no = schedule.period_no if schedule else '-'
        note = (row.note or '').strip()
        lines.append(f'- {row.taught_date.strftime("%d/%m/%Y")} คาบ {period_no} | {subject_name} | {room_name} | {lesson.title if lesson else "-"}' + (f' | หมายเหตุ: {note[:140]}' if note else ''))
    return '\n'.join(lines)


def ai_build_prompt(task, topic='', level='', extra='', subject_context='', recent_context=''):
    topic = (topic or '').strip()
    level = (level or '').strip()
    extra = (extra or '').strip()
    common = f"""
ข้อมูลจากระบบ KRURUKSORN
{subject_context or '-'}

ประวัติ/บริบทการสอนล่าสุด
{recent_context or '-'}

หัวข้อที่ครูกรอกเพิ่ม: {topic or '-'}
ระดับชั้น/กลุ่มเป้าหมาย: {level or '-'}
รายละเอียดเพิ่มเติมจากครู: {extra or '-'}
""".strip()
    if task == 'summary':
        return common + """

งานที่ต้องทำ:
สรุปให้ครูเห็นว่า "สอนถึงไหนแล้ว" จากข้อมูลที่มี โดยจัดเป็นหัวข้อสั้น อ่านง่าย ประกอบด้วย
1) สถานะการสอนล่าสุด
2) บทเรียน/คาบที่สอนไปแล้ว
3) สิ่งที่ควรสอนต่อในคาบถัดไป
4) งาน/ใบงาน/คะแนนที่ควรติดตาม
5) ข้อเสนอแนะสำหรับครู
ถ้าข้อมูลไม่พอ ให้ระบุว่าข้อมูลส่วนใดขาด ห้ามเดาเป็นข้อเท็จจริง
"""
    if task == 'knowledge':
        return common + """

งานที่ต้องทำ:
สร้างใบความรู้ภาษาไทยสำหรับนักเรียน แบบพร้อมคัดลอกไปวางในระบบ โดยจัดรูปแบบดังนี้
- ชื่อใบความรู้
- รายวิชา/ระดับชั้น
- จุดประสงค์การเรียนรู้ 3 ข้อ
- สาระสำคัญแบบเข้าใจง่าย
- เนื้อหาเป็นหัวข้อย่อย มีตัวอย่างใกล้ตัวนักเรียน
- คำศัพท์สำคัญ
- สรุปท้ายบท
- คำถามชวนคิด 3 ข้อ
ใช้ภาษาครูไทย อ่านง่าย ไม่ยาวเกินจำเป็น
"""
    if task == 'worksheet':
        return common + """

งานที่ต้องทำ:
สร้างใบงานภาษาไทย แบบพร้อมคัดลอกไปวางในระบบ โดยจัดรูปแบบดังนี้
- ชื่อใบงาน
- รายวิชา/ระดับชั้น
- จุดประสงค์
- คำชี้แจง
- กิจกรรมที่ 1: ตรวจความเข้าใจ 5 ข้อ
- กิจกรรมที่ 2: วิเคราะห์สถานการณ์ 3 ข้อ
- กิจกรรมที่ 3: ลงมือปฏิบัติ/ออกแบบ/สรุปคำตอบ 1 งาน
- เกณฑ์การให้คะแนนรวม 10 คะแนน
- แนวคำตอบย่อสำหรับครู
ให้เหมาะกับชั้นเรียนจริง และครูต้องนำไปตรวจแก้ก่อนบันทึก
"""
    return common


@app.route('/ai', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'teacher')
def ai_assistant():
    subjects = ai_subjects_for_user()
    classrooms = ai_classrooms_for_user()
    selected_subject_id = request.form.get('subject_id') or request.args.get('subject_id') or ''
    selected_classroom_id = request.form.get('classroom_id') or request.args.get('classroom_id') or ''
    selected_lesson_id = request.form.get('lesson_id') or request.args.get('lesson_id') or ''
    lessons = ai_lessons_for_user(selected_subject_id or None)
    result_text = ''
    prompt_text = ''
    selected_task = request.form.get('task') or ''
    topic = request.form.get('topic', '')
    level = request.form.get('level', '')
    extra = request.form.get('extra', '')

    if request.method == 'POST':
        try:
            subject_context, subject, lesson, classroom = ai_get_selected_context(
                selected_subject_id or None,
                selected_lesson_id or None,
                selected_classroom_id or None,
            )
            recent_context = ai_recent_teaching_context(
                selected_subject_id or None,
                selected_classroom_id or None,
            )
            prompt_text = ai_build_prompt(selected_task, topic=topic, level=level, extra=extra, subject_context=subject_context, recent_context=recent_context)
            result_text = call_ai_text(prompt_text)
            flash('AI สร้างข้อความให้แล้ว กรุณาตรวจแก้และกดบันทึกเองในหน้าที่เกี่ยวข้อง', 'success')
        except PermissionError as e:
            flash(str(e), 'danger')
        except Exception as e:
            flash(str(e), 'danger')
            if prompt_text:
                result_text = 'ยังไม่ได้สร้างผลลัพธ์จาก AI\n\nระบบเตรียม Prompt ไว้ให้ตรวจสอบก่อน:\n\n' + prompt_text

    return render_template(
        'ai_assistant.html',
        subjects=subjects,
        classrooms=classrooms,
        lessons=lessons,
        result_text=result_text,
        prompt_text=prompt_text,
        selected_task=selected_task,
        selected_subject_id=str(selected_subject_id or ''),
        selected_classroom_id=str(selected_classroom_id or ''),
        selected_lesson_id=str(selected_lesson_id or ''),
        topic=topic,
        level=level,
        extra=extra,
        openai_ready=openai_key_ready(),
        openai_status=openai_masked_status(),
        openai_model=os.environ.get('OPENAI_MODEL') or app.config.get('OPENAI_MODEL'),
        ai_ready=ai_key_ready(),
        ai_status=ai_status_text(),
        ai_provider=ai_selected_provider(),
        ai_model=ai_model_text(),
    )

@app.route('/')
def index():
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    if current_user.role == 'student':
        return redirect(url_for('student_dashboard'))
    if current_user.role == 'teacher':
        return redirect(url_for('teacher_dashboard'))
    return redirect(url_for('admin_dashboard'))

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        user = User.query.filter_by(username=username).first()
        if user and (not getattr(user, 'is_active', True)):
            flash('บัญชีนี้ถูกปิดใช้งานแล้ว กรุณาติดต่อผู้ดูแลระบบ', 'danger')
        elif user and user.check_password(password):
            session.permanent = True
            login_user(user, remember=True, duration=app.config['REMEMBER_COOKIE_DURATION'])
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/worksheet/<int:worksheet_id>/file')
@login_required
def worksheet_file(worksheet_id):
    worksheet = Worksheet.query.get_or_404(worksheet_id)
    if not worksheet.file_path:
        flash('ใบงานนี้ยังไม่มีไฟล์แนบ', 'warning')
        return redirect(url_for('lesson_detail', lesson_id=worksheet.lesson_id))
    # ครู/แอดมินต้องเป็นเจ้าของรายวิชา ส่วนนักเรียนต้องเข้าถึงผ่านงานของตนเองเท่านั้น
    if current_user.role in ['teacher', 'admin']:
        if not owns_lesson(worksheet.lesson):
            return deny_redirect('subjects')
    elif current_user.role == 'student':
        room_ids = [x.classroom_id for x in ClassroomStudent.query.filter_by(student_id=current_user.id).all()]
        ok = AssignmentStatus.query.join(Assignment).filter(
            AssignmentStatus.student_id == current_user.id,
            Assignment.lesson_id == worksheet.lesson_id
        ).first()
        if not ok:
            flash('ยังไม่มีสิทธิ์เปิดไฟล์ใบงานนี้', 'danger')
            return redirect(url_for('student_dashboard'))
    path = os.path.join(app.config['UPLOAD_FOLDER'], worksheet.file_path)
    if not os.path.exists(path):
        flash('ไม่พบไฟล์ใบงาน', 'danger')
        return redirect(url_for('lesson_detail', lesson_id=worksheet.lesson_id))
    return send_file(path, as_attachment=False, download_name=worksheet_display_name(worksheet, with_ext=True))


@app.route('/school-settings', methods=['GET', 'POST'])
@login_required
@role_required('teacher', 'admin')
def school_settings():
    setting = get_school_setting()
    if request.method == 'POST':
        if current_user.role == 'admin':
            setting.school_name = request.form.get('school_name', '').strip()
            setting.director_name = request.form.get('director_name', '').strip()
            setting.deputy_director_name = request.form.get('deputy_director_name', '').strip()
            logo = request.files.get('logo')
            if logo and logo.filename:
                try:
                    setting.logo_path = save_school_logo(logo)
                except ValueError as e:
                    flash(str(e), 'danger')
                    return redirect(url_for('school_settings'))
        current_user.position = request.form.get('position', current_user.position or 'ครู').strip() or 'ครู'
        db.session.commit()
        flash('บันทึกการตั้งค่าโรงเรียนแล้ว', 'success')
        return redirect(url_for('school_settings'))
    return render_template('school_settings.html', setting=setting, position_choices=TEACHER_POSITION_CHOICES)



def get_active_semester():
    sem = Semester.query.filter_by(is_active=True).first()
    if not sem:
        sem = Semester.query.order_by(Semester.start_date.desc()).first()
    return sem

def thai_month_name(m):
    return {1:'มกราคม',2:'กุมภาพันธ์',3:'มีนาคม',4:'เมษายน',5:'พฤษภาคม',6:'มิถุนายน',7:'กรกฎาคม',8:'สิงหาคม',9:'กันยายน',10:'ตุลาคม',11:'พฤศจิกายน',12:'ธันวาคม'}.get(m, str(m))

def buddhist_year(d):
    return d.year + 543

def month_range_between(start, end):
    months = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        months.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return months

def build_calendar_dashboard(year=2026, months=(5,6,7,8,9,10), teacher_id=None, semester=None):
    semester = semester or get_active_semester()
    if semester:
        start_date, end_date = semester.start_date, semester.end_date
        month_pairs = month_range_between(start_date, end_date)
    else:
        month_pairs = [(year, m) for m in months]
        start_date, end_date = date(year, min(months), 1), date(year, max(months), 31)

    q = CalendarEvent.query.filter(CalendarEvent.event_date>=start_date, CalendarEvent.event_date<=end_date)
    if teacher_id:
        q = q.filter((CalendarEvent.teacher_id==teacher_id) | (CalendarEvent.teacher_id==None))
    events_by_date = {}
    for e in q.order_by(CalendarEvent.event_date).all():
        events_by_date.setdefault(e.event_date, []).append(e)

    # ใช้วันที่จริงของเครื่อง ไม่บังคับให้กระโดดไปวันเปิดเทอม
    # ถ้าวันจริงอยู่นอกช่วงเดือนที่แสดง จะไม่ไฮไลต์ผิดวัน
    real_today = local_today()
    today = real_today
    blocks = []
    today_position = None
    for y, m in month_pairs:
        weeks = py_calendar.Calendar(firstweekday=6).monthdayscalendar(y, m)
        rows = []
        for wi, week in enumerate(weeks, start=1):
            cells = []
            for di, day_no in enumerate(week):
                d = date(y, m, day_no) if day_no else None
                evs = events_by_date.get(d, []) if d else []
                if d == today:
                    today_position = {
                        'date': d, 'month_name': thai_month_name(m), 'week_no': wi,
                        'weekday_name': ['อาทิตย์','จันทร์','อังคาร','พุธ','พฤหัสบดี','ศุกร์','เสาร์'][di]
                    }
                cells.append({'day': day_no, 'date': d, 'events': evs, 'weekday': di, 'is_today': d == today})
            rows.append(cells)
        blocks.append({'year': y, 'be_year': y+543, 'month': m, 'month_name': thai_month_name(m), 'weeks': rows})
    upcoming_base = today if start_date <= today <= end_date else start_date
    upcoming = [e for day in sorted(events_by_date) for e in events_by_date[day] if day >= upcoming_base][:8]
    if today_position is None:
        today_position = {
            'date': today,
            'month_name': thai_month_name(today.month),
            'week_no': '-',
            'weekday_name': ['จันทร์','อังคาร','พุธ','พฤหัสบดี','ศุกร์','เสาร์','อาทิตย์'][today.weekday()],
            'out_of_range': True
        }
    return blocks, upcoming, today_position



# -----------------------------------------------------------------------------
# KRURUKSORN SCHOOL ERP HUB
# ศูนย์รวมโมดูลระบบบริหารโรงเรียน: เพิ่มเป็น prototype แบบไม่รื้อระบบเดิม
# หมายเหตุ: ไม่ใส่ KRURUK SPORTS ตามที่ผู้ใช้ระบุว่า "13 ไม่เอา"
# -----------------------------------------------------------------------------

def _erp_feature(title, detail='', status='พร้อมใช้งาน', endpoint=None, href=None, tag=''):
    return {
        'title': title,
        'detail': detail,
        'status': status,
        'endpoint': endpoint,
        'href': href,
        'tag': tag,
    }


def build_school_erp_modules():
    """คืนค่าโครงสร้างเมนู School ERP ประมาณ 50 หน้า โดยผูกบางหน้าเข้ากับระบบเดิมแล้ว"""
    modules = [
        {
            'key': 'academic-grading',
            'icon': '📚',
            'title': 'งานวิชาการและวัดผล',
            'subtitle': 'หลักสูตร รายวิชา คะแนน เกรด และเอกสาร ปพ.',
            'color': 'green',
            'features': [
                _erp_feature('รายวิชา/หลักสูตร', 'จัดการรายวิชา หน่วยการเรียนรู้ ตัวชี้วัด และบทเรียน', endpoint='subjects', tag='เดิม'),
                _erp_feature('บันทึกคะแนน/เช็กชื่อรายวิชา', 'เลือกห้องและรายวิชาเพื่อบันทึกเวลาเรียน คะแนนในคาบ และคะแนนรวม', endpoint='records_center', tag='เดิม'),
                _erp_feature('สมุดคะแนนและคำนวณเกรด', 'คำนวณจากใบงาน แบบทดสอบ คะแนนในคาบ กลางภาค ปลายภาค และจิตพิสัย', endpoint='records_center', tag='เดิม'),
                _erp_feature('เอกสาร ปพ.5', 'ต้นแบบหน้ารวมเพื่อพิมพ์สมุดบันทึกผลการพัฒนาคุณภาพผู้เรียน', status='ต้นแบบ'),
                _erp_feature('เอกสาร ปพ.6', 'ต้นแบบรายงานผลการเรียนรายบุคคลสำหรับผู้ปกครอง', status='ต้นแบบ'),
                _erp_feature('เอกสาร ปพ.7', 'ต้นแบบหนังสือรับรองผลการศึกษา/สถานภาพนักเรียน', status='ต้นแบบ'),
                _erp_feature('สถิติผลสัมฤทธิ์', 'ภาพรวมค่าเฉลี่ย เกรด และจำนวนผู้เรียนรายห้อง/รายวิชา', endpoint='records_center', tag='เดิม'),
            ],
        },
        {
            'key': 'teacher-work',
            'icon': '👨‍🏫',
            'title': 'ระบบครูผู้สอน',
            'subtitle': 'ตารางสอน แผนการสอน ใบงาน แบบทดสอบ และงานที่สั่ง',
            'color': 'blue',
            'features': [
                _erp_feature('ตารางสอน', 'ตารางสอนครูและห้องเรียน 8 คาบ/วัน', endpoint='schedule', tag='เดิม'),
                _erp_feature('สอนแทน', 'ดูคาบที่ต้องสอนแทนและบันทึกการเข้าสอน', endpoint='substitute_schedule', tag='เดิม'),
                _erp_feature('ใบความรู้/บทเรียน', 'จัดการบทเรียน สื่อ วิดีโอ PDF รูปภาพ และเอกสารประกอบ', endpoint='subjects', tag='เดิม'),
                _erp_feature('ใบงานออนไลน์', 'สร้างใบงาน ตรวจงาน และให้คะแนน', endpoint='subjects', tag='เดิม'),
                _erp_feature('แบบทดสอบออนไลน์', 'สร้างแบบทดสอบและตรวจคะแนนอัตโนมัติ', endpoint='subjects', tag='เดิม'),
                _erp_feature('สั่งงานนักเรียน', 'มอบหมายงานรายห้อง/รายวิชา', endpoint='assign', tag='เดิม'),
            ],
        },
        {
            'key': 'student-care',
            'icon': '🫶',
            'title': 'งานกิจการนักเรียนและระบบดูแลช่วยเหลือ',
            'subtitle': 'ประวัตินักเรียน SDQ เยี่ยมบ้าน พฤติกรรม และกลุ่มเสี่ยง',
            'color': 'pink',
            'features': [
                _erp_feature('ฐานข้อมูลนักเรียน', 'ข้อมูลพื้นฐาน นักเรียน ผู้ปกครอง สุขภาพ และข้อมูลติดต่อ', endpoint='classrooms', tag='เดิม'),
                _erp_feature('เช็กชื่อเข้าเรียน', 'มา สาย ลา ขาด ไปกิจกรรม พร้อมรายงานรายวัน', endpoint='attendance_dashboard', tag='เดิม'),
                _erp_feature('แจ้งเตือนผู้ปกครอง', 'เตรียมต่อ LINE/Email จากข้อมูลการขาดเรียนและสาย', status='ต่อยอด'),
                _erp_feature('เยี่ยมบ้านออนไลน์', 'แบบฟอร์มครูและผู้ปกครอง พร้อมรูปบ้านและแผนที่', status='ต้นแบบ'),
                _erp_feature('แบบประเมิน SDQ', 'นักเรียน ผู้ปกครอง และครูทำแบบประเมินออนไลน์', status='ต้นแบบ'),
                _erp_feature('พฤติกรรม/ความประพฤติ', 'บันทึกความดี ความผิด คะแนนพฤติกรรม และหมายเหตุรายบุคคล', status='ต้นแบบ'),
                _erp_feature('ทุนการศึกษา/กลุ่มช่วยเหลือ', 'จัดกลุ่มนักเรียนที่ต้องติดตามและช่วยเหลือ', status='ต้นแบบ'),
            ],
        },
        {
            'key': 'parent-portal',
            'icon': '👪',
            'title': 'ระบบผู้ปกครอง',
            'subtitle': 'ผู้ปกครองดูคะแนน การมาเรียน การบ้าน และประกาศ',
            'color': 'orange',
            'features': [
                _erp_feature('Parent Portal', 'หน้าเข้าสู่ระบบสำหรับผู้ปกครอง', status='ต้นแบบ'),
                _erp_feature('ดูผลการเรียน', 'แสดงคะแนนและเกรดของนักเรียนแบบอ่านง่าย', status='ต้นแบบ'),
                _erp_feature('ดูการมาเรียน', 'สรุปมา สาย ลา ขาด และไปกิจกรรม', status='ต้นแบบ'),
                _erp_feature('ดูการบ้าน/งานค้าง', 'ติดตามงานที่ครูสั่งและสถานะการส่งงาน', status='ต้นแบบ'),
                _erp_feature('แจ้งเตือน LINE', 'แจ้งขาดเรียน งานค้าง คะแนนต่ำ และประกาศใหม่', status='ต่อยอด'),
            ],
        },
        {
            'key': 'student-activities',
            'icon': '🎯',
            'title': 'กิจกรรมพัฒนาผู้เรียน',
            'subtitle': 'ชุมนุม ลูกเสือ เนตรนารี จิตอาสา และกิจกรรมโรงเรียน',
            'color': 'purple',
            'features': [
                _erp_feature('ลงทะเบียนชุมนุม/ชมรม', 'นักเรียนเลือกชุมนุมผ่านเว็บ จำกัดจำนวน และตรวจสอบสิทธิ์', status='ต้นแบบ'),
                _erp_feature('รายชื่อนักเรียนในชุมนุม', 'ครูที่ปรึกษาดูรายชื่อและส่งออก Excel', status='ต้นแบบ'),
                _erp_feature('บันทึกกิจกรรมลูกเสือ/เนตรนารี', 'บันทึกการเข้าร่วมรายวัน/รายกิจกรรม', status='ต้นแบบ'),
                _erp_feature('กิจกรรมเพื่อสังคม/จิตอาสา', 'เก็บชั่วโมงกิจกรรมและรายงานนักเรียน', status='ต้นแบบ'),
                _erp_feature('กิจกรรมประจำห้อง', 'เช็กชื่อกิจกรรมหน้าเสาธง/กิจกรรมพิเศษ', endpoint='classrooms', tag='เดิม'),
            ],
        },
        {
            'key': 'student-registry',
            'icon': '🗂️',
            'title': 'งานทะเบียนนักเรียน',
            'subtitle': 'รับนักเรียน จัดห้อง ย้ายเข้า ย้ายออก และเอกสารรับรอง',
            'color': 'green',
            'features': [
                _erp_feature('จัดการห้องเรียน', 'สร้างห้อง ครูประจำชั้น และรายชื่อนักเรียน', endpoint='classrooms', tag='เดิม'),
                _erp_feature('นำเข้านักเรียน', 'นำเข้ารายชื่อนักเรียนจาก Excel', endpoint='import_students', tag='เดิม'),
                _erp_feature('ย้ายห้อง/เลื่อนชั้น', 'ย้ายห้องรายบุคคลและเตรียมต่อยอดเลื่อนชั้นทั้งระบบ', endpoint='classrooms', tag='เดิม'),
                _erp_feature('ย้ายเข้า/ย้ายออก', 'แบบฟอร์มทะเบียนรับย้ายและจำหน่ายนักเรียน', status='ต้นแบบ'),
                _erp_feature('หนังสือรับรองนักเรียน', 'ต้นแบบเอกสารรับรองสถานภาพนักเรียน', status='ต้นแบบ'),
            ],
        },
        {
            'key': 'savings',
            'icon': '🏦',
            'title': 'ระบบออมทรัพย์นักเรียน',
            'subtitle': 'ฝาก ถอน สมุดบัญชี และรายงานยอดเงิน',
            'color': 'blue',
            'features': [
                _erp_feature('บัญชีออมทรัพย์นักเรียน', 'เปิดบัญชีรายคนและดูยอดคงเหลือ', status='ต้นแบบ'),
                _erp_feature('บันทึกฝากเงิน', 'บันทึกเงินฝากรายวัน รายห้อง หรือรายบุคคล', status='ต้นแบบ'),
                _erp_feature('บันทึกถอนเงิน', 'อนุมัติและบันทึกการถอนเงิน', status='ต้นแบบ'),
                _erp_feature('สมุดบัญชี', 'พิมพ์ประวัติฝากถอนรายคน', status='ต้นแบบ'),
                _erp_feature('รายงานการเงิน', 'สรุปรายวัน รายเดือน รายห้อง และทั้งโรงเรียน', status='ต้นแบบ'),
            ],
        },
        {
            'key': 'e-saraban',
            'icon': '📨',
            'title': 'ระบบสารบรรณอิเล็กทรอนิกส์',
            'subtitle': 'หนังสือรับ หนังสือส่ง คำสั่ง ประกาศ บันทึกข้อความ และเกษียน',
            'color': 'orange',
            'features': [
                _erp_feature('ทะเบียนหนังสือรับ', 'ลงรับหนังสือ เลขรับ วันที่รับ และผู้รับผิดชอบ', status='ต้นแบบ'),
                _erp_feature('ทะเบียนหนังสือส่ง', 'เลขส่ง หน่วยงานปลายทาง และไฟล์แนบ PDF', status='ต้นแบบ'),
                _erp_feature('คำสั่งโรงเรียน', 'สร้างคำสั่งจาก Template และออกเลขอัตโนมัติ', status='ต้นแบบ'),
                _erp_feature('ประกาศโรงเรียน', 'ประกาศภายใน/ภายนอก พร้อมไฟล์แนบ', status='ต้นแบบ'),
                _erp_feature('บันทึกข้อความ', 'ร่างและพิมพ์บันทึกข้อความรูปแบบราชการ', status='ต้นแบบ'),
                _erp_feature('เกษียนออนไลน์', 'เกษียนบนเอกสารจริงและติดตามสถานะ', status='ต้นแบบ'),
            ],
        },
        {
            'key': 'personnel',
            'icon': '🧑‍💼',
            'title': 'ระบบบุคลากร',
            'subtitle': 'ข้อมูลครู ภาระงาน เวรประจำวัน ไปราชการ และลางาน',
            'color': 'pink',
            'features': [
                _erp_feature('ข้อมูลครูและบุคลากร', 'บัญชีครู ตำแหน่ง และสิทธิ์การใช้งาน', endpoint='users', tag='เดิม'),
                _erp_feature('มอบหมายครู', 'ผูกครูกับรายวิชาและห้องเรียน', endpoint='teacher_assignments', tag='เดิม'),
                _erp_feature('ภาระงานสอน', 'ดึงข้อมูลจากตารางสอนและรายวิชา', endpoint='schedule', tag='เดิม'),
                _erp_feature('เวรประจำวัน', 'ต้นแบบจัดเวรครูและพิมพ์ตารางเวร', status='ต้นแบบ'),
                _erp_feature('ไปราชการ/ลางาน', 'บันทึกวันลาและวันไปราชการ', status='ต้นแบบ'),
            ],
        },
        {
            'key': 'executive-dashboard',
            'icon': '📊',
            'title': 'ระบบผู้บริหาร',
            'subtitle': 'Dashboard และรายงานภาพรวมโรงเรียน',
            'color': 'purple',
            'features': [
                _erp_feature('Dashboard ผู้บริหาร', 'นักเรียน ครู ห้องเรียน รายวิชา และกิจกรรมวันนี้', endpoint='admin_dashboard', tag='เดิม'),
                _erp_feature('สถิติการมาเรียน', 'ภาพรวมมา สาย ลา ขาด แยกห้อง/รายวัน', endpoint='attendance_dashboard', tag='เดิม'),
                _erp_feature('สถิติผลสัมฤทธิ์', 'คะแนนเฉลี่ย เกรด และรายงานรายวิชา', endpoint='records_center', tag='เดิม'),
                _erp_feature('รายงาน PDF/Excel', 'ส่งออกเอกสารสำหรับประชุมและรายงานผู้บริหาร', status='ต่อยอด'),
                _erp_feature('ปฏิทินบริหารงานโรงเรียน', 'ปฏิทินกิจกรรม วันหยุด และกำหนดการสำคัญ', endpoint='calendar_events', tag='เดิม'),
            ],
        },
        {
            'key': 'online-exam',
            'icon': '📝',
            'title': 'ระบบสอบออนไลน์',
            'subtitle': 'คลังข้อสอบ สุ่มข้อสอบ ตรวจคะแนน และวิเคราะห์ข้อสอบ',
            'color': 'green',
            'features': [
                _erp_feature('คลังข้อสอบ', 'จัดเก็บข้อสอบตามรายวิชา หน่วย และตัวชี้วัด', endpoint='subjects', tag='เดิม'),
                _erp_feature('แบบทดสอบออนไลน์', 'สร้างข้อสอบปรนัย/อัตนัยและให้นักเรียนทำผ่านเว็บ', endpoint='subjects', tag='เดิม'),
                _erp_feature('สุ่มข้อสอบ', 'ตั้งค่าจำนวนข้อและชุดข้อสอบ', status='ต่อยอด'),
                _erp_feature('ตรวจอัตโนมัติ', 'ตรวจคำตอบปรนัยและรวมคะแนน', endpoint='assignments', tag='เดิม'),
                _erp_feature('วิเคราะห์ข้อสอบ', 'ความยาก อำนาจจำแนก และข้อที่นักเรียนผิดมาก', status='ต้นแบบ'),
            ],
        },
        {
            'key': 'activity-finance',
            'icon': '💳',
            'title': 'ระบบการเงินกิจกรรม',
            'subtitle': 'ค่ากิจกรรม ทัศนศึกษา ระดมทรัพยากร และใบเสร็จ',
            'color': 'blue',
            'features': [
                _erp_feature('รายการเก็บเงินกิจกรรม', 'สร้างรายการเก็บเงินตามห้อง/ระดับชั้น', status='ต้นแบบ'),
                _erp_feature('บันทึกการชำระเงิน', 'รับเงินสด/โอน/แนบสลิป', status='ต้นแบบ'),
                _erp_feature('ติดตามค้างชำระ', 'แสดงรายชื่อคนยังไม่จ่ายและส่งออก Excel', status='ต้นแบบ'),
                _erp_feature('ใบเสร็จรับเงิน', 'พิมพ์ใบเสร็จแบบ Manual Payment', status='ต้นแบบ'),
                _erp_feature('รายงานการเงินกิจกรรม', 'สรุปรายวัน รายห้อง และทั้งกิจกรรม', status='ต้นแบบ'),
            ],
        },
    ]
    total_features = sum(len(m['features']) for m in modules)
    linked_features = sum(1 for m in modules for f in m['features'] if f.get('endpoint') or f.get('href'))
    prototype_features = total_features - linked_features
    return modules, {
        'total_modules': len(modules),
        'total_features': total_features,
        'linked_features': linked_features,
        'prototype_features': prototype_features,
    }


def _resolve_erp_links(modules):
    for module in modules:
        module['href'] = url_for('school_erp_module', module_key=module['key'])
        for feature in module['features']:
            if feature.get('endpoint'):
                try:
                    feature['url'] = url_for(feature['endpoint'])
                except Exception:
                    feature['url'] = module['href']
            else:
                feature['url'] = module['href']
    return modules


@app.route('/school-erp')
@login_required
@role_required('admin', 'teacher')
def school_erp():
    modules, summary = build_school_erp_modules()
    modules = _resolve_erp_links(modules)
    return render_template('school_erp.html', modules=modules, summary=summary)


@app.route('/school-erp/<module_key>')
@login_required
@role_required('admin', 'teacher')
def school_erp_module(module_key):
    modules, summary = build_school_erp_modules()
    modules = _resolve_erp_links(modules)
    module = next((m for m in modules if m['key'] == module_key), None)
    if not module:
        flash('ไม่พบโมดูลที่เลือก', 'danger')
        return redirect(url_for('school_erp'))
    return render_template('school_erp_module.html', module=module, modules=modules, summary=summary)


@app.route('/admin')
@login_required
@role_required('admin')
def admin_dashboard():
    active_semester = get_active_semester()
    cal_months, upcoming, today_position = build_calendar_dashboard(semester=active_semester)
    return render_template('admin_dashboard.html', users=active_users_query().count(), teachers=active_teachers_query().count(), students=active_students_query().count(), cal_months=cal_months, upcoming=upcoming, today_position=today_position, active_semester=active_semester, period_info=current_period_info(local_today()))

@app.route('/teacher')
@login_required
@role_required('teacher','admin')
def teacher_dashboard():
    cur, nxt = get_current_schedule()
    subjects = teacher_filter(Subject).all()
    classrooms = teacher_filter(Classroom).all()
    assignments = Assignment.query.filter_by(teacher_id=current_user.id).order_by(Assignment.created_at.desc()).limit(5).all() if current_user.role=='teacher' else Assignment.query.order_by(Assignment.created_at.desc()).limit(5).all()
    active_semester = get_active_semester()
    cal_months, upcoming, today_position = build_calendar_dashboard(teacher_id=current_user.id if current_user.role=='teacher' else None, semester=active_semester)
    today = local_today()
    cur_lesson_log = get_period_lesson_log(cur, today) if cur else None
    nxt_lesson_log = get_period_lesson_log(nxt, today) if nxt else None
    return render_template('teacher_dashboard.html', cur=cur, nxt=nxt, subjects=subjects, classrooms=classrooms, assignments=assignments, cal_months=cal_months, upcoming=upcoming, today_position=today_position, active_semester=active_semester, period_info=current_period_info(today, current_user.id), cur_lesson_log=cur_lesson_log, nxt_lesson_log=nxt_lesson_log, today_date=today.isoformat())

@app.route('/student')
@login_required
@role_required('student')
def student_dashboard():
    statuses = AssignmentStatus.query.filter_by(student_id=current_user.id).join(Assignment).filter(Assignment.assignment_type=='special').order_by(Assignment.created_at.desc()).all()
    room_ids = [x.classroom_id for x in ClassroomStudent.query.filter_by(student_id=current_user.id).all()]
    schedules = TeachingSchedule.query.filter(TeachingSchedule.classroom_id.in_(room_ids or [-1])).order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all()
    today = local_today()
    today_idx = today.weekday()
    today_schedules = [x for x in schedules if x.weekday == today_idx and not is_school_blocked_day(today, x.teacher_id)]
    for row in schedules:
        row.auto_lesson = lesson_for_schedule(row, today, for_student=True)
    for row in today_schedules:
        row.auto_lesson = lesson_for_schedule(row, today, for_student=True)
    learning_plan = build_student_learning_plan(current_user, limit=120)
    return render_template('student_dashboard.html', statuses=statuses, schedules=schedules, today_schedules=today_schedules, learning_plan=learning_plan, day_names=['จันทร์','อังคาร','พุธ','พฤหัสบดี','ศุกร์','เสาร์','อาทิตย์'])

@app.route('/users', methods=['GET','POST'])
@login_required
@role_required('admin')
def users():
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        full_name = (request.form.get('full_name') or '').strip()
        role = (request.form.get('role') or 'teacher').strip()
        password = request.form.get('password') or '1234'

        if not username or not full_name:
            flash('กรุณากรอกชื่อ-สกุล และ username', 'danger')
            return redirect(url_for('users'))

        exists = User.query.filter_by(username=username).first()
        if exists:
            flash(f'username "{username}" มีอยู่แล้วในระบบ กรุณาใช้ username อื่น หรือแก้ไขผู้ใช้เดิม', 'danger')
            return redirect(url_for('users'))

        u = User(username=username, full_name=full_name, role=role)
        if role == 'student':
            fill_student_personal_fields(u, request.form)
        u.set_password(password)
        db.session.add(u)
        db.session.commit()
        flash('เพิ่มผู้ใช้แล้ว', 'success')
        return redirect(url_for('users'))
    return render_template('users.html', users=User.query.order_by(User.is_active.desc(), User.role, User.full_name).all())

@app.route('/users/<int:user_id>/edit', methods=['GET','POST'])
@login_required
@role_required('admin')
def user_edit(user_id):
    u = User.query.get_or_404(user_id)
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        full_name = (request.form.get('full_name') or '').strip()
        role = (request.form.get('role') or u.role).strip()

        if not username or not full_name:
            flash('กรุณากรอกชื่อ-สกุล และ username', 'danger')
            return render_template('user_form.html', u=u)

        exists = User.query.filter(User.username == username, User.id != u.id).first()
        if exists:
            flash(f'username "{username}" มีอยู่แล้วกับผู้ใช้ {exists.full_name} กรุณาใช้ username อื่น', 'danger')
            return render_template('user_form.html', u=u)

        u.full_name = full_name
        u.username = username
        u.role = role
        if role == 'student':
            fill_student_personal_fields(u, request.form)
        if request.form.get('password'):
            u.set_password(request.form.get('password',''))
        db.session.commit()
        flash('แก้ไขผู้ใช้แล้ว', 'success')
        return redirect(url_for('users'))
    return render_template('user_form.html', u=u)

def user_has_related_data(user_id):
    checks = [
        ('ห้องประจำชั้น', Classroom.query.filter_by(teacher_id=user_id).first()),
        ('การผูกครูกับห้อง', TeacherClassroom.query.filter_by(teacher_id=user_id).first()),
        ('นักเรียนในห้อง', ClassroomStudent.query.filter_by(student_id=user_id).first()),
        ('รายวิชาที่รับผิดชอบ', Subject.query.filter_by(teacher_id=user_id).first()),
        ('การผูกครูกับรายวิชา', TeacherSubject.query.filter_by(teacher_id=user_id).first()),
        ('งานที่สั่ง', Assignment.query.filter_by(teacher_id=user_id).first()),
        ('สถานะงานนักเรียน', AssignmentStatus.query.filter_by(student_id=user_id).first()),
        ('คำตอบใบงาน', WorksheetAnswer.query.filter_by(student_id=user_id).first()),
        ('คำตอบแบบทดสอบ', QuizAnswer.query.filter_by(student_id=user_id).first()),
        ('ตารางสอน', TeachingSchedule.query.filter_by(teacher_id=user_id).first()),
        ('ปฏิทินครู', CalendarEvent.query.filter_by(teacher_id=user_id).first()),
        ('เช็กชื่อกิจกรรม', ActivityAttendance.query.filter_by(student_id=user_id).first()),
        ('เช็กชื่อรายวิชา', Attendance.query.filter_by(student_id=user_id).first()),
        ('ประวัติผู้เช็กชื่อ', Attendance.query.filter_by(checked_by_id=user_id).first()),
        ('ประวัติสอนแทน', Attendance.query.filter_by(substitute_for_teacher_id=user_id).first()),
        ('คะแนนรายวิชา', ManualScore.query.filter_by(student_id=user_id).first()),
    ]
    return [name for name, found in checks if found is not None]

def detach_teacher_before_delete(user_id):
    """ปลด FK ของครูออกก่อนลบ User โดยไม่ล้างประวัติการเช็กชื่อ/คะแนน

    แนวคิด: ครูเป็นบัญชีล็อกอิน จึงลบได้ แต่ข้อมูลประวัติยังอยู่กับรายวิชา/ห้อง/นักเรียน
    - ตารางผูกครู-วิชา/ครู-ห้อง: ลบเฉพาะแถวผูก
    - ห้อง/รายวิชาประจำ: ตั้ง teacher_id = NULL
    - เช็กชื่อที่อ้างคนเช็ก/สอนแทน: ตั้งเป็น NULL
    - ตารางสอนของครูที่ถูกลบ: ปลด attendance.schedule_id ก่อน แล้วลบตารางสอน
    - งานที่เคยสั่ง: โอนไปให้ admin คนแรก เพื่อไม่ให้ข้อมูลการส่งงานหาย
    """
    fallback_admin = User.query.filter(User.role == 'admin', User.id != user_id).order_by(User.id).first()
    fallback_id = fallback_admin.id if fallback_admin else current_user.id

    TeacherSubject.query.filter_by(teacher_id=user_id).delete(synchronize_session=False)
    TeacherClassroom.query.filter_by(teacher_id=user_id).delete(synchronize_session=False)

    Subject.query.filter_by(teacher_id=user_id).update({Subject.teacher_id: None}, synchronize_session=False)
    Classroom.query.filter_by(teacher_id=user_id).update({Classroom.teacher_id: None}, synchronize_session=False)
    CalendarEvent.query.filter_by(teacher_id=user_id).update({CalendarEvent.teacher_id: None}, synchronize_session=False)

    Attendance.query.filter_by(checked_by_id=user_id).update({Attendance.checked_by_id: None}, synchronize_session=False)
    Attendance.query.filter_by(substitute_for_teacher_id=user_id).update({Attendance.substitute_for_teacher_id: None}, synchronize_session=False)

    # Assignment.teacher_id ยังเป็น NOT NULL ในฐานข้อมูลเดิม จึงโอนไป admin แทนการตั้ง NULL
    Assignment.query.filter_by(teacher_id=user_id).update({Assignment.teacher_id: fallback_id}, synchronize_session=False)

    schedule_ids = [x.id for x in TeachingSchedule.query.with_entities(TeachingSchedule.id).filter_by(teacher_id=user_id).all()]
    if schedule_ids:
        Attendance.query.filter(Attendance.schedule_id.in_(schedule_ids)).update({Attendance.schedule_id: None}, synchronize_session=False)
        TeachingSchedule.query.filter(TeachingSchedule.id.in_(schedule_ids)).delete(synchronize_session=False)


@app.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def user_delete(user_id):
    if user_id == current_user.id:
        flash('ไม่สามารถลบบัญชีที่กำลังใช้งานได้', 'danger')
        return redirect(url_for('users'))
    u = User.query.get_or_404(user_id)

    # ถ้าเป็นครู ให้ปลดการผูกออกก่อน แล้วลบบัญชีได้จริง
    # ประวัติการเช็กชื่อยังอยู่ เพราะอยู่กับ subject/classroom/student/date ไม่ได้ลบ attendance
    if u.role == 'teacher':
        old_username = u.username
        detach_teacher_before_delete(user_id)
        db.session.delete(u)
        db.session.commit()
        flash(f'ลบไอดีครู {old_username} แล้ว และปลดการผูกกับรายวิชา/ห้อง/ตารางสอนเรียบร้อย โดยไม่ลบประวัติเช็กชื่อเดิม', 'success')
        return redirect(url_for('users'))

    related = user_has_related_data(user_id)
    if related:
        # นักเรียน/แอดมินที่มีประวัติส่งงาน คะแนน หรือเช็กชื่อ ให้ปิดใช้งานแทน เพื่อกันข้อมูลนักเรียนหาย
        old_username = u.username
        u.is_active = False
        if not u.username.startswith('deleted_'):
            u.username = f'deleted_{u.id}_{u.username}'[:100]
        u.set_password(uuid.uuid4().hex)
        db.session.commit()
        flash(f'บัญชี {old_username} มีข้อมูลผูกอยู่ ({", ".join(related[:3])}{"..." if len(related) > 3 else ""}) จึงปิดใช้งานแทนการลบถาวร และสามารถสร้าง username เดิมใหม่ได้แล้ว', 'warning')
    else:
        db.session.delete(u)
        db.session.commit()
        flash('ลบผู้ใช้แล้ว', 'success')
    return redirect(url_for('users'))

@app.route('/users/<int:user_id>/restore', methods=['POST'])
@login_required
@role_required('admin')
def user_restore(user_id):
    u = User.query.get_or_404(user_id)
    u.is_active = True
    if u.username.startswith(f'deleted_{u.id}_'):
        original = u.username.replace(f'deleted_{u.id}_', '', 1)
        if not User.query.filter(User.username == original, User.id != u.id).first():
            u.username = original
    db.session.commit()
    flash('เปิดใช้งานบัญชีแล้ว', 'success')
    return redirect(url_for('users'))


@app.route('/classrooms', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def classrooms():
    if request.method == 'POST':
        teacher_id_raw = request.form.get('teacher_id')
        teacher_id = current_user.id if current_user.role == 'teacher' else (int(teacher_id_raw) if teacher_id_raw else None)
        room = Classroom(name=request.form.get('name',''), teacher_id=teacher_id)
        db.session.add(room); db.session.flush()

        activity_teacher_ids = request.form.getlist('activity_teacher_ids')
        if current_user.role == 'teacher':
            activity_teacher_ids = [current_user.id]
        sync_activity_teachers_for_classroom(room.id, activity_teacher_ids, keep_teacher_id=teacher_id)

        db.session.commit()
        flash('สร้างห้องเรียนแล้ว', 'success')
        return redirect(url_for('classrooms'))
    rooms = teacher_filter(Classroom).order_by(Classroom.name).all()
    teacher_room_map = build_classroom_teacher_map([r.id for r in rooms])
    return render_template('classrooms.html', rooms=rooms, teachers=active_teachers_query().order_by(User.full_name).all(), teacher_room_links=TeacherClassroom.query.all(), teacher_room_map=teacher_room_map)

@app.route('/classrooms/<int:classroom_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def classroom_edit(classroom_id):
    room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room): return deny_redirect('classrooms')
    if request.method == 'POST':
        room.name = request.form.get('name','')
        if current_user.role == 'admin':
            teacher_id_raw = request.form.get('teacher_id')
            room.teacher_id = int(teacher_id_raw) if teacher_id_raw else None
            sync_activity_teachers_for_classroom(room.id, request.form.getlist('activity_teacher_ids'), keep_teacher_id=room.teacher_id)
        db.session.commit(); flash('แก้ไขห้องเรียนแล้ว', 'success')
        return redirect(url_for('classrooms'))
    assigned_activity_teacher_ids = {x.teacher_id for x in TeacherClassroom.query.filter_by(classroom_id=room.id).all()}
    return render_template('classroom_form.html', room=room, teachers=active_teachers_query().order_by(User.full_name).all(), assigned_activity_teacher_ids=assigned_activity_teacher_ids)

def classroom_has_history(classroom_id):
    return any([
        ClassroomStudent.query.filter_by(classroom_id=classroom_id).first(),
        SubjectClassroom.query.filter_by(classroom_id=classroom_id).first(),
        Assignment.query.filter_by(classroom_id=classroom_id).first(),
        TeachingSchedule.query.filter_by(classroom_id=classroom_id).first(),
        Attendance.query.filter_by(classroom_id=classroom_id).first(),
    ])



def _delete_assignments(query):
    """ลบงานและข้อมูลส่งงาน/คำตอบที่ผูกกับงานนั้นทั้งหมด"""
    assignment_ids = [a.id for a in query.all()]
    if not assignment_ids:
        return 0
    WorksheetAnswer.query.filter(WorksheetAnswer.assignment_id.in_(assignment_ids)).delete(synchronize_session=False)
    QuizAnswer.query.filter(QuizAnswer.assignment_id.in_(assignment_ids)).delete(synchronize_session=False)
    AssignmentStatus.query.filter(AssignmentStatus.assignment_id.in_(assignment_ids)).delete(synchronize_session=False)
    deleted = Assignment.query.filter(Assignment.id.in_(assignment_ids)).delete(synchronize_session=False)
    return deleted


def force_delete_worksheet_data(worksheet_id, delete_file=True):
    """ลบใบงานพร้อมไฟล์ ข้อคำถาม และคำตอบนักเรียนที่อ้างถึงข้อในใบงานนี้"""
    ws = Worksheet.query.get(worksheet_id)
    if ws and delete_file:
        safe_remove_upload_file(ws.file_path)
    question_ids = [q.id for q in WorksheetQuestion.query.filter_by(worksheet_id=worksheet_id).all()]
    if question_ids:
        WorksheetAnswer.query.filter(WorksheetAnswer.question_id.in_(question_ids)).delete(synchronize_session=False)
        WorksheetQuestion.query.filter(WorksheetQuestion.id.in_(question_ids)).delete(synchronize_session=False)


def force_delete_quiz_data(quiz_id):
    """ลบแบบทดสอบพร้อมข้อสอบและคำตอบนักเรียนที่อ้างถึงข้อสอบนั้น"""
    question_ids = [q.id for q in QuizQuestion.query.filter_by(quiz_id=quiz_id).all()]
    if question_ids:
        QuizAnswer.query.filter(QuizAnswer.question_id.in_(question_ids)).delete(synchronize_session=False)
        QuizQuestion.query.filter(QuizQuestion.id.in_(question_ids)).delete(synchronize_session=False)


def force_delete_lesson_data(lesson_id, delete_files=True):
    """ลบบทเรียนแบบรวมศูนย์ เพื่อไม่ให้มีข้อมูลลูกค้าง: ไฟล์ ใบงาน แบบทดสอบ งาน คาบเรียน และคะแนนคาบ"""
    _delete_assignments(Assignment.query.filter_by(lesson_id=lesson_id))
    PeriodLessonLog.query.filter_by(lesson_id=lesson_id).delete(synchronize_session=False)
    TeachingSchedule.query.filter_by(lesson_id=lesson_id).update({TeachingSchedule.lesson_id: None}, synchronize_session=False)
    try:
        ClassworkScoreItem.query.filter_by(lesson_id=lesson_id).update({ClassworkScoreItem.lesson_id: None}, synchronize_session=False)
    except Exception:
        pass

    for lf in LessonFile.query.filter_by(lesson_id=lesson_id).all():
        if delete_files:
            safe_remove_upload_file(lf.file_path)
        db.session.delete(lf)

    for ws in Worksheet.query.filter_by(lesson_id=lesson_id).all():
        force_delete_worksheet_data(ws.id, delete_file=delete_files)
        db.session.delete(ws)

    for quiz in Quiz.query.filter_by(lesson_id=lesson_id).all():
        force_delete_quiz_data(quiz.id)
        db.session.delete(quiz)

def force_delete_classroom_data(classroom_id):
    """ลบห้องเรียนแบบตัดข้อมูลที่ผูกอยู่ทั้งหมด เพื่อให้ลบได้จริงตามคำขอ"""
    _delete_assignments(Assignment.query.filter_by(classroom_id=classroom_id))

    item_ids = [x.id for x in ClassworkScoreItem.query.filter_by(classroom_id=classroom_id).all()]
    if item_ids:
        ClassworkScore.query.filter(ClassworkScore.item_id.in_(item_ids)).delete(synchronize_session=False)
        ClassworkScoreItem.query.filter(ClassworkScoreItem.id.in_(item_ids)).delete(synchronize_session=False)

    Attendance.query.filter_by(classroom_id=classroom_id).delete(synchronize_session=False)
    TeachingSchedule.query.filter_by(classroom_id=classroom_id).delete(synchronize_session=False)
    SubjectClassroom.query.filter_by(classroom_id=classroom_id).delete(synchronize_session=False)
    ClassroomStudent.query.filter_by(classroom_id=classroom_id).delete(synchronize_session=False)
    TeacherClassroom.query.filter_by(classroom_id=classroom_id).delete(synchronize_session=False)

def force_delete_subject_data(subject_id):
    """ลบรายวิชาแบบตัดข้อมูลที่ผูกอยู่ทั้งหมด เพื่อให้ลบได้จริงตามคำขอ"""
    _delete_assignments(Assignment.query.filter_by(subject_id=subject_id))

    # คะแนนรายคาบต้องลบคะแนนลูกก่อนหัวข้อคะแนน
    item_ids = [x.id for x in ClassworkScoreItem.query.filter_by(subject_id=subject_id).all()]
    if item_ids:
        ClassworkScore.query.filter(ClassworkScore.item_id.in_(item_ids)).delete(synchronize_session=False)
        ClassworkScoreItem.query.filter(ClassworkScoreItem.id.in_(item_ids)).delete(synchronize_session=False)

    Attendance.query.filter_by(subject_id=subject_id).delete(synchronize_session=False)
    ManualScore.query.filter_by(subject_id=subject_id).delete(synchronize_session=False)
    GradeSetting.query.filter_by(subject_id=subject_id).delete(synchronize_session=False)
    TeachingSchedule.query.filter_by(subject_id=subject_id).delete(synchronize_session=False)
    SubjectClassroom.query.filter_by(subject_id=subject_id).delete(synchronize_session=False)
    TeacherSubject.query.filter_by(subject_id=subject_id).delete(synchronize_session=False)

    units = Unit.query.filter_by(subject_id=subject_id).all()
    for unit in units:
        force_delete_unit_data(unit.id)
        db.session.delete(unit)


def force_delete_unit_data(unit_id):
    """ลบหน่วยการเรียนรู้พร้อมบทเรียน ไฟล์ ใบงาน แบบทดสอบ งาน และคำตอบทั้งหมดที่ผูกอยู่"""
    lessons = Lesson.query.filter_by(unit_id=unit_id).all()
    for lesson in lessons:
        force_delete_lesson_data(lesson.id, delete_files=True)
        db.session.delete(lesson)

@app.route('/classrooms/<int:classroom_id>/deactivate', methods=['POST'])
@login_required
@role_required('teacher','admin')
def classroom_deactivate(classroom_id):
    room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room): return deny_redirect('classrooms')
    room.is_active = False
    db.session.commit(); flash('ปิดใช้งานห้องเรียนแล้ว ข้อมูลเดิมยังอยู่ ไม่หาย', 'success')
    return redirect(url_for('classrooms'))

@app.route('/classrooms/<int:classroom_id>/restore', methods=['POST'])
@login_required
@role_required('admin')
def classroom_restore(classroom_id):
    room = Classroom.query.get_or_404(classroom_id)
    room.is_active = True
    db.session.commit(); flash('เปิดใช้งานห้องเรียนแล้ว', 'success')
    return redirect(url_for('classrooms'))

@app.route('/classrooms/<int:classroom_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def classroom_delete(classroom_id):
    room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room): return deny_redirect('classrooms')
    room_name = room.name
    force_delete_classroom_data(room.id)
    db.session.delete(room)
    db.session.commit()
    flash(f'ลบห้องเรียน {room_name} และข้อมูลที่ผูกอยู่ทั้งหมดแล้ว', 'success')
    return redirect(url_for('classrooms'))

@app.route('/classroom_students/<int:link_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def classroom_student_delete(link_id):
    link = ClassroomStudent.query.get_or_404(link_id)
    if not owns_classroom(link.classroom): return deny_redirect('classrooms')
    room_id = link.classroom_id
    db.session.delete(link); db.session.commit(); flash('นำนักเรียนออกจากห้องแล้ว', 'success')
    return redirect(url_for('classroom_students', classroom_id=room_id))


@app.route('/teacher_assignments', methods=['GET','POST'])
@login_required
@role_required('admin')
def teacher_assignments():
    if request.method == 'POST':
        teacher_id = int(request.form.get('teacher_id',''))
        new_subject_ids = {int(x) for x in request.form.getlist('subject_ids') if x}
        new_classroom_ids = {int(x) for x in request.form.getlist('classroom_ids') if x}

        # อัปเดตแบบแทนที่จริง: ติ๊กอะไรไว้ = ครูเห็น/ใช้สิ่งนั้น, เอาติ๊กออก = ถอดสิทธิ์
        old_subject_links = TeacherSubject.query.filter_by(teacher_id=teacher_id).all()
        old_classroom_links = TeacherClassroom.query.filter_by(teacher_id=teacher_id).all()
        for link in old_subject_links:
            if link.subject_id not in new_subject_ids:
                db.session.delete(link)
        for link in old_classroom_links:
            if link.classroom_id not in new_classroom_ids:
                db.session.delete(link)
        for sid in new_subject_ids:
            if not TeacherSubject.query.filter_by(teacher_id=teacher_id, subject_id=sid).first():
                db.session.add(TeacherSubject(teacher_id=teacher_id, subject_id=sid))
        for rid in new_classroom_ids:
            if not TeacherClassroom.query.filter_by(teacher_id=teacher_id, classroom_id=rid).first():
                db.session.add(TeacherClassroom(teacher_id=teacher_id, classroom_id=rid))

        # ตั้งครูประจำชั้นให้รายการที่ยังไม่มี เพื่อให้หน้าอื่นแสดงผลชัดเจน
        for sid in new_subject_ids:
            sub = Subject.query.get(sid)
            if sub and not sub.teacher_id:
                sub.teacher_id = teacher_id
        for rid in new_classroom_ids:
            room = Classroom.query.get(rid)
            if room and not room.teacher_id:
                room.teacher_id = teacher_id
        db.session.commit(); flash('บันทึกการมอบหมายแล้ว ครูจะเห็นเฉพาะวิชา/ห้องที่ได้รับมอบหมาย', 'success')
        return redirect(url_for('teacher_assignments', teacher_id=teacher_id))
    selected_teacher_id = request.args.get('teacher_id', type=int)
    teachers = active_teachers_query().order_by(User.full_name).all()
    subjects = Subject.query.filter_by(is_active=True).order_by(Subject.name).all()
    classrooms = Classroom.query.filter_by(is_active=True).order_by(Classroom.name).all()
    assigned_subject_ids = set()
    assigned_classroom_ids = set()
    assigned_subject_links = []
    assigned_classroom_links = []
    if selected_teacher_id:
        assigned_subject_links = TeacherSubject.query.filter_by(teacher_id=selected_teacher_id).all()
        assigned_classroom_links = TeacherClassroom.query.filter_by(teacher_id=selected_teacher_id).all()
        assigned_subject_ids = {x.subject_id for x in assigned_subject_links}
        assigned_classroom_ids = {x.classroom_id for x in assigned_classroom_links}
    return render_template('teacher_assignments.html', teachers=teachers, subjects=subjects, classrooms=classrooms, selected_teacher_id=selected_teacher_id, assigned_subject_ids=assigned_subject_ids, assigned_classroom_ids=assigned_classroom_ids, assigned_subject_links=assigned_subject_links, assigned_classroom_links=assigned_classroom_links)

@app.route('/teacher_assignments/subject/<int:link_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def teacher_subject_delete(link_id):
    link = TeacherSubject.query.get_or_404(link_id)
    teacher_id = link.teacher_id
    sub = link.subject
    db.session.delete(link)
    if sub and sub.teacher_id == teacher_id and not TeacherSubject.query.filter(TeacherSubject.teacher_id!=teacher_id, TeacherSubject.subject_id==sub.id).first():
        sub.teacher_id = None
    db.session.commit(); flash('ยกเลิกมอบหมายรายวิชาแล้ว ข้อมูลรายวิชายังอยู่', 'success')
    return redirect(url_for('teacher_assignments', teacher_id=teacher_id))

@app.route('/teacher_assignments/classroom/<int:link_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def teacher_classroom_delete(link_id):
    link = TeacherClassroom.query.get_or_404(link_id)
    teacher_id = link.teacher_id
    room = link.classroom
    db.session.delete(link)
    if room and room.teacher_id == teacher_id and not TeacherClassroom.query.filter(TeacherClassroom.teacher_id!=teacher_id, TeacherClassroom.classroom_id==room.id).first():
        room.teacher_id = None
    db.session.commit(); flash('ยกเลิกมอบหมายห้องเรียนแล้ว ข้อมูลห้องยังอยู่', 'success')
    return redirect(url_for('teacher_assignments', teacher_id=teacher_id))


@app.route('/classroom/<int:classroom_id>/students', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def classroom_students(classroom_id):
    room = Classroom.query.get_or_404(classroom_id)
    if current_user.role != 'admin' and room.teacher_id != current_user.id:
        flash('ไม่มีสิทธิ์', 'danger'); return redirect(url_for('classrooms'))
    if request.method == 'POST':
        action = request.form.get('action') or 'link_existing'
        try:
            if action == 'quick_add':
                create_or_update_student_from_form(request.form, default_room_id=room.id)
                db.session.commit()
                flash('เพิ่ม/อัปเดตนักเรียนและนำเข้าห้องแล้ว', 'success')
            else:
                student_id = int(request.form.get('student_id',''))
                if not ClassroomStudent.query.filter_by(classroom_id=room.id, student_id=student_id).first():
                    db.session.add(ClassroomStudent(classroom_id=room.id, student_id=student_id))
                    db.session.commit()
                    flash('เพิ่มนักเรียนเข้าห้องแล้ว', 'success')
        except Exception as e:
            db.session.rollback(); flash(str(e), 'danger')
        return redirect(url_for('classroom_students', classroom_id=room.id))
    links = ClassroomStudent.query.filter_by(classroom_id=room.id).join(User, ClassroomStudent.student_id == User.id).order_by(User.username.asc(), User.full_name.asc()).all()
    students = active_students_query().order_by(User.username.asc(), User.full_name.asc()).all()
    classrooms_all = Classroom.query.filter(db.or_(Classroom.is_active == True, Classroom.is_active.is_(None))).order_by(Classroom.name).all()
    return render_template('classroom_students.html', room=room, links=links, students=students, classrooms_all=classrooms_all)


@app.route('/student/<int:student_id>/move-room', methods=['POST'])
@login_required
@role_required('teacher','admin')
def student_move_room(student_id):
    student = User.query.get_or_404(student_id)
    new_room_id = request.form.get('new_classroom_id', type=int)
    old_room_id = request.form.get('old_classroom_id', type=int)
    redirect_to = request.form.get('redirect_to') or request.referrer or url_for('classrooms')
    if student.role != 'student' or not new_room_id:
        flash('ข้อมูลนักเรียนหรือห้องเรียนไม่ถูกต้อง', 'danger')
        return redirect(redirect_to)
    new_room = Classroom.query.get_or_404(new_room_id)
    if current_user.role != 'admin' and not owns_classroom(new_room):
        return deny_redirect('classrooms')
    # ย้ายห้องปัจจุบัน: เอานักเรียนออกจากห้องเดิมทั้งหมด แล้วผูกเข้าห้องใหม่
    ClassroomStudent.query.filter_by(student_id=student.id).delete(synchronize_session=False)
    db.session.add(ClassroomStudent(classroom_id=new_room.id, student_id=student.id))
    db.session.commit()
    flash(f'ย้าย {student.full_name} ไปห้อง {new_room.name} แล้ว ประวัติเช็กชื่อ/คะแนนเดิมยังอยู่', 'success')
    return redirect(redirect_to)

@app.route('/attendance/<int:subject_id>/<int:classroom_id>/quick-add-student', methods=['POST'])
@login_required
@role_required('teacher','admin')
def attendance_quick_add_student(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room):
        return deny_redirect('records_center')
    try:
        create_or_update_student_from_form(request.form, default_room_id=room.id)
        db.session.commit()
        flash('เพิ่มนักเรียนเข้าห้องนี้แล้ว', 'success')
    except Exception as e:
        db.session.rollback(); flash(str(e), 'danger')
    return redirect(request.form.get('redirect_to') or url_for('attendance', subject_id=subject.id, classroom_id=room.id, date=request.form.get('date') or local_today().isoformat(), period_no=request.form.get('period_no') or None, schedule_id=request.form.get('schedule_id') or None))


@app.route('/classroom/<int:classroom_id>/activities', methods=['GET', 'POST'])
@login_required
@role_required('teacher', 'admin')
def classroom_activities(classroom_id):
    room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room):
        return deny_redirect('classrooms')
    ensure_default_classroom_activities(room.id)

    if request.method == 'POST':
        title = (request.form.get('title') or '').strip()
        event_date_text = (request.form.get('event_date') or '').strip()
        note = (request.form.get('note') or '').strip()
        target_scope = (request.form.get('target_scope') or '').strip() or 'classroom'
        if target_scope not in ('classroom', 'all', 'scout_m1_m3'):
            target_scope = 'classroom'
        # ถ้าชื่อกิจกรรมมีคำว่า รวม ให้ถือว่าเป็นรวมทั้งโรงเรียนทันที
        if 'รวม' in title or 'ทั้งหมด' in title:
            target_scope = 'all'
        if 'ลูกเสือ' in title or 'เนตรนารี' in title:
            target_scope = 'scout_m1_m3'
        if not title:
            flash('กรุณากรอกชื่อกิจกรรม', 'danger')
            return redirect(url_for('classroom_activities', classroom_id=room.id))
        event_date = datetime.strptime(event_date_text, '%Y-%m-%d').date() if event_date_text else None
        db.session.add(ClassroomActivity(
            classroom_id=room.id,
            title=title,
            activity_type='special',
            target_scope=target_scope,
            event_date=event_date,
            note=note,
            is_active=True,
        ))
        db.session.commit()
        flash('เพิ่มกิจกรรมพิเศษแล้ว', 'success')
        return redirect(url_for('classroom_activities', classroom_id=room.id))

    activities = ClassroomActivity.query.filter_by(classroom_id=room.id, is_active=True).all()
    activities = sorted(activities, key=lambda a: (0 if a.activity_type == 'assembly' else 1, a.event_date or date.min, a.title or ''))
    today = date.today()
    return render_template(
        'classroom_activities.html',
        room=room,
        activities=activities,
        activity_teachers=classroom_activity_teachers(room.id),
        today=today,
        today_iso=today.isoformat(),
        today_label=thai_date_label(today),
        is_today_blocked=is_school_blocked_day(today, room.teacher_id),
    )


@app.route('/classroom/<int:classroom_id>/activities/<int:activity_id>/delete', methods=['POST'])
@login_required
@role_required('teacher', 'admin')
def classroom_activity_delete(classroom_id, activity_id):
    room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room):
        return deny_redirect('classrooms')
    activity = ClassroomActivity.query.filter_by(id=activity_id, classroom_id=room.id).first_or_404()
    if activity.activity_type == 'assembly':
        flash('รายการเข้าแถวเป็นรายการหลักของห้อง ไม่ให้ลบ แต่สามารถไม่ใช้งานได้โดยไม่เช็กชื่อ', 'warning')
        return redirect(url_for('classroom_activities', classroom_id=room.id))
    ActivityAttendance.query.filter_by(activity_id=activity.id).delete()
    db.session.delete(activity)
    db.session.commit()
    flash('ลบกิจกรรมพิเศษแล้ว', 'success')
    return redirect(url_for('classroom_activities', classroom_id=room.id))


@app.route('/classroom/<int:classroom_id>/activities/<int:activity_id>/attendance', methods=['GET', 'POST'])
@login_required
@role_required('teacher', 'admin')
def classroom_activity_attendance(classroom_id, activity_id):
    room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room):
        return deny_redirect('classrooms')
    activity = ClassroomActivity.query.filter_by(id=activity_id, classroom_id=room.id).first_or_404()
    links = activity_student_links(activity, room)
    activity_scope = get_activity_scope(activity)
    show_classroom_column = activity_scope in ('all', 'scout_m1_m3')

    default_date = activity.event_date or local_today()
    selected_date = datetime.strptime(request.args.get('date', default_date.isoformat()), '%Y-%m-%d').date()
    if activity.event_date:
        selected_date = activity.event_date

    if request.method == 'POST':
        selected_date = datetime.strptime(request.form.get('date',''), '%Y-%m-%d').date()
        if activity.event_date:
            selected_date = activity.event_date
        bulk_status = normalize_attendance_status(request.form.get('bulk_status')) if request.form.get('bulk_status') else None
        for link in links:
            st = bulk_status if bulk_status else normalize_attendance_status(request.form.get(f's_{link.student_id}', 'มา'))
            student_room_id = link.classroom_id
            row = ActivityAttendance.query.filter_by(
                activity_id=activity.id,
                student_id=link.student_id,
                date=selected_date,
            ).first()
            if not row:
                row = ActivityAttendance(activity_id=activity.id, classroom_id=student_room_id, student_id=link.student_id, date=selected_date)
                db.session.add(row)
            row.classroom_id = student_room_id
            row.status = st
        db.session.commit()
        flash('บันทึกเช็กชื่อกิจกรรมเรียบร้อยแล้ว', 'success')
        return redirect(url_for('classroom_activity_attendance', classroom_id=room.id, activity_id=activity.id, date=selected_date.isoformat()))

    old = {
        a.student_id: normalize_attendance_status(a.status)
        for a in ActivityAttendance.query.filter_by(activity_id=activity.id, date=selected_date).all()
    }
    date_headers, att_data, summary = build_activity_attendance_report(activity.id, room.id, links)

    return render_template(
        'classroom_activity_attendance.html',
        room=room,
        activity=activity,
        links=links,
        old=old,
        statuses=ATTENDANCE_STATUSES,
        att_date=selected_date,
        selected_day_label=thai_date_label(selected_date),
        is_blocked_day=is_school_blocked_day(selected_date, room.teacher_id),
        date_headers=date_headers,
        att_data=att_data,
        summary=summary,
        status_symbols=ATTENDANCE_SYMBOLS,
        activity_scope=activity_scope,
        activity_scope_label=activity_scope_label(activity),
        show_classroom_column=show_classroom_column,
    )

@app.route('/classroom/<int:classroom_id>/activities/<int:activity_id>/attendance/delete-day', methods=['POST'])
@login_required
@role_required('teacher', 'admin')
def classroom_activity_attendance_delete_day(classroom_id, activity_id):
    room = Classroom.query.get_or_404(classroom_id)
    if not owns_classroom(room):
        return deny_redirect('classrooms')
    activity = ClassroomActivity.query.filter_by(id=activity_id, classroom_id=room.id).first_or_404()
    att_date = datetime.strptime(request.form.get('date',''), '%Y-%m-%d').date()
    ActivityAttendance.query.filter_by(activity_id=activity.id, date=att_date).delete()
    db.session.commit()
    flash('ลบเช็กชื่อกิจกรรมของวันที่เลือกแล้ว', 'success')
    return redirect(url_for('classroom_activity_attendance', classroom_id=room.id, activity_id=activity.id, date=att_date.isoformat()))

@app.route('/import_students', methods=['GET','POST'])
@login_required
@role_required('admin','teacher')
def import_students():
    rooms = teacher_filter(Classroom).all()
    if request.method == 'POST':
        f = request.files.get('excel')
        classroom_id = int(request.form.get('classroom_id',''))
        if not f:
            flash('กรุณาเลือกไฟล์ Excel', 'danger'); return redirect(url_for('import_students'))
        path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(f.filename))
        f.save(path)
        wb = load_workbook(path)
        ws = wb.active
        created = updated = 0
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1: continue
            citizen_id, full_name, birth_date = row[0], row[1], row[2]
            if not citizen_id or not full_name: continue
            citizen_id = str(citizen_id).strip()
            if isinstance(birth_date, datetime):
                password = birth_date.strftime('%d%m%Y')
                birth_text = birth_date.strftime('%d/%m/%Y')
            else:
                birth_text = str(birth_date).strip()
                password = birth_text.replace('/','').replace('-','')
            user = User.query.filter_by(username=citizen_id).first()
            if not user:
                user = User(username=citizen_id, full_name=str(full_name), role='student', citizen_id=citizen_id, birth_date=birth_text, must_change_password=True)
                user.set_password(password)
                db.session.add(user); db.session.flush(); created += 1
            else:
                user.full_name = str(full_name); user.birth_date = birth_text; updated += 1
            if not ClassroomStudent.query.filter_by(classroom_id=classroom_id, student_id=user.id).first():
                db.session.add(ClassroomStudent(classroom_id=classroom_id, student_id=user.id))
        db.session.commit()
        flash(f'นำเข้าเรียบร้อย สร้างใหม่ {created} คน อัปเดต {updated} คน', 'success')
        return redirect(url_for('import_students'))

    return render_template('import_students.html', rooms=rooms)


def parse_excel_date_value(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    txt = str(value).strip().replace('/', '-').replace('.', '-')
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y'):
        try:
            d = datetime.strptime(txt, fmt).date()
            if d.year > 2400:
                d = date(d.year-543, d.month, d.day)
            elif d.year < 100:
                d = date(d.year+2000, d.month, d.day)
            return d
        except Exception:
            pass
    return None

def birth_password(value):
    d = parse_excel_date_value(value)
    if d:
        return f"{d.day:02d}{d.month:02d}{d.year+543}"
    return str(value).replace('/','').replace('-','').strip() if value else '1234'

def sheet_rows(wb, names):
    for name in names:
        if name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return [], {}
            headers = [str(x).strip().lower() if x is not None else '' for x in rows[0]]
            return rows[1:], {h:i for i,h in enumerate(headers) if h}
    return [], {}

def cell(row, headers, *names, default=''):
    for n in names:
        i = headers.get(n.lower())
        if i is not None and i < len(row) and row[i] is not None:
            return row[i]
    return default

def get_or_create_teacher(username, full_name=None):
    username = str(username).strip()
    u = User.query.filter_by(username=username).first()
    if not u:
        u = User(username=username, full_name=full_name or username, role='teacher')
        u.set_password('1234')
        db.session.add(u); db.session.flush()
    else:
        u.role = 'teacher'
        if full_name: u.full_name = full_name
    return u

def get_or_create_classroom(name):
    name = str(name).strip()
    room = Classroom.query.filter_by(name=name).first()
    if not room:
        room = Classroom(name=name)
        db.session.add(room); db.session.flush()
    return room

def normalize_subject_display(text):
    """ทำชื่อรายวิชาให้สะอาดขึ้นก่อนบันทึก/เทียบซ้ำ"""
    import re
    s = str(text or '').strip()
    replacements = {
        'อาชิพ':'อาชีพ', 'เทคโนโลอยี':'เทคโนโลยี', 'ออกเบบ':'ออกแบบ',
        'ออกเแบบ':'ออกแบบ', 'เเฉะ':'และ', 'เและ':'และ', 'เเละ':'และ',
        'การสือสาร':'การสื่อสาร', 'ภาษาจิน':'ภาษาจีน', 'หน้าทิพลเมือง':'หน้าที่พลเมือง',
        'วัทยาศาสตร์':'วิทยาศาสตร์'
    }
    for a,b in replacements.items():
        s = s.replace(a,b)
    s = re.sub(r'[!|\[\]"]+', '', s)
    s = re.sub(r'\s*\.\s*', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip(' /-')
    s = re.sub(r'\s*และ\s*', 'และ', s)
    return s.strip()

def extract_subject_code(text):
    import re
    m = re.search(r'([ก-๙A-Za-z]{1,3}\d{5})', str(text or '').strip())
    return m.group(1) if m else ''

def subject_identity_key(text):
    """คีย์สำหรับรวมวิชา: ถ้ามีรหัสวิชา ใช้รหัสเป็นหลัก ถ้าไม่มีใช้ชื่อที่ normalize แล้ว"""
    import re
    s = normalize_subject_display(text)
    code = extract_subject_code(s)
    if code:
        return code.lower()
    return re.sub(r'\s+', '', s).lower()

def make_subject_name(code='', name='', fallback=''):
    code = str(code or '').strip() or extract_subject_code(fallback)
    name = normalize_subject_display(name or fallback)
    if code and name.startswith(code):
        name = normalize_subject_display(name[len(code):])
    return ' '.join([x for x in [code, name] if x]).strip()

def ensure_grade_setting(subject_id):
    if subject_id and not GradeSetting.query.filter_by(subject_id=subject_id).first():
        db.session.add(GradeSetting(subject_id=subject_id))

def subject_is_linked_to_teacher(subject_id, teacher_id):
    if not teacher_id:
        return False
    return TeacherSubject.query.filter_by(teacher_id=teacher_id, subject_id=subject_id).first() is not None

def ensure_teacher_subject(teacher_id, subject_id):
    if teacher_id and subject_id and not TeacherSubject.query.filter_by(teacher_id=teacher_id, subject_id=subject_id).first():
        db.session.add(TeacherSubject(teacher_id=teacher_id, subject_id=subject_id))

def get_or_create_subject(name, credit=1.0, total_periods=40):
    name = make_subject_name(fallback=name)
    key = subject_identity_key(name)
    sub = None
    for cand in Subject.query.all():
        if subject_identity_key(cand.name) == key and not TeacherSubject.query.filter_by(subject_id=cand.id).first():
            sub = cand
            break
    if not sub:
        sub = Subject(name=name, credit=float(credit or 1), total_periods=int(total_periods or 40))
        db.session.add(sub); db.session.flush()
        ensure_grade_setting(sub.id)
    else:
        # ถ้าชื่อใหม่สะอาดกว่า ให้ปรับชื่อหลัก แต่ยังเป็นวิชาเดิม
        if len(name) >= len(sub.name or ''):
            sub.name = name
    return sub

def get_or_create_subject_for_teacher(name, teacher_id, code='', subject_name='', credit=1.0, total_periods=40):
    """
    รวมรายวิชาซ้ำเฉพาะครูคนเดียวกัน
    - ครูเดียวกัน + รหัสวิชาเดียวกัน = ใช้วิชาเดียว เพื่ออัปงานครั้งเดียวได้หลายห้อง
    - คนละครูสอน = แยกเป็นคนละวิชา แม้รหัส/ชื่อเหมือนกัน
    """
    display_name = make_subject_name(code, subject_name, fallback=name)
    key = subject_identity_key(display_name)
    sub = None
    if teacher_id:
        links = TeacherSubject.query.filter_by(teacher_id=teacher_id).all()
        for link in links:
            cand = Subject.query.get(link.subject_id)
            if cand and subject_identity_key(cand.name) == key:
                sub = cand
                break
    else:
        for cand in Subject.query.all():
            if subject_identity_key(cand.name) == key and not TeacherSubject.query.filter_by(subject_id=cand.id).first():
                sub = cand
                break
    if not sub:
        sub = Subject(name=display_name, teacher_id=teacher_id if teacher_id else None, credit=float(credit or 1), total_periods=int(total_periods or 40))
        db.session.add(sub); db.session.flush()
    else:
        if len(display_name) >= len(sub.name or ''):
            sub.name = display_name
        if teacher_id and not sub.teacher_id:
            sub.teacher_id = teacher_id
    ensure_teacher_subject(teacher_id, sub.id)
    ensure_grade_setting(sub.id)
    return sub

def merge_duplicate_subjects_for_teacher():
    """ยุบวิชาซ้ำที่เกิดจาก OCR/ช่องว่าง เฉพาะภายในครูคนเดียวกันเท่านั้น"""
    merged = 0
    teacher_ids = {x.teacher_id for x in TeacherSubject.query.all()}
    for teacher_id in teacher_ids:
        groups = {}
        for link in TeacherSubject.query.filter_by(teacher_id=teacher_id).all():
            sub = Subject.query.get(link.subject_id)
            if not sub:
                continue
            groups.setdefault(subject_identity_key(sub.name), []).append(sub)
        for key, subs in groups.items():
            if len(subs) <= 1:
                continue
            subs = sorted(subs, key=lambda x: (0 if TeachingSchedule.query.filter_by(subject_id=x.id).first() else 1, x.id))
            main = subs[0]
            for dup in subs[1:]:
                # ย้ายข้อมูลทุกตารางมาไว้ที่วิชาหลัก
                for model in [Unit, Assignment, TeachingSchedule, Attendance, ManualScore]:
                    model.query.filter_by(subject_id=dup.id).update({'subject_id': main.id})
                for gs in GradeSetting.query.filter_by(subject_id=dup.id).all():
                    if not GradeSetting.query.filter_by(subject_id=main.id).first():
                        gs.subject_id = main.id
                    else:
                        db.session.delete(gs)
                for sc in SubjectClassroom.query.filter_by(subject_id=dup.id).all():
                    if not SubjectClassroom.query.filter_by(subject_id=main.id, classroom_id=sc.classroom_id).first():
                        sc.subject_id = main.id
                    else:
                        db.session.delete(sc)
                TeacherSubject.query.filter_by(subject_id=dup.id).delete()
                # ถ้าวิชาซ้ำไม่มีครูอื่นใช้แล้ว ลบออก
                if not TeacherSubject.query.filter_by(subject_id=dup.id).first():
                    db.session.delete(dup)
                merged += 1
    db.session.flush()
    return merged

def period_time(period_no):
    times = {1:('08:40','09:30'),2:('09:31','10:20'),3:('10:31','11:20'),4:('11:21','12:10'),5:('13:00','13:50'),6:('13:51','14:40'),7:('14:41','15:30'),8:('15:41','16:00')}
    return times.get(int(period_no or 1), ('08:40','09:30'))

def weekday_value(v):
    txt = str(v).strip()
    mp = {'จันทร์':0,'mon':0,'monday':0,'0':0,'อังคาร':1,'tue':1,'tuesday':1,'1':1,'พุธ':2,'wed':2,'wednesday':2,'2':2,'พฤหัสบดี':3,'พฤหัส':3,'thu':3,'thursday':3,'3':3,'ศุกร์':4,'fri':4,'friday':4,'4':4,'เสาร์':5,'sat':5,'5':5,'อาทิตย์':6,'sun':6,'6':6}
    return mp.get(txt.lower(), mp.get(txt, 0))


IMPORT_MODULES = {
    'teachers': {'title':'นำเข้าครู', 'sheet':'Teachers', 'desc':'เพิ่ม/อัปเดตบัญชีครู'},
    'students': {'title':'นำเข้านักเรียน', 'sheet':'Students', 'desc':'เพิ่ม/อัปเดตนักเรียนและผูกห้องเรียน'},
    'subjects': {'title':'นำเข้ารายวิชา', 'sheet':'Subjects', 'desc':'เพิ่ม/อัปเดตรายวิชา'},
    'classrooms': {'title':'นำเข้าห้องเรียน', 'sheet':'Classrooms', 'desc':'เพิ่ม/อัปเดตห้องเรียน'},
    'teacher_assignments': {'title':'นำเข้าการมอบหมายครู', 'sheet':'TeacherAssignments', 'desc':'กำหนดว่าครูคนไหนสอนวิชา/ห้องไหน'},
    'units': {'title':'นำเข้าหน่วยการเรียนรู้', 'sheet':'Units', 'desc':'เพิ่มหน่วย ตัวชี้วัด และจำนวนคาบ'},
    'lessons': {'title':'นำเข้าบทเรียน', 'sheet':'Lessons', 'desc':'เพิ่มบทเรียน จุดประสงค์ เนื้อหา สื่อ'},
    'worksheets': {'title':'นำเข้าใบงาน', 'sheet':'Worksheets', 'desc':'เพิ่มใบงานตามบทเรียน'},
    'worksheet_questions': {'title':'นำเข้าข้อใบงาน', 'sheet':'WorksheetQuestions', 'desc':'เพิ่มข้อคำถามและคะแนนใบงาน'},
    'quizzes': {'title':'นำเข้าแบบทดสอบ', 'sheet':'Quizzes', 'desc':'เพิ่มแบบทดสอบและเกณฑ์ผ่าน'},
    'quiz_questions': {'title':'นำเข้าข้อสอบ', 'sheet':'QuizQuestions', 'desc':'เพิ่มข้อสอบ ตัวเลือก เฉลย คะแนน'},
    'schedule': {'title':'นำเข้าตารางสอน', 'sheet':'Schedule', 'desc':'เพิ่ม/อัปเดตตารางสอนรายคาบ'},
    'calendar': {'title':'นำเข้าปฏิทิน', 'sheet':'Calendar', 'desc':'เพิ่มวันหยุด กิจกรรม วันทำงาน'},
}
IMPORT_SHEET_ALIASES = {
    'teachers':['Teachers','ครู'],
    'students':['Students','นักเรียน'],
    'subjects':['Subjects','รายวิชา'],
    'classrooms':['Classrooms','ห้องเรียน'],
    'teacher_assignments':['TeacherAssignments','มอบหมายครู'],
    'units':['Units','หน่วย'],
    'lessons':['Lessons','บทเรียน'],
    'worksheets':['Worksheets','ใบงาน'],
    'worksheet_questions':['WorksheetQuestions','ข้อใบงาน'],
    'quizzes':['Quizzes','แบบทดสอบ'],
    'quiz_questions':['QuizQuestions','ข้อสอบ'],
    'schedule':['Schedule','ตารางสอน'],
    'calendar':['Calendar','ปฏิทิน'],
}

@app.route('/imports')
@login_required
@role_required('admin')
def imports_center():
    return render_template('imports.html', modules=IMPORT_MODULES)

@app.route('/imports/<kind>')
@login_required
@role_required('admin')
def import_module_page(kind):
    module = IMPORT_MODULES.get(kind)
    if not module:
        flash('ไม่พบหมวดนำเข้า', 'danger')
        return redirect(url_for('imports_center'))
    return render_template('import_module.html', kind=kind, module=module)

@app.route('/download_import_template/<kind>')
@login_required
@role_required('admin')
def download_import_template_kind(kind):
    if kind not in IMPORT_MODULES:
        flash('ไม่พบไฟล์ตัวอย่าง', 'danger')
        return redirect(url_for('imports_center'))
    path = os.path.join(BASE_DIR, 'import_templates', f'{kind}.xlsx')
    return send_file(path, as_attachment=True)

@app.route('/import_all', methods=['GET','POST'])
@app.route('/admin/import-system-excel', methods=['GET','POST'])
@login_required
@role_required('admin')
def import_all():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('กรุณาเลือกไฟล์ Excel', 'danger')
            return redirect(url_for('import_all'))
        wb = load_workbook(file)
        kind = request.form.get('kind', '').strip()
        if kind:
            allowed = set(IMPORT_SHEET_ALIASES.get(kind, []))
            for sheet_name in list(wb.sheetnames):
                if sheet_name not in allowed:
                    wb.remove(wb[sheet_name])
        report = []
        replace_schedule = bool(request.form.get('replace_schedule'))
        # Teachers
        n=0
        rows,h = sheet_rows(wb, ['Teachers','ครู'])
        for r in rows:
            username = cell(r,h,'username','teacher_username','รหัสครู') or cell(r,h,'citizen_id','เลขบัตร')
            full_name = cell(r,h,'full_name','name','ชื่อครู')
            if username and full_name:
                u = get_or_create_teacher(username, full_name)
                if cell(r,h,'password','รหัสผ่าน'):
                    u.set_password(cell(r,h,'password','รหัสผ่าน'))
                n += 1
        report.append(f'ครู {n} รายการ')
        # Classrooms
        n=0
        rows,h = sheet_rows(wb, ['Classrooms','ห้องเรียน'])
        for r in rows:
            name = cell(r,h,'classroom','classroom_name','ห้องเรียน','ห้อง')
            if name:
                room = get_or_create_classroom(name)
                teacher_username = cell(r,h,'teacher_username','ครู')
                if teacher_username:
                    t = User.query.filter_by(username=str(teacher_username).strip()).first()
                    if t:
                        room.teacher_id = t.id
                        if not TeacherClassroom.query.filter_by(teacher_id=t.id, classroom_id=room.id).first():
                            db.session.add(TeacherClassroom(teacher_id=t.id, classroom_id=room.id))
                n += 1
        report.append(f'ห้องเรียน {n} รายการ')
        # Students
        n=0
        rows,h = sheet_rows(wb, ['Students','นักเรียน'])
        for r in rows:
            full_name = cell(r,h,'full_name','name','ชื่อสกุล','ชื่อ-สกุล')
            citizen_id = str(cell(r,h,'citizen_id','เลขบัตร','เลขบัตรประชาชน') or '').strip()
            username = str(cell(r,h,'username','user') or citizen_id or cell(r,h,'student_code','เลขที่','รหัสนักเรียน')).strip()
            if not username or not full_name: continue
            u = User.query.filter_by(username=username).first()
            if not u:
                u = User(username=username, full_name=str(full_name).strip(), role='student', citizen_id=citizen_id, birth_date=str(cell(r,h,'birth_date','วันเกิด') or ''))
                u.set_password(birth_password(cell(r,h,'birth_date','วันเกิด')))
                db.session.add(u); db.session.flush()
            else:
                u.full_name=str(full_name).strip(); u.role='student'; u.citizen_id=citizen_id or u.citizen_id
            classroom_name = cell(r,h,'classroom','ห้องเรียน','ห้อง')
            if classroom_name:
                room = get_or_create_classroom(classroom_name)
                if not ClassroomStudent.query.filter_by(classroom_id=room.id, student_id=u.id).first():
                    db.session.add(ClassroomStudent(classroom_id=room.id, student_id=u.id))
            n += 1
        report.append(f'นักเรียน {n} รายการ')
        # Subjects
        n=0
        rows,h = sheet_rows(wb, ['Subjects','รายวิชา'])
        for r in rows:
            subject_code = cell(r,h,'subject_code','code','รหัสวิชา','รหัส')
            subject_name_only = cell(r,h,'subject_name','วิชา','ชื่อวิชา')
            name = cell(r,h,'subject','รายวิชา') or make_subject_name(subject_code, subject_name_only)
            if name:
                teacher_username = cell(r,h,'teacher_username','ครู')
                t = User.query.filter_by(username=str(teacher_username).strip()).first() if teacher_username else None
                if t:
                    sub = get_or_create_subject_for_teacher(name, t.id, subject_code, subject_name_only, cell(r,h,'credit','หน่วยกิต', default=1), cell(r,h,'total_periods','จำนวนคาบ', default=40))
                else:
                    sub = get_or_create_subject(name, cell(r,h,'credit','หน่วยกิต', default=1), cell(r,h,'total_periods','จำนวนคาบ', default=40))
                n += 1
        report.append(f'รายวิชา {n} รายการ')
        # TeacherAssignments
        n=0
        rows,h = sheet_rows(wb, ['TeacherAssignments','มอบหมายครู'])
        for r in rows:
            t = User.query.filter_by(username=str(cell(r,h,'teacher_username','ครู')).strip()).first()
            if not t: continue
            subject_name = cell(r,h,'subject','subject_name','รายวิชา')
            classroom_name = cell(r,h,'classroom','classroom_name','ห้องเรียน')
            if subject_name:
                sub = get_or_create_subject_for_teacher(subject_name, t.id)
                if not TeacherSubject.query.filter_by(teacher_id=t.id, subject_id=sub.id).first(): db.session.add(TeacherSubject(teacher_id=t.id, subject_id=sub.id))
                n += 1
            if classroom_name:
                room = get_or_create_classroom(classroom_name)
                if not TeacherClassroom.query.filter_by(teacher_id=t.id, classroom_id=room.id).first(): db.session.add(TeacherClassroom(teacher_id=t.id, classroom_id=room.id))
                n += 1
        report.append(f'มอบหมายครู {n} รายการ')
        # Units
        n=0
        rows,h = sheet_rows(wb, ['Units','หน่วยการเรียนรู้'])
        for r in rows:
            subject_name = cell(r,h,'subject','subject_name','รายวิชา')
            title = cell(r,h,'unit','unit_title','หน่วย','ชื่อหน่วย')
            if subject_name and title:
                sub = get_or_create_subject(subject_name)
                u = Unit.query.filter_by(subject_id=sub.id, title=str(title).strip()).first()
                if not u:
                    u = Unit(subject_id=sub.id, title=str(title).strip())
                    db.session.add(u)
                u.indicators = str(cell(r,h,'indicators','ตัวชี้วัด') or '')
                u.required_periods = int(cell(r,h,'required_periods','จำนวนคาบ', default=1) or 1)
                n += 1
        report.append(f'หน่วยการเรียนรู้ {n} รายการ')
        db.session.flush()
        # Lessons
        n=0
        rows,h = sheet_rows(wb, ['Lessons','บทเรียน'])
        for r in rows:
            subject_name = cell(r,h,'subject','subject_name','รายวิชา')
            unit_title = cell(r,h,'unit','unit_title','หน่วย')
            title = cell(r,h,'lesson','lesson_title','บทเรียน','ชื่อบทเรียน')
            if subject_name and unit_title and title:
                sub = get_or_create_subject(subject_name)
                u = Unit.query.filter_by(subject_id=sub.id, title=str(unit_title).strip()).first() or Unit(subject_id=sub.id, title=str(unit_title).strip())
                db.session.add(u); db.session.flush()
                l = Lesson.query.filter_by(unit_id=u.id, title=str(title).strip()).first()
                if not l:
                    l = Lesson(unit_id=u.id, title=str(title).strip())
                    db.session.add(l)
                l.objective=str(cell(r,h,'objective','จุดประสงค์') or '')
                l.content=str(cell(r,h,'content','เนื้อหา') or '')
                l.media_url=str(cell(r,h,'media_url','สื่อ') or '')
                l.required_minutes=int(cell(r,h,'required_minutes','นาที', default=50) or 50)
                n += 1
        report.append(f'บทเรียน {n} รายการ')
        db.session.flush()
        # Worksheets
        n=0
        rows,h = sheet_rows(wb, ['Worksheets','ใบงาน'])
        for r in rows:
            lesson_title = cell(r,h,'lesson','lesson_title','บทเรียน')
            title = cell(r,h,'worksheet','worksheet_title','ใบงาน','ชื่อใบงาน')
            l = Lesson.query.filter_by(title=str(lesson_title).strip()).first() if lesson_title else None
            if l and title:
                ws = Worksheet.query.filter_by(lesson_id=l.id, title=str(title).strip()).first()
                if not ws:
                    ws = Worksheet(lesson_id=l.id, title=str(title).strip())
                    db.session.add(ws)
                ws.worksheet_type=str(cell(r,h,'worksheet_type','ประเภท', default='text') or 'text')
                ws.description=str(cell(r,h,'description','คำชี้แจง') or '')
                n += 1
        report.append(f'ใบงาน {n} รายการ')
        db.session.flush()
        # WorksheetQuestions
        n=0
        rows,h = sheet_rows(wb, ['WorksheetQuestions','ข้อใบงาน'])
        for r in rows:
            ws_title = cell(r,h,'worksheet','worksheet_title','ใบงาน')
            qtext = cell(r,h,'question_text','คำถาม')
            ws = Worksheet.query.filter_by(title=str(ws_title).strip()).first() if ws_title else None
            if ws and qtext:
                num=int(cell(r,h,'number','ข้อ', default=1) or 1)
                q = WorksheetQuestion.query.filter_by(worksheet_id=ws.id, number=num).first()
                if not q:
                    q = WorksheetQuestion(worksheet_id=ws.id, number=num, question_text=str(qtext))
                    db.session.add(q)
                q.question_text=str(qtext); q.answer_type=str(cell(r,h,'answer_type','ประเภทคำตอบ', default='text') or 'text'); q.max_score=float(cell(r,h,'max_score','คะแนน', default=1) or 1)
                n += 1
        report.append(f'ข้อใบงาน {n} รายการ')
        # Quizzes
        n=0
        rows,h = sheet_rows(wb, ['Quizzes','แบบทดสอบ'])
        for r in rows:
            lesson_title = cell(r,h,'lesson','lesson_title','บทเรียน')
            title = cell(r,h,'quiz','quiz_title','แบบทดสอบ')
            l = Lesson.query.filter_by(title=str(lesson_title).strip()).first() if lesson_title else None
            if l and title:
                qz = Quiz.query.filter_by(lesson_id=l.id, title=str(title).strip()).first()
                if not qz:
                    qz = Quiz(lesson_id=l.id, title=str(title).strip())
                    db.session.add(qz)
                qz.pass_percent=int(cell(r,h,'pass_percent','ผ่านร้อยละ', default=60) or 60)
                n += 1
        report.append(f'แบบทดสอบ {n} รายการ')
        db.session.flush()
        # QuizQuestions
        n=0
        rows,h = sheet_rows(wb, ['QuizQuestions','ข้อสอบ'])
        for r in rows:
            quiz_title = cell(r,h,'quiz','quiz_title','แบบทดสอบ')
            qtext = cell(r,h,'question_text','คำถาม')
            qz = Quiz.query.filter_by(title=str(quiz_title).strip()).first() if quiz_title else None
            if qz and qtext:
                qq = QuizQuestion(quiz_id=qz.id, question_text=str(qtext), question_type=str(cell(r,h,'question_type','ประเภท', default='choice') or 'choice'), choices=str(cell(r,h,'choices','ตัวเลือก') or '').replace('|','\n'), correct_answer=str(cell(r,h,'correct_answer','เฉลย') or ''), score=float(cell(r,h,'score','คะแนน', default=1) or 1))
                db.session.add(qq); n += 1
        report.append(f'ข้อสอบ {n} รายการ')
        # Schedule
        n=0; made_schedule_teachers=0; skipped_schedule=0
        rows,h = sheet_rows(wb, ['Schedule','ตารางสอน'])
        if rows and replace_schedule:
            TeachingSchedule.query.delete()
            db.session.flush()

        def import_username_from_name(name):
            import re
            base = re.sub(r'[^a-zA-Z0-9ก-๙]+', '', str(name or '').replace('ครู','').strip()) or 'teacher'
            username = base[:40]
            if not User.query.filter_by(username=username).first():
                return username
            i = 2
            while User.query.filter_by(username=f'{username}{i}').first():
                i += 1
            return f'{username}{i}'

        def get_or_create_import_teacher(username, full_name, password='1234'):
            username = str(username or '').strip()
            full_name = str(full_name or '').strip()
            teacher = None
            if username:
                teacher = User.query.filter_by(username=username).first()
            if not teacher and full_name:
                teacher = User.query.filter_by(full_name=full_name).first()
            created = False
            if not teacher and (username or full_name):
                teacher = User(username=username or import_username_from_name(full_name), full_name=full_name or username, role='teacher')
                teacher.set_password(str(password or '1234'))
                db.session.add(teacher); db.session.flush()
                created = True
            if teacher:
                teacher.role = 'teacher'
                if full_name: teacher.full_name = full_name
            return teacher, created

        for r in rows:
            teacher_username = cell(r,h,'teacher_username','username','รหัสครู')
            teacher_full_name = cell(r,h,'teacher_full_name','teacher_name','full_name','ครู','ชื่อครู','ผู้สอน')
            teacher_password = cell(r,h,'password','รหัสผ่าน', default='1234')
            t, created_teacher = get_or_create_import_teacher(teacher_username, teacher_full_name, teacher_password)
            if created_teacher: made_schedule_teachers += 1

            subject_code = str(cell(r,h,'subject_code','code','รหัสวิชา','รหัส', default='') or '').strip()
            subject_name_only = str(cell(r,h,'subject_name','วิชา','ชื่อวิชา', default='') or '').strip()
            subject_name = cell(r,h,'subject','รายวิชา') or ' '.join([x for x in [subject_code, subject_name_only] if x]).strip()
            classroom_name = cell(r,h,'classroom','classroom_name','ห้องเรียน','ห้อง','ชั้น/ห้อง')
            if not (t and subject_name and classroom_name):
                skipped_schedule += 1
                continue
            sub = get_or_create_subject_for_teacher(subject_name, t.id, subject_code, subject_name_only); room = get_or_create_classroom(classroom_name)
            if not TeacherSubject.query.filter_by(teacher_id=t.id, subject_id=sub.id).first():
                db.session.add(TeacherSubject(teacher_id=t.id, subject_id=sub.id))
            if not TeacherClassroom.query.filter_by(teacher_id=t.id, classroom_id=room.id).first():
                db.session.add(TeacherClassroom(teacher_id=t.id, classroom_id=room.id))
            link_subject_classroom(sub.id, room.id)
            try:
                period_no = int(float(cell(r,h,'period_no','period','คาบ', default=1) or 1))
            except Exception:
                skipped_schedule += 1
                continue
            wd = weekday_value(cell(r,h,'weekday','day','วัน'))
            st, et = period_time(period_no)
            st = str(cell(r,h,'start_time','เวลาเริ่ม', default=st) or st)[:5]
            et = str(cell(r,h,'end_time','เวลาจบ', default=et) or et)[:5]
            sched = TeachingSchedule.query.filter_by(teacher_id=t.id, weekday=wd, period_no=period_no, classroom_id=room.id).first()
            if not sched:
                sched = TeachingSchedule(teacher_id=t.id, subject_id=sub.id, classroom_id=room.id, weekday=wd, period_no=period_no, start_time=st, end_time=et)
                db.session.add(sched)
            sched.subject_id=sub.id; sched.classroom_id=room.id; sched.start_time=st; sched.end_time=et; sched.room_name=str(cell(r,h,'room_name','สถานที่') or classroom_name or ''); sched.topic=str(cell(r,h,'topic','หัวข้อ') or '')
            n += 1
        report.append(f'ตารางสอน {n} รายการ')
        if made_schedule_teachers:
            report.append(f'สร้างครูจากตารางสอน {made_schedule_teachers} คน')
        if skipped_schedule:
            report.append(f'ข้ามแถวตารางสอน {skipped_schedule} แถว')
        merged_subjects = merge_duplicate_subjects_for_teacher()
        if merged_subjects:
            report.append(f'ยุบรายวิชาซ้ำ {merged_subjects} รายการ')
        # Calendar
        n=0
        rows,h = sheet_rows(wb, ['Calendar','ปฏิทิน'])
        for r in rows:
            d = parse_excel_date_value(cell(r,h,'date','event_date','วันที่'))
            title = cell(r,h,'title','รายการ','กิจกรรม')
            if d and title:
                typ = str(cell(r,h,'event_type','ประเภท', default='กิจกรรมโรงเรียน') or 'กิจกรรมโรงเรียน')
                note = str(cell(r,h,'note','หมายเหตุ') or '')
                if not CalendarEvent.query.filter_by(teacher_id=None, event_date=d, title=str(title).strip()).first():
                    db.session.add(CalendarEvent(teacher_id=None, event_date=d, title=str(title).strip(), event_type=typ, note=note))
                n += 1
        report.append(f'ปฏิทิน {n} รายการ')
        db.session.commit()
        flash('นำเข้าข้อมูลสำเร็จ: ' + ' / '.join(report), 'success')
        if kind:
            return redirect(request.referrer or url_for('import_module_page', kind=kind))
        return redirect(url_for('import_all'))
    return render_template('import_all.html')


@app.route('/download_term_import_file')
@login_required
@role_required('admin')
def download_term_import_file():
    path = os.path.join(BASE_DIR, 'krurakson_term_1_2569_import.xlsx')
    return send_file(path, as_attachment=True)

@app.route('/download_import_template')
@login_required
@role_required('admin')
def download_import_template():
    return send_file(os.path.join(BASE_DIR, 'krurakson_import_template.xlsx'), as_attachment=True)

@app.route('/subjects', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def subjects():
    if request.method == 'POST':
        teacher_id_raw = request.form.get('teacher_id')
        teacher_id = current_user.id if current_user.role == 'teacher' else (int(teacher_id_raw) if teacher_id_raw else None)
        sub = Subject(name=request.form.get('name',''), teacher_id=teacher_id, credit=float(request.form.get('credit',1)), total_periods=int(request.form.get('total_periods',40)))
        db.session.add(sub); db.session.flush()
        if teacher_id and not TeacherSubject.query.filter_by(teacher_id=teacher_id, subject_id=sub.id).first():
            db.session.add(TeacherSubject(teacher_id=teacher_id, subject_id=sub.id))
        db.session.add(GradeSetting(subject_id=sub.id))
        db.session.commit()
        flash('สร้างรายวิชาแล้ว', 'success')
        return redirect(url_for('subjects'))
    return render_template('subjects.html', subjects=teacher_filter(Subject).all(), teachers=active_teachers_query().all(), classrooms=teacher_filter(Classroom).order_by(Classroom.name).all(), subject_links=TeacherSubject.query.all(), classroom_links=SubjectClassroom.query.all())


@app.route('/subjects/<int:subject_id>/classrooms/add', methods=['POST'])
@login_required
@role_required('teacher','admin')
def subject_classroom_add(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject):
        return deny_redirect('subjects')
    try:
        classroom_id = int(request.form.get('classroom_id') or 0)
    except Exception:
        classroom_id = 0
    classroom = Classroom.query.get(classroom_id)
    if not classroom:
        flash('ไม่พบห้องเรียนที่เลือก', 'danger')
        return redirect(url_for('subjects'))
    if current_user.role == 'teacher' and not TeacherClassroom.query.filter_by(teacher_id=current_user.id, classroom_id=classroom.id).first():
        flash('ยังไม่ได้มอบหมายห้องนี้ให้ครู', 'danger')
        return redirect(url_for('subjects'))
    link_subject_classroom(subject.id, classroom.id)
    if current_user.role == 'teacher' and not TeacherSubject.query.filter_by(teacher_id=current_user.id, subject_id=subject.id).first():
        db.session.add(TeacherSubject(teacher_id=current_user.id, subject_id=subject.id))
    db.session.commit()
    flash(f'เพิ่มห้อง {classroom.name} ให้รายวิชา {subject.name} แล้ว', 'success')
    return redirect(url_for('subjects'))

@app.route('/subjects/<int:subject_id>/classrooms/<int:classroom_id>/remove', methods=['POST'])
@login_required
@role_required('teacher','admin')
def subject_classroom_remove(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject):
        return deny_redirect('subjects')
    link = SubjectClassroom.query.filter_by(subject_id=subject_id, classroom_id=classroom_id).first()
    if link:
        db.session.delete(link)
        db.session.commit()
        flash('นำห้องเรียนออกจากรายวิชาแล้ว', 'success')
    return redirect(url_for('subjects'))

@app.route('/subjects/<int:subject_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def subject_edit(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject): return deny_redirect('subjects')
    if request.method == 'POST':
        subject.name = request.form.get('name','')
        subject.credit = float(request.form.get('credit',1))
        subject.total_periods = int(request.form.get('total_periods',40))
        if current_user.role == 'admin':
            teacher_id_raw = request.form.get('teacher_id')
            subject.teacher_id = int(teacher_id_raw) if teacher_id_raw else None
        db.session.commit(); flash('แก้ไขรายวิชาแล้ว', 'success')
        return redirect(url_for('subjects'))
    return render_template('subject_form.html', subject=subject, teachers=active_teachers_query().all())

def subject_has_history(subject_id):
    return any([
        Unit.query.filter_by(subject_id=subject_id).first(),
        Assignment.query.filter_by(subject_id=subject_id).first(),
        TeachingSchedule.query.filter_by(subject_id=subject_id).first(),
        Attendance.query.filter_by(subject_id=subject_id).first(),
        ManualScore.query.filter_by(subject_id=subject_id).first(),
        SubjectClassroom.query.filter_by(subject_id=subject_id).first(),
    ])

@app.route('/subjects/<int:subject_id>/deactivate', methods=['POST'])
@login_required
@role_required('teacher','admin')
def subject_deactivate(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject): return deny_redirect('subjects')
    subject.is_active = False
    # ถอดจากตารางสอนใหม่ไม่ได้ แต่เก็บประวัติไว้
    db.session.commit(); flash('ปิดใช้งานรายวิชาแล้ว ข้อมูลเดิมยังอยู่ ไม่หาย', 'success')
    return redirect(url_for('subjects'))

@app.route('/subjects/<int:subject_id>/restore', methods=['POST'])
@login_required
@role_required('admin')
def subject_restore(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    subject.is_active = True
    db.session.commit(); flash('เปิดใช้งานรายวิชาแล้ว', 'success')
    return redirect(url_for('subjects'))

@app.route('/subjects/<int:subject_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def subject_delete(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject): return deny_redirect('subjects')
    subject_name = subject.name
    force_delete_subject_data(subject.id)
    db.session.delete(subject)
    db.session.commit()
    flash(f'ลบรายวิชา {subject_name} และข้อมูลที่ผูกอยู่ทั้งหมดแล้ว', 'success')
    return redirect(url_for('subjects'))


@app.route('/subject/<int:subject_id>/download_lesson_import_template')
@login_required
@role_required('teacher','admin')
def download_subject_lesson_import_template(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject):
        return deny_redirect('subjects')
    path = os.path.join(BASE_DIR, 'import_templates', 'subject_lessons.xlsx')
    return send_file(path, as_attachment=True)


@app.route('/subject/<int:subject_id>/import_lessons_excel', methods=['POST'])
@login_required
@role_required('teacher','admin')
def subject_import_lessons_excel(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject):
        return deny_redirect('subjects')
    file = request.files.get('file')
    if not file or not file.filename:
        flash('กรุณาเลือกไฟล์ Excel สำหรับนำเข้าหน่วย/บทเรียน', 'danger')
        return redirect(url_for('subject_detail', subject_id=subject.id))
    try:
        wb = load_workbook(file)
    except Exception:
        flash('เปิดไฟล์ Excel ไม่สำเร็จ กรุณาใช้ไฟล์ .xlsx', 'danger')
        return redirect(url_for('subject_detail', subject_id=subject.id))

    # โหมดนำเข้า:
    # update = อัปเดตข้อมูลเดิมเมื่อชื่อหน่วย/บทเรียนตรงกัน และเพิ่มรายการใหม่
    # keep = คงข้อมูลเดิมไว้ ถ้าชื่อตรงกันจะไม่ทับ แต่ยังเพิ่มรายการใหม่
    # replace_matching_units = ถ้าชื่อหน่วยตรงกัน ให้ลบหน่วยเดิมพร้อมข้อมูลลูก แล้วนำเข้าใหม่
    # replace_all_units = ลบหน่วยทั้งหมดในรายวิชานี้ก่อน แล้วนำเข้าใหม่ทั้งไฟล์
    import_mode = request.form.get('import_mode', 'update').strip() or 'update'
    if import_mode not in ['update', 'keep', 'replace_matching_units', 'replace_all_units']:
        import_mode = 'update'

    imported_units = 0
    imported_lessons = 0
    updated_units = 0
    updated_lessons = 0
    kept_units = 0
    kept_lessons = 0
    replaced_units = 0
    skipped = 0
    replaced_title_cache = set()

    if import_mode == 'replace_all_units':
        old_units = Unit.query.filter_by(subject_id=subject.id).all()
        for old_unit in old_units:
            force_delete_unit_data(old_unit.id)
            db.session.delete(old_unit)
            replaced_units += 1
        db.session.flush()

    def get_or_prepare_unit(unit_title):
        nonlocal imported_units, replaced_units, kept_units
        unit_title = str(unit_title).strip()
        unit = Unit.query.filter_by(subject_id=subject.id, title=unit_title).first()
        if unit and import_mode == 'replace_matching_units' and unit_title not in replaced_title_cache:
            force_delete_unit_data(unit.id)
            db.session.delete(unit)
            db.session.flush()
            replaced_title_cache.add(unit_title)
            replaced_units += 1
            unit = None
        if not unit:
            unit = Unit(subject_id=subject.id, title=unit_title)
            db.session.add(unit)
            db.session.flush()
            imported_units += 1
        else:
            kept_units += 1
        return unit

    def apply_unit_fields(unit, r, h):
        nonlocal updated_units
        if import_mode == 'keep' and unit.id:
            return
        changed = False
        indicators = cell(r,h,'indicators','indicator','ตัวชี้วัด','ผลการเรียนรู้')
        if indicators not in (None, '') and str(unit.indicators or '') != str(indicators):
            unit.indicators = str(indicators); changed = True
        rp = cell(r,h,'required_periods','periods','จำนวนคาบ','ชั่วโมง')
        if rp not in (None, ''):
            try:
                rp_int = int(float(rp))
                if unit.required_periods != rp_int:
                    unit.required_periods = rp_int; changed = True
            except Exception:
                pass
        if changed:
            updated_units += 1

    def apply_lesson_fields(lesson, r, h):
        nonlocal updated_lessons, kept_lessons
        if import_mode == 'keep':
            kept_lessons += 1
            return
        changed = False
        objective = cell(r,h,'objective','จุดประสงค์')
        content = cell(r,h,'content','เนื้อหา','รายละเอียด')
        media_url = cell(r,h,'media_url','สื่อ','ลิงก์สื่อ')
        required_minutes = cell(r,h,'required_minutes','minutes','นาที')
        if objective not in (None, '') and str(lesson.objective or '') != str(objective):
            lesson.objective = str(objective); changed = True
        if content not in (None, '') and str(lesson.content or '') != str(content):
            lesson.content = str(content); changed = True
        if media_url not in (None, '') and str(lesson.media_url or '') != str(media_url):
            lesson.media_url = str(media_url); changed = True
        if required_minutes not in (None, ''):
            try:
                minutes_int = int(float(required_minutes))
                if lesson.required_minutes != minutes_int:
                    lesson.required_minutes = minutes_int; changed = True
            except Exception:
                pass
        if changed:
            updated_lessons += 1
        else:
            kept_lessons += 1

    # รองรับทั้งชีตเดียว LessonImport และไฟล์เดิมที่แยก Units / Lessons
    rows, h = sheet_rows(wb, ['LessonImport','บทเรียนรายวิชา','Lessons','บทเรียน'])
    for r in rows:
        unit_title = cell(r,h,'unit_title','unit','หน่วย','ชื่อหน่วย')
        lesson_title = cell(r,h,'lesson_title','lesson','บทเรียน','ชื่อบทเรียน')
        if not unit_title and not lesson_title:
            continue
        if not unit_title:
            skipped += 1
            continue

        unit = get_or_prepare_unit(unit_title)
        apply_unit_fields(unit, r, h)

        if lesson_title:
            lesson_title = str(lesson_title).strip()
            lesson = Lesson.query.filter_by(unit_id=unit.id, title=lesson_title).first()
            if not lesson:
                lesson = Lesson(unit_id=unit.id, title=lesson_title)
                db.session.add(lesson)
                imported_lessons += 1
            apply_lesson_fields(lesson, r, h)

    # ถ้าไฟล์เป็นรูปแบบแยก Units ให้เก็บหน่วยจากชีต Units เพิ่มเติมด้วย
    rows_u, h_u = sheet_rows(wb, ['Units','หน่วยการเรียนรู้','หน่วย'])
    for r in rows_u:
        unit_title = cell(r,h_u,'unit_title','unit','หน่วย','ชื่อหน่วย')
        if not unit_title:
            continue
        unit = get_or_prepare_unit(unit_title)
        apply_unit_fields(unit, r, h_u)

    db.session.commit()
    mode_label = {
        'update': 'อัปเดตของเดิม + เพิ่มรายการใหม่',
        'keep': 'คงข้อมูลเดิมไว้ + เพิ่มเฉพาะรายการใหม่',
        'replace_matching_units': 'ทำทับเฉพาะหน่วยที่ชื่อซ้ำ',
        'replace_all_units': 'ลบหน่วยเดิมทั้งหมดแล้วนำเข้าใหม่'
    }.get(import_mode, import_mode)
    msg = (f'นำเข้าหน่วย/บทเรียนให้รายวิชา {subject.name} สำเร็จ '
           f'({mode_label}): หน่วยใหม่ {imported_units} / บทเรียนใหม่ {imported_lessons} '
           f'/ อัปเดตหน่วย {updated_units} / อัปเดตบทเรียน {updated_lessons} '
           f'/ คงไว้ {kept_lessons} รายการ / ลบแทนที่ {replaced_units} หน่วย')
    if skipped:
        msg += f' / ข้าม {skipped} แถว'
    flash(msg, 'success')
    return redirect(url_for('subject_detail', subject_id=subject.id))


@app.route('/subject/<int:subject_id>', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def subject_detail(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject): flash('ไม่มีสิทธิ์','danger'); return redirect(url_for('subjects'))
    if request.method == 'POST':
        db.session.add(Unit(subject_id=subject.id, title=request.form.get('title',''), indicators=request.form.get('indicators',''), required_periods=int(request.form.get('required_periods',1))))
        db.session.commit(); return redirect(url_for('subject_detail', subject_id=subject.id))

    # ใช้เส้นทางเดียวให้ชัดเจน: รายวิชา -> หน่วย -> บทเรียน -> ใบงาน/แบบทดสอบ
    # หน้านี้จึงไม่แสดงแค่จำนวนรวม แต่แตกบทเรียนย่อยใต้หน่วย เพื่อกันอัปใบงานผิดบทเรียนแล้วหาไม่เจอ
    units = Unit.query.filter_by(subject_id=subject.id).order_by(Unit.id.asc()).all()
    completeness = []
    for u in units:
        lessons = Lesson.query.filter_by(unit_id=u.id).order_by(Lesson.id.asc()).all()
        lesson_rows = []
        worksheet_total = 0
        quiz_total = 0
        file_total = 0
        for lesson in lessons:
            worksheet_count = Worksheet.query.filter_by(lesson_id=lesson.id).count()
            quiz_count = Quiz.query.filter_by(lesson_id=lesson.id).count()
            file_count = LessonFile.query.filter_by(lesson_id=lesson.id).count()
            worksheet_total += worksheet_count
            quiz_total += quiz_count
            file_total += file_count
            lesson_rows.append(SimpleNamespace(
                lesson=lesson,
                worksheet_count=worksheet_count,
                quiz_count=quiz_count,
                file_count=file_count,
            ))

        lesson_count = len(lessons)
        percent = min(100, int((lesson_count / max(1, u.required_periods)) * 100))
        completeness.append(SimpleNamespace(
            unit=u,
            lesson_count=lesson_count,
            required_periods=u.required_periods,
            worksheets=worksheet_total,
            quizzes=quiz_total,
            files=file_total,
            percent=percent,
            missing=max(0, u.required_periods - lesson_count),
            lesson_rows=lesson_rows,
        ))
    subject_rooms = subject_classrooms_for_user(subject.id)
    return render_template('subject_detail.html', subject=subject, units=units, completeness=completeness, subject_rooms=subject_rooms)

@app.route('/unit/<int:unit_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def unit_edit(unit_id):
    unit = Unit.query.get_or_404(unit_id)
    if not owns_unit(unit): return deny_redirect('subjects')
    if request.method == 'POST':
        unit.title = request.form.get('title','')
        unit.indicators = request.form.get('indicators','')
        unit.required_periods = int(request.form.get('required_periods',1))
        db.session.commit(); flash('แก้ไขหน่วยการเรียนรู้แล้ว', 'success')
        return redirect(url_for('subject_detail', subject_id=unit.subject_id))
    return render_template('unit_form.html', unit=unit)

@app.route('/unit/<int:unit_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def unit_delete(unit_id):
    unit = Unit.query.get_or_404(unit_id)
    if not owns_unit(unit): return deny_redirect('subjects')
    subject_id = unit.subject_id
    unit_title = unit.title
    force_delete_unit_data(unit.id)
    db.session.delete(unit)
    db.session.commit()
    flash(f'ลบหน่วยการเรียนรู้ {unit_title} พร้อมบทเรียน/ใบงาน/แบบทดสอบที่ผูกอยู่แล้ว', 'success')
    return redirect(url_for('subject_detail', subject_id=subject_id))


@app.route('/subject/<int:subject_id>/units/bulk-delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def subject_units_bulk_delete(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject):
        return deny_redirect('subjects')
    raw_ids = request.form.getlist('unit_ids')
    unit_ids = []
    for x in raw_ids:
        try:
            unit_ids.append(int(x))
        except Exception:
            pass
    if not unit_ids:
        flash('กรุณาติ๊กเลือกหน่วยการเรียนรู้ที่ต้องการลบก่อน', 'danger')
        return redirect(url_for('subject_detail', subject_id=subject.id))
    units = Unit.query.filter(Unit.subject_id==subject.id, Unit.id.in_(unit_ids)).all()
    deleted_titles = []
    for unit in units:
        deleted_titles.append(unit.title)
        force_delete_unit_data(unit.id)
        db.session.delete(unit)
    db.session.commit()
    flash(f'ลบหน่วยการเรียนรู้ {len(deleted_titles)} หน่วย พร้อมบทเรียน/ใบงาน/แบบทดสอบที่ผูกอยู่แล้ว', 'success')
    return redirect(url_for('subject_detail', subject_id=subject.id))



def normalize_thai_title(value):
    """ทำชื่อบทเรียนให้เทียบกันง่ายขึ้น เช่น ตัดช่องว่าง/สัญลักษณ์/คำหน้าท้ายที่ไม่จำเป็น"""
    txt = str(value or '').strip().lower()
    txt = re.sub(r'\.(jpg|jpeg|png|webp|gif|pdf|docx?|pptx?|md)$', '', txt, flags=re.I)
    txt = txt.replace('ใบงานครูรัก', '').replace('ใบความรู้ครูรัก', '').replace('บทเรียนลงระบบ', '')
    txt = re.sub(r'ม\.?\s*[123]', '', txt)
    txt = re.sub(r'คาบ\s*\d+', '', txt)
    txt = re.sub(r'[_\-–—:：|/\\\s]+', '', txt)
    return txt


def detect_subject_level(subject):
    """เดาระดับชั้นจากชื่อ/รหัสวิชา เพื่อกรองไฟล์ใน ZIP ม.1-ม.3 ให้ตรงรายวิชา"""
    txt = f"{getattr(subject, 'code', '')} {getattr(subject, 'name', '')}".lower()
    if re.search(r'ม\.?\s*1|ม\.1|ม1|ว21', txt):
        return 'ม1'
    if re.search(r'ม\.?\s*2|ม\.2|ม2|ว22', txt):
        return 'ม2'
    if re.search(r'ม\.?\s*3|ม\.3|ม3|ว23', txt):
        return 'ม3'
    return ''


def parse_lesson_doc_filename(name):
    """อ่านชื่อไฟล์เอกสารครูรัก เช่น ม.1_คาบ01_เทคโนโลยีคืออะไร_ใบงานครูรัก.jpg"""
    base = os.path.basename(name)
    base_no_ext = os.path.splitext(base)[0]
    level = ''
    if re.search(r'ม\.?\s*1|ม1', name): level = 'ม1'
    elif re.search(r'ม\.?\s*2|ม2', name): level = 'ม2'
    elif re.search(r'ม\.?\s*3|ม3', name): level = 'ม3'
    m = re.search(r'คาบ\s*0*(\d+)', base_no_ext)
    period = int(m.group(1)) if m else None
    kind = 'other'
    if 'ใบงาน' in base_no_ext:
        kind = 'worksheet'
    elif 'ใบความรู้' in base_no_ext:
        kind = 'knowledge'
    elif 'บทเรียนลงระบบ' in base_no_ext or base.lower().endswith('.md'):
        kind = 'lesson_md'
    title = base_no_ext
    title = re.sub(r'^ม\.?\s*[123][_-]?', '', title)
    title = re.sub(r'คาบ\s*0*\d+[_-]?', '', title)
    title = title.replace('ใบงานครูรัก', '').replace('ใบความรู้ครูรัก', '').replace('บทเรียนลงระบบ', '')
    title = re.sub(r'[_\-–—]+', ' ', title).strip()
    return SimpleNamespace(level=level, period=period, title=title, kind=kind, basename=base)


def extract_markdown_section(md, heading):
    """ดึงเนื้อหาใต้หัวข้อ markdown ระดับ ## แบบง่าย"""
    if not md:
        return ''
    pattern = rf'(?ms)^##\s*{re.escape(heading)}\s*\n(.*?)(?=^##\s+|\Z)'
    m = re.search(pattern, md)
    return m.group(1).strip() if m else ''


def save_bytes_to_upload(data, original_name, subdir):
    """บันทึกไฟล์จาก ZIP ลง uploads โดยไม่พึ่ง FileStorage"""
    ext = os.path.splitext(original_name)[1].lower().lstrip('.')
    if ext not in ALLOWED_WORKSHEET_EXTENSIONS and ext not in ALLOWED_IMAGE_EXTENSIONS and ext not in {'md', 'txt'}:
        raise ValueError(f'ชนิดไฟล์ไม่รองรับ: {original_name}')
    target_dir = os.path.join(app.config['UPLOAD_FOLDER'], subdir)
    os.makedirs(target_dir, exist_ok=True)
    safe = secure_filename(original_name) or f'upload.{ext or "file"}'
    stamp = datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
    filename = f"{stamp}_{safe}"
    with open(os.path.join(target_dir, filename), 'wb') as out:
        out.write(data)
    return f"{subdir}/{filename}", original_name


def build_lesson_match_map(subject):
    lessons = (Lesson.query.join(Unit)
               .filter(Unit.subject_id == subject.id)
               .order_by(Unit.id.asc(), Lesson.id.asc())
               .all())
    by_norm = {normalize_thai_title(l.title): l for l in lessons if normalize_thai_title(l.title)}
    by_period = {i: l for i, l in enumerate(lessons, start=1)}
    return lessons, by_norm, by_period


@app.route('/subject/<int:subject_id>/import_lesson_docs_zip', methods=['POST'])
@login_required
@role_required('teacher','admin')
def subject_import_lesson_docs_zip(subject_id):
    """นำเข้า ZIP เอกสารบทเรียนครูรัก: .md เป็นเนื้อหาบทเรียน, ใบความรู้เป็นไฟล์แนบ, ใบงานเป็น Worksheet"""
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject):
        return deny_redirect('subjects')

    upload = request.files.get('docs_zip') or request.files.get('file')
    if not upload or not upload.filename:
        flash('กรุณาเลือกไฟล์ ZIP เอกสารบทเรียน', 'danger')
        return redirect(url_for('subject_detail', subject_id=subject.id))
    if not upload.filename.lower().endswith('.zip'):
        flash('กรุณาอัปโหลดไฟล์ .zip เท่านั้น', 'danger')
        return redirect(url_for('subject_detail', subject_id=subject.id))

    mode = request.form.get('docs_import_mode', 'update').strip() or 'update'
    if mode not in ['update', 'keep', 'replace']:
        mode = 'update'

    target_level = detect_subject_level(subject)
    lessons, by_norm, by_period = build_lesson_match_map(subject)
    matched_lessons = set()
    updated_content = 0
    added_files = 0
    updated_worksheets = 0
    created_worksheets = 0
    skipped = []

    try:
        zdata = BytesIO(upload.read())
        with zipfile.ZipFile(zdata) as zf:
            entries = [n for n in zf.namelist() if not n.endswith('/') and not os.path.basename(n).startswith('.')]
            # กรองระดับชั้นถ้าเดาได้ เช่น รายวิชา ว21102/ม.1 จะอ่านเฉพาะโฟลเดอร์/ไฟล์ ม1
            if target_level:
                level_entries = []
                for n in entries:
                    meta = parse_lesson_doc_filename(n)
                    if not meta.level or meta.level == target_level:
                        # ถ้าใน ZIP รวมหลายชั้น ให้รับเฉพาะไฟล์ที่ชื่อ/โฟลเดอร์ตรงชั้น หรือไฟล์กลาง README/CSV ข้ามอยู่แล้ว
                        if meta.kind != 'other' and meta.level == target_level:
                            level_entries.append(n)
                entries = level_entries

            for name in entries:
                meta = parse_lesson_doc_filename(name)
                if meta.kind == 'other':
                    continue
                lesson = by_norm.get(normalize_thai_title(meta.title))
                if not lesson and meta.period:
                    lesson = by_period.get(meta.period)
                if not lesson:
                    skipped.append(meta.basename)
                    continue
                matched_lessons.add(lesson.id)
                raw = zf.read(name)

                if meta.kind == 'lesson_md':
                    if mode == 'keep' and lesson.content:
                        continue
                    try:
                        md = raw.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        md = raw.decode('utf-8', errors='ignore')
                    objective = extract_markdown_section(md, 'จุดประสงค์การเรียนรู้')
                    content_parts = []
                    for h in ['สาระสำคัญ', 'เนื้อหาบทเรียน', 'กิจกรรมการเรียนรู้', 'การวัดและประเมินผล']:
                        sec = extract_markdown_section(md, h)
                        if sec:
                            content_parts.append(f"## {h}\n{sec}")
                    content = '\n\n'.join(content_parts).strip() or md.strip()
                    if objective and mode != 'keep':
                        lesson.objective = objective
                    if mode in ['update', 'replace'] or not lesson.content:
                        lesson.content = content
                        updated_content += 1

                elif meta.kind == 'knowledge':
                    # ใบความรู้ให้เป็นสื่อประกอบบทเรียน ถ้า replace ให้ลบสื่อเดิมที่ชื่อคล้ายใบความรู้ก่อน
                    if mode == 'replace':
                        for old in LessonFile.query.filter_by(lesson_id=lesson.id).all():
                            if 'ใบความรู้' in (old.original_file_name or '') or 'ความรู้' in (old.original_file_name or ''):
                                safe_remove_upload_file(old.file_path)
                                db.session.delete(old)
                    if mode == 'keep' and LessonFile.query.filter_by(lesson_id=lesson.id, original_file_name=meta.basename).first():
                        continue
                    path, original = save_bytes_to_upload(raw, meta.basename, 'lesson_files')
                    db.session.add(LessonFile(lesson_id=lesson.id, file_path=path, original_file_name=original, file_type=detect_lesson_file_type(path)))
                    added_files += 1

                elif meta.kind == 'worksheet':
                    title = f"ใบงานครูรัก {subject.name} คาบ {meta.period or ''} - {lesson.title}".replace('  ', ' ').strip()
                    existing = Worksheet.query.filter(Worksheet.lesson_id == lesson.id, Worksheet.title.like('ใบงานครูรัก%')).first()
                    if existing and mode == 'keep':
                        continue
                    path, original = save_bytes_to_upload(raw, meta.basename, 'worksheets')
                    if existing:
                        if existing.file_path:
                            safe_remove_upload_file(existing.file_path)
                        existing.title = title
                        existing.description = f"ใบงานประกอบบทเรียน {lesson.title} สำหรับรายวิชา {subject.name}"
                        existing.worksheet_type = 'academic'
                        existing.file_path = path
                        existing.original_file_name = original
                        updated_worksheets += 1
                    else:
                        ws = Worksheet(
                            lesson_id=lesson.id,
                            title=title,
                            worksheet_type='academic',
                            description=f"ใบงานประกอบบทเรียน {lesson.title} สำหรับรายวิชา {subject.name}",
                            file_path=path,
                            original_file_name=original,
                        )
                        db.session.add(ws)
                        db.session.flush()
                        db.session.add(WorksheetQuestion(worksheet_id=ws.id, number=1, question_text='ทำใบงานตามไฟล์แนบ แล้วส่งตามที่ครูกำหนด', answer_type='file', max_score=10))
                        created_worksheets += 1

        db.session.commit()
        msg = f'นำเข้าเอกสารบทเรียนสำเร็จ: อัปเดตเนื้อหา {updated_content} บท, แนบใบความรู้ {added_files} ไฟล์, สร้างใบงาน {created_worksheets} ใบ, อัปเดตใบงาน {updated_worksheets} ใบ'
        if target_level:
            msg += f' (กรองเฉพาะ {target_level})'
        if skipped:
            msg += f' | ข้าม {len(skipped)} ไฟล์ที่จับคู่บทเรียนไม่เจอ'
        flash(msg, 'success')
    except zipfile.BadZipFile:
        db.session.rollback()
        flash('เปิดไฟล์ ZIP ไม่สำเร็จ กรุณาตรวจไฟล์อีกครั้ง', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'นำเข้าเอกสารไม่สำเร็จ: {e}', 'danger')
    return redirect(url_for('subject_detail', subject_id=subject.id))

def add_lesson_uploads(lesson):
    files = request.files.getlist('lesson_files')
    added = 0
    for f in files:
        if not f or not f.filename:
            continue
        try:
            fp, original = save_uploaded_file(f, 'lesson_files')
        except ValueError as e:
            flash(str(e), 'danger')
            continue
        if fp:
            db.session.add(LessonFile(lesson_id=lesson.id, file_path=fp, original_file_name=original, file_type=detect_lesson_file_type(fp)))
            added += 1
    return added


@app.route('/unit/<int:unit_id>/lessons')
@login_required
@role_required('teacher','admin')
def unit_lessons(unit_id):
    unit = Unit.query.get_or_404(unit_id)
    if not owns_unit(unit):
        return deny_redirect('subjects')
    lessons = Lesson.query.filter_by(unit_id=unit.id).order_by(Lesson.id.asc()).all()
    lesson_rows = []
    for lesson in lessons:
        worksheet_count = Worksheet.query.filter_by(lesson_id=lesson.id).count()
        quiz_count = Quiz.query.filter_by(lesson_id=lesson.id).count()
        file_count = LessonFile.query.filter_by(lesson_id=lesson.id).count()
        lesson_rows.append((lesson, worksheet_count, quiz_count, file_count))
    return render_template('unit_lessons.html', unit=unit, lesson_rows=lesson_rows)

@app.route('/unit/<int:unit_id>/lesson', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def lesson_create(unit_id):
    unit = Unit.query.get_or_404(unit_id)
    subject = unit.subject
    if not owns_subject(subject): flash('ไม่มีสิทธิ์','danger'); return redirect(url_for('subjects'))
    if request.method == 'POST':
        l = Lesson(unit_id=unit.id, title=request.form.get('title',''), objective=request.form.get('objective',''), content=request.form.get('content',''), media_url=request.form.get('media_url',''))
        db.session.add(l); db.session.flush()
        add_lesson_uploads(l)
        db.session.commit()
        flash('สร้างบทเรียนและแนบไฟล์แล้ว นักเรียนจะเห็นทันทีเมื่อเปิดบทเรียน/ตารางสอนที่ผูกบทเรียนนี้', 'success')
        return redirect(url_for('lesson_detail', lesson_id=l.id))
    return render_template('lesson_form.html', unit=unit, lesson=None, files=[])

@app.route('/lesson/<int:lesson_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def lesson_edit(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    if not owns_lesson(lesson): return deny_redirect('subjects')
    if request.method == 'POST':
        lesson.title = request.form.get('title','')
        lesson.objective = request.form.get('objective','')
        lesson.content = request.form.get('content','')
        lesson.media_url = request.form.get('media_url','')
        add_lesson_uploads(lesson)
        db.session.commit(); flash('แก้ไขบทเรียนแล้ว เนื้อหา/ไฟล์แนบจะแสดงในหน้าบทเรียนทันที', 'success')
        return redirect(url_for('lesson_detail', lesson_id=lesson.id))
    files = LessonFile.query.filter_by(lesson_id=lesson.id).order_by(LessonFile.created_at.desc()).all()
    return render_template('lesson_form.html', unit=lesson.unit, lesson=lesson, files=files)

@app.route('/lesson_file/<int:file_id>/delete', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def lesson_file_delete(file_id):
    lf = LessonFile.query.get_or_404(file_id)
    lesson = lf.lesson
    if not owns_lesson(lesson): return deny_redirect('subjects')
    safe_remove_upload_file(lf.file_path)
    db.session.delete(lf); db.session.commit(); flash('ลบไฟล์บทเรียนแล้ว', 'success')
    return redirect(url_for('lesson_edit', lesson_id=lesson.id))

@app.route('/lesson/<int:lesson_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def lesson_delete(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    if not owns_lesson(lesson): return deny_redirect('subjects')
    subject_id = lesson.unit.subject_id
    unit_id = lesson.unit_id
    title = lesson.title
    try:
        force_delete_lesson_data(lesson.id, delete_files=True)
        db.session.delete(lesson)
        db.session.commit()
        flash(f'ลบบทเรียน “{title}” พร้อมไฟล์ ใบงาน แบบทดสอบ และคำตอบที่ผูกอยู่แล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ลบบทเรียนไม่สำเร็จ: {e}', 'danger')
        return redirect(url_for('unit_lessons', unit_id=unit_id))
    return redirect(url_for('subject_detail', subject_id=subject_id))


@app.route('/lesson/<int:lesson_id>')
@login_required
def lesson_detail(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    if current_user.role == 'student' and not student_can_view_lesson(lesson, current_user):
        flash('ครูยังไม่ได้เปิดบทเรียนนี้ให้นักเรียนในห้องของคุณเห็น', 'warning')
        return redirect(url_for('student_dashboard'))
    worksheets = Worksheet.query.filter_by(lesson_id=lesson.id).all()
    quizzes = Quiz.query.filter_by(lesson_id=lesson.id).all()
    files = LessonFile.query.filter_by(lesson_id=lesson.id).order_by(LessonFile.created_at.desc()).all()
    lesson_status = get_or_create_lesson_status(lesson, current_user) if current_user.role == 'student' else None
    return render_template('lesson_detail.html', lesson=lesson, worksheets=worksheets, quizzes=quizzes, files=files, lesson_status=lesson_status, youtube_embed=youtube_embed_url(lesson.media_url))


@app.route('/subject/<int:subject_id>/worksheets')
@login_required
@role_required('teacher','admin')
def subject_worksheets(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    if not owns_subject(subject): return deny_redirect('subjects')
    items = []
    units = Unit.query.filter_by(subject_id=subject.id).all()
    for unit in units:
        lessons = Lesson.query.filter_by(unit_id=unit.id).all()
        for lesson in lessons:
            for worksheet in Worksheet.query.filter_by(lesson_id=lesson.id).all():
                items.append(SimpleNamespace(unit=unit, lesson=lesson, worksheet=worksheet))
    return render_template('subject_worksheets.html', subject=subject, items=items)

@app.route('/worksheet/<int:worksheet_id>/review')
@login_required
@role_required('teacher','admin')
def worksheet_review(worksheet_id):
    worksheet = Worksheet.query.get_or_404(worksheet_id)
    if not owns_lesson(worksheet.lesson): return deny_redirect('subjects')
    subject = worksheet.lesson.unit.subject
    rows = []
    # หา assignment ที่ผูกกับบทเรียนนี้ แล้วดึงนักเรียนของห้องนั้น ๆ
    assignments = Assignment.query.filter_by(lesson_id=worksheet.lesson_id).all()
    seen = set()
    for assignment in assignments:
        memberships = ClassroomStudent.query.filter_by(classroom_id=assignment.classroom_id).all()
        for m in memberships:
            key = (assignment.id, m.student_id)
            if key in seen:
                continue
            seen.add(key)
            status = AssignmentStatus.query.filter_by(assignment_id=assignment.id, student_id=m.student_id).first()
            answers = WorksheetAnswer.query.filter_by(assignment_id=assignment.id, student_id=m.student_id).join(WorksheetQuestion).filter(WorksheetQuestion.worksheet_id==worksheet.id).order_by(WorksheetQuestion.number).all()
            submitted = bool(status and status.worksheet_submitted) or bool(answers)
            score = sum((a.score or 0) for a in answers)
            rows.append(SimpleNamespace(student=m.student, classroom=assignment.classroom, status=status, answers=answers, submitted=submitted, score=score))
    # ถ้ายังไม่เคยสั่งงาน ให้แสดงข้อความว่างอย่างชัดเจน
    return render_template('worksheet_review.html', worksheet=worksheet, rows=rows)

@app.route('/lesson/<int:lesson_id>/worksheet', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def worksheet_create(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    if not owns_lesson(lesson): return deny_redirect('subjects')
    if request.method == 'POST':
        ws = Worksheet(lesson_id=lesson.id, title=request.form.get('title',''), worksheet_type=request.form.get('worksheet_type','academic'), description=request.form.get('description',''))
        try:
            fp, original = save_uploaded_file(request.files.get('worksheet_file'), 'worksheets')
            ws.file_path, ws.original_file_name = fp, original
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('worksheet_create', lesson_id=lesson.id))
        db.session.add(ws); db.session.flush()
        questions = request.form.get('questions','').splitlines()
        for idx, q in enumerate([x.strip() for x in questions if x.strip()], start=1):
            db.session.add(WorksheetQuestion(worksheet_id=ws.id, number=idx, question_text=q, answer_type='text', max_score=float(request.form.get('default_score', 1) or 1)))
        # petanque template: ระยะ 6.5/7.5/8.5/9.5 x สถานี 1-5
        if ws.worksheet_type == 'petanque_score' and not questions:
            distances = ['6.5 m','7.5 m','8.5 m','9.5 m']
            n=1
            for d in distances:
                for station in range(1,6):
                    db.session.add(WorksheetQuestion(worksheet_id=ws.id, number=n, question_text=f'{d} - สถานีที่ {station}', answer_type='score', max_score=5)); n+=1
        if ws.worksheet_type == 'upload' and not questions:
            db.session.add(WorksheetQuestion(worksheet_id=ws.id, number=1, question_text='แนบไฟล์คำตอบ', answer_type='file', max_score=float(request.form.get('default_score', 1) or 1)))
        db.session.commit(); flash('สร้างใบงานแล้ว', 'success'); return redirect(url_for('lesson_detail', lesson_id=lesson.id))
    return render_template('worksheet_form.html', lesson=lesson, worksheet=None)

@app.route('/worksheet/<int:worksheet_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def worksheet_edit(worksheet_id):
    ws = Worksheet.query.get_or_404(worksheet_id)
    if not owns_lesson(ws.lesson): return deny_redirect('subjects')
    if request.method == 'POST':
        old_type = ws.worksheet_type
        ws.title = request.form.get('title','')
        ws.worksheet_type = request.form.get('worksheet_type','academic')
        ws.description = request.form.get('description','')
        try:
            fp, original = save_uploaded_file(request.files.get('worksheet_file'), 'worksheets')
            if fp:
                ws.file_path, ws.original_file_name = fp, original
        except ValueError as e:
            flash(str(e), 'danger')
            return redirect(url_for('worksheet_edit', worksheet_id=ws.id))
        # ถ้าเปลี่ยนเป็นใบงานเปตอง/อัปโหลดและยังไม่มีข้อ ให้สร้างฟอร์มอัตโนมัติ
        if ws.worksheet_type == 'petanque_score' and old_type != 'petanque_score' and not WorksheetQuestion.query.filter_by(worksheet_id=ws.id).first():
            n=1
            for d in ['6.5 m','7.5 m','8.5 m','9.5 m']:
                for station in range(1,6):
                    db.session.add(WorksheetQuestion(worksheet_id=ws.id, number=n, question_text=f'{d} - สถานีที่ {station}', answer_type='score', max_score=5)); n+=1
        if ws.worksheet_type == 'upload' and not WorksheetQuestion.query.filter_by(worksheet_id=ws.id).first():
            db.session.add(WorksheetQuestion(worksheet_id=ws.id, number=1, question_text='แนบไฟล์คำตอบ', answer_type='file', max_score=1))
        db.session.commit(); flash('แก้ไขใบงานแล้ว', 'success')
        return redirect(url_for('lesson_detail', lesson_id=ws.lesson_id))
    return render_template('worksheet_form.html', lesson=ws.lesson, worksheet=ws)

@app.route('/worksheet/<int:worksheet_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def worksheet_delete(worksheet_id):
    ws = Worksheet.query.get_or_404(worksheet_id)
    if not owns_lesson(ws.lesson): return deny_redirect('subjects')
    lesson_id = ws.lesson_id
    try:
        # ต้องลบคำตอบนักเรียนและข้อคำถามก่อน ไม่อย่างนั้นฐานข้อมูลจะกันไม่ให้ลบใบงาน
        force_delete_worksheet_data(ws.id)
        db.session.delete(ws)
        db.session.commit()
        flash('ลบใบงานพร้อมข้อคำถามและคำตอบที่ผูกอยู่แล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ลบใบงานไม่สำเร็จ: {e}', 'danger')
    return redirect(url_for('lesson_detail', lesson_id=lesson_id))

@app.route('/worksheet_question/<int:question_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def worksheet_question_delete(question_id):
    q = WorksheetQuestion.query.get_or_404(question_id)
    if not owns_lesson(q.worksheet.lesson): return deny_redirect('subjects')
    lesson_id = q.worksheet.lesson_id
    try:
        WorksheetAnswer.query.filter_by(question_id=q.id).delete(synchronize_session=False)
        db.session.delete(q)
        db.session.commit()
        flash('ลบข้อคำถามใบงานและคำตอบที่ผูกอยู่แล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ลบข้อคำถามไม่สำเร็จ: {e}', 'danger')
    return redirect(url_for('lesson_detail', lesson_id=lesson_id))


@app.route('/lesson/<int:lesson_id>/quiz', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def quiz_create(lesson_id):
    lesson = Lesson.query.get_or_404(lesson_id)
    if request.method == 'POST':
        qz = Quiz(lesson_id=lesson.id, title=request.form.get('title',''), pass_percent=int(request.form.get('pass_percent',60)))
        db.session.add(qz); db.session.commit(); return redirect(url_for('quiz_questions', quiz_id=qz.id))
    return render_template('quiz_form.html', lesson=lesson)

@app.route('/quiz/<int:quiz_id>/questions', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def quiz_questions(quiz_id):
    quiz = Quiz.query.get_or_404(quiz_id)
    if request.method == 'POST':
        question_type = request.form.get('question_type','abcd')
        if question_type == 'abcd':
            labels = ['ก','ข','ค','ง']
            choices = '\n'.join([f"{lb}. {request.form.get('choice_'+lb, '').strip()}" for lb in labels if request.form.get('choice_'+lb, '').strip()])
            correct_answer = request.form.get('correct_letter','ก')
        else:
            choices = request.form.get('choices','')
            correct_answer = request.form.get('correct_answer','')
        db.session.add(QuizQuestion(quiz_id=quiz.id, question_text=request.form.get('question_text',''), question_type=question_type, choices=choices, correct_answer=correct_answer, score=float(request.form.get('score',1))))
        db.session.commit(); flash('เพิ่มข้อสอบแล้ว', 'success'); return redirect(url_for('quiz_questions', quiz_id=quiz.id))
    questions = QuizQuestion.query.filter_by(quiz_id=quiz.id).all()
    return render_template('quiz_questions.html', quiz=quiz, questions=questions)

@app.route('/quiz/<int:quiz_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def quiz_edit(quiz_id):
    quiz = Quiz.query.get_or_404(quiz_id)
    if not owns_lesson(quiz.lesson): return deny_redirect('subjects')
    if request.method == 'POST':
        quiz.title = request.form.get('title','')
        quiz.pass_percent = int(request.form.get('pass_percent',60))
        db.session.commit(); flash('แก้ไขแบบทดสอบแล้ว', 'success')
        return redirect(url_for('lesson_detail', lesson_id=quiz.lesson_id))
    return render_template('quiz_form.html', lesson=quiz.lesson, quiz=quiz)

@app.route('/quiz/<int:quiz_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def quiz_delete(quiz_id):
    quiz = Quiz.query.get_or_404(quiz_id)
    if not owns_lesson(quiz.lesson): return deny_redirect('subjects')
    lesson_id = quiz.lesson_id
    try:
        force_delete_quiz_data(quiz.id)
        db.session.delete(quiz)
        db.session.commit()
        flash('ลบแบบทดสอบพร้อมข้อสอบและคำตอบที่ผูกอยู่แล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ลบแบบทดสอบไม่สำเร็จ: {e}', 'danger')
    return redirect(url_for('lesson_detail', lesson_id=lesson_id))

@app.route('/quiz_question/<int:question_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def quiz_question_edit(question_id):
    q = QuizQuestion.query.get_or_404(question_id)
    if not owns_lesson(q.quiz.lesson): return deny_redirect('subjects')
    if request.method == 'POST':
        q.question_text = request.form.get('question_text','')
        q.question_type = request.form.get('question_type','abcd')
        if q.question_type == 'abcd':
            labels = ['ก','ข','ค','ง']
            q.choices = '\n'.join([f"{lb}. {request.form.get('choice_'+lb, '').strip()}" for lb in labels if request.form.get('choice_'+lb, '').strip()])
            q.correct_answer = request.form.get('correct_letter','ก')
        else:
            q.choices = request.form.get('choices','')
            q.correct_answer = request.form.get('correct_answer','')
        q.score = float(request.form.get('score',1))
        db.session.commit(); flash('แก้ไขข้อสอบแล้ว', 'success')
        return redirect(url_for('quiz_questions', quiz_id=q.quiz_id))
    return render_template('quiz_question_form.html', quiz=q.quiz, q=q)

@app.route('/quiz_question/<int:question_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def quiz_question_delete(question_id):
    q = QuizQuestion.query.get_or_404(question_id)
    if not owns_lesson(q.quiz.lesson): return deny_redirect('subjects')
    quiz_id = q.quiz_id
    try:
        QuizAnswer.query.filter_by(question_id=q.id).delete(synchronize_session=False)
        db.session.delete(q)
        db.session.commit()
        flash('ลบข้อสอบพร้อมคำตอบที่ผูกอยู่แล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ลบข้อสอบไม่สำเร็จ: {e}', 'danger')
    return redirect(url_for('quiz_questions', quiz_id=quiz_id))


@app.route('/assign', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def assign():
    subjects = teacher_filter(Subject).all()
    rooms = teacher_filter(Classroom).all()
    lessons = Lesson.query.join(Unit).join(Subject).filter(Subject.id.in_(teacher_subject_ids() or [-1])).all() if current_user.role=='teacher' else Lesson.query.all()
    if request.method == 'POST':
        a = Assignment(teacher_id=current_user.id, subject_id=int(request.form.get('subject_id','')), classroom_id=int(request.form.get('classroom_id','')), lesson_id=int(request.form.get('lesson_id','')), title=request.form.get('title',''), due_date=datetime.strptime(request.form.get('due_date',''),'%Y-%m-%d').date() if request.form.get('due_date') else None)
        db.session.add(a); db.session.flush()
        links = ClassroomStudent.query.filter_by(classroom_id=a.classroom_id).all()
        for link in links:
            db.session.add(AssignmentStatus(assignment_id=a.id, student_id=link.student_id))
        db.session.commit(); flash('สั่งงานให้นักเรียนในห้องแล้ว', 'success'); return redirect(url_for('teacher_dashboard'))
    return render_template('assign.html', subjects=subjects, rooms=rooms, lessons=lessons, assignment=None)

@app.route('/assignments')
@login_required
@role_required('teacher','admin')
def assignments():
    rows = Assignment.query.filter_by(teacher_id=current_user.id).order_by(Assignment.created_at.desc()).all() if current_user.role=='teacher' else Assignment.query.order_by(Assignment.created_at.desc()).all()
    return render_template('assignments.html', assignments=rows)

@app.route('/assignment/<int:assignment_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def assignment_edit(assignment_id):
    a = Assignment.query.get_or_404(assignment_id)
    if current_user.role != 'admin' and a.teacher_id != current_user.id: return deny_redirect('assignments')
    subjects = teacher_filter(Subject).all(); rooms = teacher_filter(Classroom).all()
    lessons = Lesson.query.join(Unit).join(Subject).filter(Subject.id.in_(teacher_subject_ids() or [-1])).all() if current_user.role=='teacher' else Lesson.query.all()
    if request.method == 'POST':
        a.title = request.form.get('title','')
        a.subject_id = int(request.form.get('subject_id',''))
        a.classroom_id = int(request.form.get('classroom_id',''))
        a.lesson_id = int(request.form.get('lesson_id',''))
        a.due_date = datetime.strptime(request.form.get('due_date',''),'%Y-%m-%d').date() if request.form.get('due_date') else None
        db.session.commit(); flash('แก้ไขงานที่สั่งแล้ว', 'success')
        return redirect(url_for('assignments'))
    return render_template('assign.html', subjects=subjects, rooms=rooms, lessons=lessons, assignment=a)

@app.route('/assignment/<int:assignment_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def assignment_delete(assignment_id):
    a = Assignment.query.get_or_404(assignment_id)
    if current_user.role != 'admin' and a.teacher_id != current_user.id: return deny_redirect('assignments')
    try:
        _delete_assignments(Assignment.query.filter_by(id=a.id))
        db.session.commit()
        flash('ลบงานที่สั่ง พร้อมสถานะ/คำตอบนักเรียนที่ผูกอยู่แล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ลบงานไม่สำเร็จ: {e}', 'danger')
    return redirect(url_for('assignments'))


@app.route('/student/assignment/<int:status_id>')
@login_required
@role_required('student')
def student_assignment(status_id):
    status = AssignmentStatus.query.get_or_404(status_id)
    if status.student_id != current_user.id: flash('ไม่มีสิทธิ์','danger'); return redirect(url_for('student_dashboard'))
    status.lesson_viewed = True
    update_assignment_complete(status)
    db.session.commit()
    lesson = status.assignment.lesson
    worksheets = Worksheet.query.filter_by(lesson_id=lesson.id).all()
    quiz = Quiz.query.filter_by(lesson_id=lesson.id).first()
    return render_template('student_assignment.html', status=status, lesson=lesson, worksheets=worksheets, quiz=quiz)

@app.route('/student/assignment/<int:status_id>/worksheet/<int:worksheet_id>', methods=['GET','POST'])
@login_required
@role_required('student')
def do_worksheet(status_id, worksheet_id):
    status = AssignmentStatus.query.get_or_404(status_id)
    worksheet = Worksheet.query.get_or_404(worksheet_id)
    questions = WorksheetQuestion.query.filter_by(worksheet_id=worksheet.id).order_by(WorksheetQuestion.number).all()
    if request.method == 'POST':
        for q in questions:
            ans = request.form.get(f'q_{q.id}','')
            row = WorksheetAnswer.query.filter_by(assignment_id=status.assignment_id, student_id=current_user.id, question_id=q.id).first()
            if not row:
                row = WorksheetAnswer(assignment_id=status.assignment_id, student_id=current_user.id, question_id=q.id)
                db.session.add(row)
            row.answer_text = ans
            file_obj = request.files.get(f'file_{q.id}')
            if file_obj and file_obj.filename:
                try:
                    fp, original = save_uploaded_file(file_obj, 'student_works')
                    row.file_path, row.original_file_name = fp, original
                except ValueError as e:
                    flash(str(e), 'danger')
                    return redirect(url_for('do_worksheet', status_id=status.id, worksheet_id=worksheet.id))
            try: row.score = float(ans) if q.answer_type=='score' and ans else 0
            except: row.score = 0
        status.worksheet_submitted = True; status.submitted_at = datetime.utcnow(); update_assignment_complete(status)
        db.session.commit(); flash('ส่งใบงานแล้ว', 'success'); return redirect(url_for('student_assignment', status_id=status.id))
    answers = WorksheetAnswer.query.filter_by(assignment_id=status.assignment_id, student_id=current_user.id).all()
    old = {a.question_id:a.answer_text for a in answers}
    old_files = {a.question_id:a for a in answers if a.file_path}
    return render_template('do_worksheet.html', status=status, worksheet=worksheet, questions=questions, old=old, old_files=old_files)

@app.route('/student/assignment/<int:status_id>/quiz/<int:quiz_id>', methods=['GET','POST'])
@login_required
@role_required('student')
def do_quiz(status_id, quiz_id):
    status = AssignmentStatus.query.get_or_404(status_id)
    quiz = Quiz.query.get_or_404(quiz_id)
    questions = QuizQuestion.query.filter_by(quiz_id=quiz.id).all()
    if request.method == 'POST':
        total = sum(q.score for q in questions) or 1
        got = 0
        QuizAnswer.query.filter_by(assignment_id=status.assignment_id, student_id=current_user.id).delete()
        for q in questions:
            ans = request.form.get(f'q_{q.id}','')
            if q.question_type == 'abcd':
                correct = ans.strip() == (q.correct_answer or '').strip()
            else:
                correct = ans.strip().lower() == (q.correct_answer or '').strip().lower()
            score = q.score if correct else 0
            got += score
            db.session.add(QuizAnswer(assignment_id=status.assignment_id, student_id=current_user.id, question_id=q.id, answer_text=ans, is_correct=correct, score=score))
        status.quiz_submitted = True
        status.quiz_score = round((got/total)*100,2)
        update_assignment_complete(status)
        db.session.commit(); flash(f'ทำแบบทดสอบแล้ว คะแนน {status.quiz_score}%', 'success'); return redirect(url_for('student_assignment', status_id=status.id))
    return render_template('do_quiz.html', status=status, quiz=quiz, questions=questions)



def default_period_times():
    return {
        1: ('08:40','09:30'),
        2: ('09:31','10:20'),
        3: ('10:31','11:20'),
        4: ('11:21','12:10'),
        5: ('13:00','13:50'),
        6: ('13:51','14:40'),
        7: ('14:41','15:30'),
        8: ('15:41','16:00'),
    }

def thai_weekday_to_int(value):
    txt = str(value or '').strip().lower()
    mapping = {
        '0':0,'จันทร์':0,'จ.':0,'mon':0,'monday':0,
        '1':1,'อังคาร':1,'อ.':1,'tue':1,'tuesday':1,
        '2':2,'พุธ':2,'พ.':2,'wed':2,'wednesday':2,
        '3':3,'พฤหัสบดี':3,'พฤหัส':3,'พฤ.':3,'thu':3,'thursday':3,
        '4':4,'ศุกร์':4,'ศ.':4,'fri':4,'friday':4,
        '5':5,'เสาร์':5,'ส.':5,'sat':5,'saturday':5,
        '6':6,'อาทิตย์':6,'อา.':6,'sun':6,'sunday':6,
    }
    return mapping.get(txt)

def get_or_create_classroom_for_teacher(name, teacher_id):
    name = str(name or '').strip()
    if not name:
        return None
    room = Classroom.query.filter_by(name=name).first()
    if not room:
        room = Classroom(name=name, teacher_id=None)
        db.session.add(room); db.session.flush()
    if teacher_id and not TeacherClassroom.query.filter_by(teacher_id=teacher_id, classroom_id=room.id).first():
        db.session.add(TeacherClassroom(teacher_id=teacher_id, classroom_id=room.id))
    return room

def link_subject_classroom(subject_id, classroom_id):
    """ผูกวิชากับห้องเรียนครั้งเดียว เพื่อให้วิชาเดียว/รหัสเดียวใช้ชุดใบงานเดียวกันได้หลายห้อง"""
    if subject_id and classroom_id and not SubjectClassroom.query.filter_by(subject_id=subject_id, classroom_id=classroom_id).first():
        db.session.add(SubjectClassroom(subject_id=subject_id, classroom_id=classroom_id))

def build_schedule_grid(rows):
    """รวมคาบที่ครูสอนเวลาเดียวกันและวิชาเดียวกัน ให้แสดงชื่อห้องหลายห้องในช่องเดียว"""
    grid = {}
    for row in rows:
        slot_key = f"{row.weekday}-{row.period_no}"
        groups = grid.setdefault(slot_key, [])
        group_key = (row.teacher_id, subject_identity_key(row.subject.name if row.subject else ''), row.start_time, row.end_time)
        group = next((g for g in groups if g['group_key'] == group_key), None)
        if not group:
            group = {
                'group_key': group_key,
                'subject': row.subject,
                'lesson': row.lesson,
                'classrooms': [],
                'places': [],
                'topics': [],
                'rows': []
            }
            groups.append(group)
        if row.classroom and row.classroom.name not in group['classrooms']:
            group['classrooms'].append(row.classroom.name)
        if row.room_name and row.room_name not in group['places']:
            group['places'].append(row.room_name)
        topic = row.topic or (row.lesson.title if row.lesson else '')
        if topic and topic not in group['topics']:
            group['topics'].append(topic)
        group['rows'].append(row)
    return grid


def user_can_open_schedule(row):
    if current_user.role == 'admin':
        return True
    if current_user.role == 'teacher':
        return row.teacher_id == current_user.id or owns_subject(row.subject) or owns_classroom(row.classroom)
    if current_user.role == 'student':
        return ClassroomStudent.query.filter_by(classroom_id=row.classroom_id, student_id=current_user.id).first() is not None
    return False

@app.route('/schedule/<int:schedule_id>/period')
@login_required
@role_required('teacher','admin','student')
def schedule_period(schedule_id):
    row = TeachingSchedule.query.get_or_404(schedule_id)
    if not user_can_open_schedule(row):
        return deny_redirect('schedule')
    selected_date = datetime.strptime(request.args.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    lesson_log = get_period_lesson_log(row, selected_date)
    lesson = lesson_for_schedule(row, selected_date, for_student=(current_user.role == 'student'))
    if current_user.role == 'student' and not lesson:
        return render_template(
            'schedule_period.html',
            row=row,
            selected_date=selected_date,
            selected_day_label=thai_date_label(selected_date),
            lesson=None,
            worksheets=[],
            quizzes=[],
            files_count=0,
            lesson_status=None,
            lesson_log=lesson_log,
            available_lessons=[]
        )
    worksheets = Worksheet.query.filter_by(lesson_id=lesson.id).all() if lesson else []
    quizzes = Quiz.query.filter_by(lesson_id=lesson.id).all() if lesson else []
    files_count = LessonFile.query.filter_by(lesson_id=lesson.id).count() if lesson else 0
    lesson_status = get_or_create_lesson_status(lesson, current_user) if (lesson and current_user.role == 'student') else None
    return render_template(
        'schedule_period.html',
        row=row,
        selected_date=selected_date,
        selected_day_label=thai_date_label(selected_date),
        lesson=lesson,
        worksheets=worksheets,
        quizzes=quizzes,
        files_count=files_count,
        lesson_status=lesson_status,
        lesson_log=lesson_log,
        available_lessons=[
            SimpleNamespace(
                lesson=l,
                worksheet_count=Worksheet.query.filter_by(lesson_id=l.id).count(),
                quiz_count=Quiz.query.filter_by(lesson_id=l.id).count(),
                file_count=LessonFile.query.filter_by(lesson_id=l.id).count(),
            )
            for l in (ordered_lessons_for_subject(row.subject_id) if current_user.role in ['teacher','admin'] else [])
        ]
    )


@app.route('/schedule/<int:schedule_id>/period/publish', methods=['POST'])
@login_required
@role_required('teacher','admin')
def schedule_period_publish(schedule_id):
    row = TeachingSchedule.query.get_or_404(schedule_id)
    if not user_can_open_schedule(row):
        return deny_redirect('schedule')
    selected_date = datetime.strptime(request.form.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    lesson_id = request.form.get('lesson_id', type=int)
    note = (request.form.get('note') or '').strip()
    lesson = Lesson.query.get_or_404(lesson_id)
    if lesson.unit.subject_id != row.subject_id:
        flash('บทเรียนนี้ไม่ได้อยู่ในรายวิชาของคาบนี้', 'danger')
        return redirect(url_for('schedule_period', schedule_id=row.id, date=selected_date.isoformat()))
    log = get_period_lesson_log(row, selected_date)
    if not log:
        log = PeriodLessonLog(schedule_id=row.id, taught_date=selected_date)
        db.session.add(log)
    log.lesson_id = lesson.id
    log.is_published = True
    log.taught_at = datetime.utcnow()
    log.taught_by_id = current_user.id
    log.note = note
    created = create_lesson_assignment_for_classroom(lesson, row)
    db.session.commit()
    flash(f'เปิดบทเรียน “{lesson.title}” ให้นักเรียนเห็นแล้ว และบันทึกว่าคาบนี้สอนไปแล้ว', 'success')
    return redirect(url_for('schedule_period', schedule_id=row.id, date=selected_date.isoformat()))


@app.route('/schedule/<int:schedule_id>/period/unpublish', methods=['POST'])
@login_required
@role_required('teacher','admin')
def schedule_period_unpublish(schedule_id):
    row = TeachingSchedule.query.get_or_404(schedule_id)
    if not user_can_open_schedule(row):
        return deny_redirect('schedule')
    selected_date = datetime.strptime(request.form.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    log = get_period_lesson_log(row, selected_date)
    if log:
        log.is_published = False
        db.session.commit()
        flash('ซ่อนบทเรียนของคาบนี้จากนักเรียนแล้ว', 'success')
    else:
        flash('ยังไม่มีบทเรียนที่เปิดในคาบนี้', 'warning')
    return redirect(url_for('schedule_period', schedule_id=row.id, date=selected_date.isoformat()))


@app.route('/schedule', methods=['GET','POST'])
@login_required
@role_required('teacher','admin','student')
def schedule():
    teachers = active_teachers_query().order_by(User.full_name).all()
    day_names = ['จันทร์','อังคาร','พุธ','พฤหัสบดี','ศุกร์','เสาร์','อาทิตย์']
    selected_teacher_id = request.args.get('teacher_id', type=int)

    if current_user.role == 'admin':
        active_teacher_id = selected_teacher_id or (teachers[0].id if teachers else None)
        subjects = Subject.query.filter(Subject.id.in_({x.subject_id for x in TeacherSubject.query.filter_by(teacher_id=active_teacher_id).all()} or {-1})).order_by(Subject.name).all() if active_teacher_id else []
        rooms = Classroom.query.filter(Classroom.id.in_({x.classroom_id for x in TeacherClassroom.query.filter_by(teacher_id=active_teacher_id).all()} or {-1})).order_by(Classroom.name).all() if active_teacher_id else []
        schedules_query = TeachingSchedule.query
        if active_teacher_id:
            schedules_query = schedules_query.filter_by(teacher_id=active_teacher_id)
        schedules = schedules_query.order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all()
        teacher_title = User.query.get(active_teacher_id).full_name if active_teacher_id else 'ยังไม่เลือกครู'
    elif current_user.role == 'teacher':
        active_teacher_id = current_user.id
        subjects = teacher_filter(Subject).order_by(Subject.name).all()
        rooms = teacher_filter(Classroom).order_by(Classroom.name).all()
        schedules = TeachingSchedule.query.filter_by(teacher_id=current_user.id).order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all()
        teacher_title = current_user.full_name
    else:
        # นักเรียนเห็นเฉพาะตารางสอนของห้องที่ตนเองอยู่
        student_room_ids = [x.classroom_id for x in ClassroomStudent.query.filter_by(student_id=current_user.id).all()]
        schedules = TeachingSchedule.query.filter(TeachingSchedule.classroom_id.in_(student_room_ids or [-1])).order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all()
        subjects = []
        rooms = []
        active_teacher_id = None
        teacher_title = current_user.full_name

    lessons = Lesson.query.join(Unit).filter(Unit.subject_id.in_([x.id for x in subjects] or [-1])).order_by(Lesson.title).all() if current_user.role != 'student' else []
    edit_row = None
    if request.method == 'POST':
        if current_user.role == 'student':
            return deny_redirect('schedule')
        post_teacher_id = int(request.form.get('teacher_id') or current_user.id)
        if current_user.role != 'admin':
            post_teacher_id = current_user.id
        subject_id = int(request.form.get('subject_id',''))
        classroom_id = int(request.form.get('classroom_id',''))
        # ตารางสอนต้องใช้เฉพาะวิชา/ห้องที่กำหนดให้ครูแล้วเท่านั้น
        if not TeacherSubject.query.filter_by(teacher_id=post_teacher_id, subject_id=subject_id).first():
            flash('ยังไม่ได้มอบหมายวิชานี้ให้ครูคนนี้ กรุณาไปหน้ามอบหมายครูก่อน', 'danger')
            return redirect(url_for('schedule', teacher_id=post_teacher_id))
        if not TeacherClassroom.query.filter_by(teacher_id=post_teacher_id, classroom_id=classroom_id).first():
            flash('ยังไม่ได้มอบหมายห้องนี้ให้ครูคนนี้ กรุณาไปหน้ามอบหมายครูก่อน', 'danger')
            return redirect(url_for('schedule', teacher_id=post_teacher_id))
        period_no = int(request.form.get('period_no',''))
        defaults = default_period_times()
        st, en = defaults.get(period_no, ('08:00','09:00'))
        db.session.add(TeachingSchedule(
            teacher_id=post_teacher_id,
            subject_id=subject_id,
            classroom_id=classroom_id,
            weekday=int(request.form.get('weekday','')),
            period_no=period_no,
            start_time=request.form.get('start_time') or st,
            end_time=request.form.get('end_time') or en,
            room_name=request.form.get('room_name',''),
            topic=request.form.get('topic',''),
            lesson_id=int(request.form.get('lesson_id','')) if request.form.get('lesson_id') else None
        ))
        db.session.commit(); return redirect(url_for('schedule', teacher_id=post_teacher_id))

    periods = list(range(1, 9))
    schedule_slots = [
        {'type':'period','no':1,'time':'08:40-09:30'},
        {'type':'period','no':2,'time':'09:31-10:20'},
        {'type':'break','label':'พัก 10 นาที','time':'10:21-10:30'},
        {'type':'period','no':3,'time':'10:31-11:20'},
        {'type':'period','no':4,'time':'11:21-12:10'},
        {'type':'break','label':'พักเที่ยง','time':'12:10-13:00'},
        {'type':'period','no':5,'time':'13:00-13:50'},
        {'type':'period','no':6,'time':'13:51-14:40'},
        {'type':'break','label':'พัก 10 นาที','time':'14:41-14:50'},
        {'type':'period','no':7,'time':'14:41-15:30'},
        {'type':'period','no':8,'time':'15:41-16:00'},
    ]
    today = local_today()
    current_lesson_logs = {row.id: get_period_lesson_log(row, today) for row in schedules}
    schedule_grid = build_schedule_grid(schedules)
    day_pairs = list(enumerate(day_names[:5]))
    schedules_by_teacher = {}
    if current_user.role == 'admin' and not selected_teacher_id:
        all_rows = TeachingSchedule.query.order_by(TeachingSchedule.teacher_id, TeachingSchedule.weekday, TeachingSchedule.period_no).all()
        for row in all_rows:
            schedules_by_teacher.setdefault(row.teacher_id, []).append(row)
    return render_template('schedule.html', schedules=schedules, subjects=subjects, rooms=rooms, lessons=lessons, edit_row=edit_row, day_names=day_names, periods=periods, schedule_grid=schedule_grid, day_pairs=day_pairs, teachers=teachers, selected_teacher_id=selected_teacher_id, active_teacher_id=active_teacher_id, teacher_title=teacher_title, schedules_by_teacher=schedules_by_teacher, today_date=today.isoformat(), schedule_slots=schedule_slots, current_lesson_logs=current_lesson_logs)

@app.route('/schedule/import', methods=['POST'])
@login_required
@role_required('teacher','admin')
def schedule_import():
    """นำเข้าตารางสอนจาก Excel
    รองรับ 2 แบบ
    1) แบบเดิม: เลือกครูจากฟอร์ม แล้วไฟล์มี วัน, คาบ, รายวิชา, ห้อง, สถานที่, หัวข้อ
    2) แบบรวมทั้งโรงเรียน: ไฟล์มี ครู/teacher/teacher_name เพิ่มมาในแต่ละแถว
       ถ้าครูยังไม่มีในระบบ จะสร้าง user อัตโนมัติ รหัสผ่านเริ่มต้น 1234
    ช่องที่รองรับเพิ่มเติม: รหัสวิชา/code, ชื่อวิชา/subject_name, ชั้นห้อง/classroom, ห้องเรียน/room_name
    """
    f = request.files.get('excel')
    if not f:
        flash('กรุณาเลือกไฟล์ Excel ตารางสอน', 'danger')
        return redirect(url_for('schedule'))
    path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(f.filename))
    f.save(path)
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash('ไฟล์ Excel ไม่มีข้อมูล', 'danger')
        return redirect(url_for('schedule'))

    headers = [str(x or '').strip().lower() for x in rows[0]]
    def idx(*names, default=None):
        for name in names:
            n = name.lower()
            if n in headers:
                return headers.index(n)
        return default

    i_teacher = idx('teacher','teacher_name','ครู','ชื่อครู','ผู้สอน', default=None)
    i_day = idx('day','วัน','weekday', default=0)
    i_period = idx('period','คาบ','period_no', default=1)
    i_code = idx('code','subject_code','รหัสวิชา','รหัส', default=None)
    i_subject = idx('subject','subject_name','รายวิชา','วิชา','ชื่อวิชา', default=2)
    i_classroom = idx('classroom','ห้อง','ชั้น/ห้อง','ชั้นเรียน', default=3)
    i_room = idx('room','room_name','สถานที่','ห้องเรียน', default=4)
    i_topic = idx('topic','หัวข้อ','เรื่อง', default=5)

    def cell(row, i, default=''):
        if i is None or i >= len(row):
            return default
        return str(row[i] or '').strip()

    def username_from_name(name):
        import re
        base = re.sub(r'[^a-zA-Z0-9ก-๙]+', '', str(name).replace('ครู','').strip())
        if not base:
            base = 'teacher'
        username = base[:40]
        if not User.query.filter_by(username=username).first():
            return username
        n = 2
        while User.query.filter_by(username=f'{username}{n}').first():
            n += 1
        return f'{username}{n}'

    def get_or_create_teacher_by_name(name):
        name = str(name or '').strip()
        if not name:
            return None
        t = User.query.filter_by(full_name=name, role='teacher').first()
        if not t:
            t = User.query.filter_by(full_name=name).first()
        if not t:
            t = User(username=username_from_name(name), full_name=name, role='teacher', must_change_password=True)
            t.set_password('1234')
            db.session.add(t)
            db.session.flush()
        return t

    defaults = default_period_times()
    default_teacher_id = int(request.form.get('teacher_id') or current_user.id)
    if current_user.role != 'admin':
        default_teacher_id = current_user.id

    replace_mode = bool(request.form.get('replace_all'))
    if replace_mode:
        if current_user.role == 'admin':
            TeachingSchedule.query.delete()
        else:
            TeachingSchedule.query.filter_by(teacher_id=current_user.id).delete()
        db.session.flush()

    created = updated = skipped = made_teachers = 0
    teacher_names_created = set()
    for row in rows[1:]:
        if not row:
            skipped += 1
            continue
        weekday = thai_weekday_to_int(cell(row, i_day))
        try:
            period_no = int(float(cell(row, i_period)))
        except Exception:
            period_no = None
        classroom_name = cell(row, i_classroom)
        subject_code = cell(row, i_code)
        subject_name = cell(row, i_subject)
        subject_full = ' '.join([x for x in [subject_code, subject_name] if x]).strip()
        teacher_name = cell(row, i_teacher)

        if current_user.role == 'admin' and teacher_name:
            before = User.query.count()
            teacher = get_or_create_teacher_by_name(teacher_name)
            teacher_id = teacher.id
            if User.query.count() > before and teacher.full_name not in teacher_names_created:
                made_teachers += 1
                teacher_names_created.add(teacher.full_name)
        else:
            teacher_id = default_teacher_id

        if weekday is None or not period_no or not subject_full or not classroom_name:
            skipped += 1
            continue

        sub = get_or_create_subject_for_teacher(subject_full, teacher_id, subject_code, subject_name)
        room = get_or_create_classroom_for_teacher(classroom_name, teacher_id)
        link_subject_classroom(sub.id, room.id)
        st, en = defaults.get(period_no, ('08:00','09:00'))
        room_name = cell(row, i_room)
        topic = cell(row, i_topic)
        old = TeachingSchedule.query.filter_by(teacher_id=teacher_id, weekday=weekday, period_no=period_no, classroom_id=room.id).first()
        if old:
            old.subject_id = sub.id
            old.start_time = st
            old.end_time = en
            old.room_name = room_name
            old.topic = topic
            updated += 1
        else:
            db.session.add(TeachingSchedule(teacher_id=teacher_id, subject_id=sub.id, classroom_id=room.id, weekday=weekday, period_no=period_no, start_time=st, end_time=en, room_name=room_name, topic=topic))
            created += 1
    merged_subjects = merge_duplicate_subjects_for_teacher()
    db.session.commit()
    extra = f', ยุบรายวิชาซ้ำ {merged_subjects} รายการ' if merged_subjects else ''
    flash(f'นำเข้าตารางสอนแล้ว: เพิ่ม {created} คาบ, อัปเดต {updated} คาบ, สร้างครูใหม่ {made_teachers} คน, ข้าม {skipped} แถว{extra}', 'success')
    return redirect(url_for('schedule'))

@app.route('/schedule/<int:schedule_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def schedule_edit(schedule_id):
    row = TeachingSchedule.query.get_or_404(schedule_id)
    if current_user.role != 'admin' and row.teacher_id != current_user.id: return deny_redirect('schedule')
    subjects = teacher_filter(Subject).all(); rooms = teacher_filter(Classroom).all()
    lessons = Lesson.query.join(Unit).join(Subject).filter(Subject.id.in_(teacher_subject_ids() or [-1])).all() if current_user.role=='teacher' else Lesson.query.all()
    if request.method == 'POST':
        row.weekday=int(request.form.get('weekday','')); row.period_no=int(request.form.get('period_no',''))
        row.start_time=request.form.get('start_time',''); row.end_time=request.form.get('end_time','')
        row.subject_id=int(request.form.get('subject_id','')); row.classroom_id=int(request.form.get('classroom_id',''))
        row.lesson_id=int(request.form.get('lesson_id','')) if request.form.get('lesson_id') else None
        row.room_name=request.form.get('room_name',''); row.topic=request.form.get('topic','')
        db.session.commit(); flash('แก้ไขตารางสอนแล้ว', 'success')
        return redirect(url_for('schedule'))
    schedules = TeachingSchedule.query.filter_by(teacher_id=current_user.id).order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all() if current_user.role=='teacher' else TeachingSchedule.query.order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all()
    day_names = ['จันทร์','อังคาร','พุธ','พฤหัสบดี','ศุกร์','เสาร์','อาทิตย์']
    periods = sorted({x.period_no for x in schedules}) or list(range(1, 9))
    period_times = {x.period_no: f"{x.start_time}-{x.end_time}" for x in schedules}
    schedule_grid = build_schedule_grid(schedules)
    schedule_slots = [
        {'type':'period','no':1,'time':'08:40-09:30'}, {'type':'period','no':2,'time':'09:31-10:20'},
        {'type':'break','label':'พัก 10 นาที','time':'10:21-10:30'}, {'type':'period','no':3,'time':'10:31-11:20'},
        {'type':'period','no':4,'time':'11:21-12:10'}, {'type':'break','label':'พักเที่ยง','time':'12:10-13:00'},
        {'type':'period','no':5,'time':'13:00-13:50'}, {'type':'period','no':6,'time':'13:51-14:40'},
        {'type':'break','label':'พัก 10 นาที','time':'14:41-14:50'}, {'type':'period','no':7,'time':'14:41-15:30'}, {'type':'period','no':8,'time':'15:41-16:00'}]
    day_pairs = list(enumerate(day_names[:5]))
    return render_template('schedule.html', schedules=schedules, subjects=subjects, rooms=rooms, lessons=lessons, edit_row=row, day_names=day_names, periods=periods, period_times=period_times, schedule_grid=schedule_grid, day_pairs=day_pairs, today_date=local_today().isoformat(), teachers=[], selected_teacher_id=None, active_teacher_id=None, teacher_title=current_user.full_name, schedules_by_teacher={}, schedule_slots=schedule_slots)

@app.route('/schedule/<int:schedule_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def schedule_delete(schedule_id):
    row = TeachingSchedule.query.get_or_404(schedule_id)
    if current_user.role != 'admin' and row.teacher_id != current_user.id: return deny_redirect('schedule')
    try:
        PeriodLessonLog.query.filter_by(schedule_id=row.id).delete(synchronize_session=False)
        ClassworkScoreItem.query.filter_by(schedule_id=row.id).update({ClassworkScoreItem.schedule_id: None}, synchronize_session=False)
        db.session.delete(row)
        db.session.commit()
        flash('ลบคาบสอนแล้ว และตัดลิงก์ข้อมูลคาบที่เกี่ยวข้องแล้ว', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'ลบคาบสอนไม่สำเร็จ: {e}', 'danger')
    return redirect(url_for('schedule'))



def academic_calendar_1_2569_events():
    """ปฏิทินการศึกษา ภาคเรียนที่ 1/2569 ตามภาพตัวอย่างที่ผู้ใช้ให้ไว้
    ใช้ปี ค.ศ. 2026 ในระบบ แต่แสดงผลเป็น พ.ศ. 2569 ในหน้าเว็บได้
    """
    items = [
        # พฤษภาคม 2569
        ('2026-05-01','วันแรงงานแห่งชาติ','วันหยุดราชการ','พฤษภาคม 2569'),
        ('2026-05-04','วันฉัตรมงคล','วันหยุดราชการ','พฤษภาคม 2569'),
        ('2026-05-11','เปิดภาคเรียนที่ 1/2569','กิจกรรมโรงเรียน','พฤษภาคม 2569'),
        ('2026-05-13','วันพืชมงคล','วันหยุดราชการ','พฤษภาคม 2569'),
        ('2026-05-15','ส่งโครงการสอนและแผนการจัดการเรียน','กิจกรรมโรงเรียน','พฤษภาคม 2569'),
        ('2026-05-18','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-19','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-20','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-21','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-22','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-25','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-26','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-27','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-28','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-29','ยื่นคำร้องแก้ 0 ร มส.','กิจกรรมโรงเรียน','18-29 พฤษภาคม 2569'),
        ('2026-05-31','วันวิสาขบูชา','วันหยุดราชการ','พฤษภาคม 2569'),
        # มิถุนายน 2569
        ('2026-06-01','วันหยุดชดเชยวันวิสาขบูชา','วันหยุดราชการ','มิถุนายน 2569'),
        ('2026-06-02','ประกาศผลการแก้ 0 ร มส.','กิจกรรมโรงเรียน','มิถุนายน 2569'),
        ('2026-06-03','วันเฉลิมพระชนมพรรษาสมเด็จพระราชินี','วันหยุดราชการ','มิถุนายน 2569'),
        ('2026-06-08','คลินิก 0 ร มส.','กิจกรรมโรงเรียน','มิถุนายน 2569'),
        ('2026-06-10','ส่ง ปถ.05 ครั้งที่ 1 กำหนดคะแนน','กิจกรรมโรงเรียน','มิถุนายน 2569'),
        ('2026-06-15','PLC รอบที่ 1 กำหนดปัญหา','กิจกรรมโรงเรียน','15-19 มิถุนายน 2569'),
        ('2026-06-16','PLC รอบที่ 1 กำหนดปัญหา','กิจกรรมโรงเรียน','15-19 มิถุนายน 2569'),
        ('2026-06-17','PLC รอบที่ 1 กำหนดปัญหา','กิจกรรมโรงเรียน','15-19 มิถุนายน 2569'),
        ('2026-06-18','PLC รอบที่ 1 กำหนดปัญหา','กิจกรรมโรงเรียน','15-19 มิถุนายน 2569'),
        ('2026-06-19','PLC รอบที่ 1 กำหนดปัญหา','กิจกรรมโรงเรียน','15-19 มิถุนายน 2569'),
        ('2026-06-26','วันสุนทรภู่','กิจกรรมโรงเรียน','หยุดชดเชย 2 กรกฎาคม'),
        # กรกฎาคม 2569
        ('2026-07-01','วันคล้ายวันสถาปนาคณะลูกเสือแห่งชาติ','กิจกรรมโรงเรียน','กรกฎาคม 2569'),
        ('2026-07-06','สำรวจติดตามสอบกลางภาค','กิจกรรมโรงเรียน','กรกฎาคม 2569'),
        ('2026-07-16','สอบกลางภาค','กิจกรรมโรงเรียน','16-17 กรกฎาคม 2569'),
        ('2026-07-17','สอบกลางภาค','กิจกรรมโรงเรียน','16-17 กรกฎาคม 2569'),
        ('2026-07-20','PLC รอบที่ 2 สะท้อนแผน','กิจกรรมโรงเรียน','20-24 กรกฎาคม 2569'),
        ('2026-07-21','PLC รอบที่ 2 สะท้อนแผน','กิจกรรมโรงเรียน','20-24 กรกฎาคม 2569'),
        ('2026-07-22','PLC รอบที่ 2 สะท้อนแผน','กิจกรรมโรงเรียน','20-24 กรกฎาคม 2569'),
        ('2026-07-23','PLC รอบที่ 2 สะท้อนแผน','กิจกรรมโรงเรียน','20-24 กรกฎาคม 2569'),
        ('2026-07-24','ส่ง ปถ.05 ครั้งที่ 2 + คะแนนกลางภาค','กิจกรรมโรงเรียน','กรกฎาคม 2569'),
        ('2026-07-28','วันเฉลิมพระชนมพรรษาพระบาทสมเด็จพระเจ้าอยู่หัว','วันหยุดราชการ','กรกฎาคม 2569'),
        ('2026-07-29','วันอาสาฬหบูชา','วันหยุดราชการ','กรกฎาคม 2569'),
        ('2026-07-30','วันเข้าพรรษา','วันหยุดราชการ','กรกฎาคม 2569'),
        # สิงหาคม 2569
        ('2026-08-03','PLC รอบที่ 3 เปิดชั้นเรียน','กิจกรรมโรงเรียน','3-7 สิงหาคม 2569'),
        ('2026-08-04','PLC รอบที่ 3 เปิดชั้นเรียน','กิจกรรมโรงเรียน','3-7 สิงหาคม 2569'),
        ('2026-08-05','PLC รอบที่ 3 เปิดชั้นเรียน','กิจกรรมโรงเรียน','3-7 สิงหาคม 2569'),
        ('2026-08-06','PLC รอบที่ 3 เปิดชั้นเรียน','กิจกรรมโรงเรียน','3-7 สิงหาคม 2569'),
        ('2026-08-07','PLC รอบที่ 3 เปิดชั้นเรียน','กิจกรรมโรงเรียน','3-7 สิงหาคม 2569'),
        ('2026-08-12','วันแม่แห่งชาติ','วันหยุดราชการ','สิงหาคม 2569'),
        ('2026-08-18','สัปดาห์วันวิทยาศาสตร์','กิจกรรมโรงเรียน','สิงหาคม 2569'),
        # กันยายน 2569
        ('2026-09-07','สำรวจข้อสอบปลายภาค','กิจกรรมโรงเรียน','กันยายน 2569'),
        ('2026-09-08','ส่งก่อนปลายเรียน 60% 80%','กิจกรรมโรงเรียน','กันยายน 2569'),
        ('2026-09-09','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-10','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-11','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-14','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-15','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-16','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-17','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-18','แก้ มส.','กิจกรรมโรงเรียน','9-18 กันยายน 2569'),
        ('2026-09-21','ขออนุมัติผล มผ.','กิจกรรมโรงเรียน','กันยายน 2569'),
        ('2026-09-23','สอบปลายภาค','กิจกรรมโรงเรียน','23-25 กันยายน 2569'),
        ('2026-09-24','สอบปลายภาค','กิจกรรมโรงเรียน','23-25 กันยายน 2569'),
        ('2026-09-25','สอบปลายภาค','กิจกรรมโรงเรียน','23-25 กันยายน 2569'),
        # ตุลาคม 2569
        ('2026-10-01','ปิดภาคเรียนที่ 1/2569','กิจกรรมโรงเรียน','ตุลาคม 2569'),
        ('2026-10-09','ส่งผลการเรียน ปถ.05 + bookmark','กิจกรรมโรงเรียน','ตุลาคม 2569'),
        ('2026-10-13','วันคล้ายวันสวรรคต ร.9','วันหยุดราชการ','ตุลาคม 2569'),
        ('2026-10-23','วันปิยมหาราช','วันหยุดราชการ','ตุลาคม 2569'),
        ('2026-10-26','วันออกพรรษา','วันหยุดราชการ','ตุลาคม 2569'),
    ]
    return [(datetime.strptime(d, '%Y-%m-%d').date(), title, typ, note) for d, title, typ, note in items]

def seed_academic_calendar_1_2569(teacher_id=None):
    created = 0
    for event_date, title, event_type, note in academic_calendar_1_2569_events():
        exists = CalendarEvent.query.filter_by(teacher_id=teacher_id, event_date=event_date, title=title, event_type=event_type).first()
        if not exists:
            db.session.add(CalendarEvent(teacher_id=teacher_id, event_date=event_date, title=title, event_type=event_type, note=note))
            created += 1
    db.session.commit()
    return created

def working_day_summary(start=None, end=None, teacher_id=None, semester=None):
    semester = semester or get_active_semester()
    if semester and not start and not end:
        start, end = semester.start_date, semester.end_date
    start = start or date(2026,5,11)
    end = end or date(2026,10,1)
    holidays = set()
    q = CalendarEvent.query.filter(CalendarEvent.event_date>=start, CalendarEvent.event_date<=end, CalendarEvent.event_type.in_(['วันหยุดราชการ','วันหยุดโรงเรียน']))
    if teacher_id:
        q = q.filter((CalendarEvent.teacher_id==teacher_id) | (CalendarEvent.teacher_id==None))
    for e in q.all():
        holidays.add(e.event_date)
    summary = {}
    cur = start
    while cur <= end:
        key = (cur.year, cur.month)
        if key not in summary:
            summary[key] = {'month': f"{thai_month_name(cur.month)} {cur.year+543}", 'working':0, 'holiday':0, 'weekend':0}
        if cur.weekday() >= 5:
            summary[key]['weekend'] += 1
        elif cur in holidays:
            summary[key]['holiday'] += 1
        else:
            summary[key]['working'] += 1
        cur += timedelta(days=1)
    return list(summary.values())


@app.route('/semesters', methods=['GET','POST'])
@login_required
@role_required('admin')
def semesters():
    if request.method == 'POST':
        sem = Semester(
            name=request.form.get('name',''),
            academic_year=request.form.get('academic_year','2569'),
            term_no=request.form.get('term_no','1'),
            start_date=datetime.strptime(request.form.get('start_date',''),'%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form.get('end_date',''),'%Y-%m-%d').date(),
            is_active=bool(request.form.get('is_active'))
        )
        if sem.is_active:
            Semester.query.update({Semester.is_active: False})
        db.session.add(sem); db.session.commit(); flash('เพิ่มภาคเรียนแล้ว', 'success')
        return redirect(url_for('semesters'))
    return render_template('semesters.html', semesters=Semester.query.order_by(Semester.start_date.desc()).all(), edit_row=None)

@app.route('/semesters/<int:semester_id>/edit', methods=['GET','POST'])
@login_required
@role_required('admin')
def semester_edit(semester_id):
    sem = Semester.query.get_or_404(semester_id)
    if request.method == 'POST':
        sem.name=request.form.get('name',''); sem.academic_year=request.form.get('academic_year','')
        sem.term_no=request.form.get('term_no','')
        sem.start_date=datetime.strptime(request.form.get('start_date',''),'%Y-%m-%d').date()
        sem.end_date=datetime.strptime(request.form.get('end_date',''),'%Y-%m-%d').date()
        sem.is_active=bool(request.form.get('is_active'))
        if sem.is_active:
            Semester.query.filter(Semester.id!=sem.id).update({Semester.is_active: False})
        db.session.commit(); flash('แก้ไขภาคเรียนแล้ว', 'success')
        return redirect(url_for('semesters'))
    return render_template('semesters.html', semesters=Semester.query.order_by(Semester.start_date.desc()).all(), edit_row=sem)

@app.route('/semesters/<int:semester_id>/active', methods=['POST'])
@login_required
@role_required('admin')
def semester_active(semester_id):
    Semester.query.update({Semester.is_active: False})
    sem = Semester.query.get_or_404(semester_id); sem.is_active=True
    db.session.commit(); flash(f'เปลี่ยนเป็น {sem.name} แล้ว', 'success')
    return redirect(url_for('semesters'))

@app.route('/semesters/<int:semester_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def semester_delete(semester_id):
    sem = Semester.query.get_or_404(semester_id)
    db.session.delete(sem); db.session.commit(); flash('ลบภาคเรียนแล้ว', 'success')
    return redirect(url_for('semesters'))

@app.route('/calendar', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def calendar_events():
    semesters = Semester.query.order_by(Semester.start_date.desc()).all()
    selected_semester = Semester.query.get(request.args.get('semester_id', type=int)) if request.args.get('semester_id') else get_active_semester()
    if request.method == 'POST':
        db.session.add(CalendarEvent(
            teacher_id=None if current_user.role=='admin' and request.form.get('global_event') else current_user.id,
            event_date=datetime.strptime(request.form.get('event_date',''),'%Y-%m-%d').date(),
            title=request.form.get('title',''), event_type=request.form.get('event_type','กิจกรรมโรงเรียน'), note=request.form.get('note','')
        ))
        db.session.commit(); flash('เพิ่มรายการปฏิทินแล้ว', 'success'); return redirect(url_for('calendar_events', semester_id=selected_semester.id if selected_semester else None) + '#calendar-preview')
    q = CalendarEvent.query if current_user.role=='admin' else CalendarEvent.query.filter((CalendarEvent.teacher_id==current_user.id) | (CalendarEvent.teacher_id==None))
    if selected_semester:
        q = q.filter(CalendarEvent.event_date>=selected_semester.start_date, CalendarEvent.event_date<=selected_semester.end_date)
    events = q.order_by(CalendarEvent.event_date.desc()).all()
    summary = working_day_summary(teacher_id=current_user.id if current_user.role=='teacher' else None, semester=selected_semester)
    images = CalendarImage.query.filter_by(semester_id=selected_semester.id).order_by(CalendarImage.uploaded_at.desc()).all() if selected_semester else []
    cal_months, upcoming, today_position = build_calendar_dashboard(teacher_id=current_user.id if current_user.role=='teacher' else None, semester=selected_semester)
    return render_template('calendar.html', events=events, event=None, summary=summary, semesters=semesters, selected_semester=selected_semester, images=images, cal_months=cal_months, upcoming=upcoming, today_position=today_position, active_semester=selected_semester)

@app.route('/calendar/import-excel', methods=['POST'])
@login_required
@role_required('teacher','admin')
def calendar_import_excel():
    file = request.files.get('file')
    semester_id = request.form.get('semester_id', type=int)
    sem = Semester.query.get(semester_id) if semester_id else get_active_semester()
    if not file or not file.filename:
        flash('กรุณาเลือกไฟล์ Excel', 'danger'); return redirect(url_for('calendar_events', semester_id=semester_id))
    wb = load_workbook(file)
    ws = wb.active
    created = updated = 0
    # รองรับหัวตาราง: date/event_date, title, event_type, note
    headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
    def col(name, fallback):
        for n in name:
            if n in headers: return headers.index(n)
        return fallback
    date_i = col(['date','event_date','วันที่'], 0)
    title_i = col(['title','รายการ','กิจกรรม'], 1)
    type_i = col(['event_type','ประเภท'], 2)
    note_i = col(['note','หมายเหตุ'], 3)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[date_i] or not row[title_i]:
            continue
        raw_date = row[date_i]
        if isinstance(raw_date, datetime):
            d = raw_date.date()
        elif isinstance(raw_date, date):
            d = raw_date
        else:
            txt = str(raw_date).strip().replace('/', '-')
            parts = txt.split('-')
            if len(parts[0]) == 4:
                d = datetime.strptime(txt, '%Y-%m-%d').date()
            else:
                d = datetime.strptime(txt, '%d-%m-%Y').date()
            if d.year > 2400:
                d = date(d.year-543, d.month, d.day)
        title = str(row[title_i]).strip()
        typ = str(row[type_i]).strip() if len(row) > type_i and row[type_i] else 'กิจกรรมโรงเรียน'
        note = str(row[note_i]).strip() if len(row) > note_i and row[note_i] else ''
        teacher_id = None if current_user.role == 'admin' else current_user.id
        e = CalendarEvent.query.filter_by(teacher_id=teacher_id, event_date=d, title=title).first()
        if e:
            e.event_type = typ; e.note = note; updated += 1
        else:
            db.session.add(CalendarEvent(teacher_id=teacher_id, event_date=d, title=title, event_type=typ, note=note)); created += 1
    db.session.commit(); flash(f'นำเข้าปฏิทินแล้ว เพิ่ม {created} รายการ / อัปเดต {updated} รายการ', 'success')
    return redirect(url_for('calendar_events', semester_id=sem.id if sem else None))

@app.route('/calendar/upload-image', methods=['POST'])
@login_required
@role_required('teacher','admin')
def calendar_upload_image():
    file = request.files.get('image')
    semester_id = request.form.get('semester_id', type=int)
    if not file or not file.filename:
        flash('กรุณาเลือกรูปปฏิทิน', 'danger'); return redirect(url_for('calendar_events', semester_id=semester_id))
    filename = secure_filename(file.filename)
    stamp = datetime.now().strftime('%Y%m%d%H%M%S')
    save_name = f'calendar_{stamp}_{filename}'
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], save_name)
    file.save(save_path)
    db.session.add(CalendarImage(semester_id=semester_id, file_path=save_name, original_name=filename))
    db.session.commit()
    flash('บันทึกรูปปฏิทินไว้เป็นหลักฐานแล้ว หากต้องการอัปเดตข้อมูลอัตโนมัติให้ใช้ไฟล์ Excel นำเข้า', 'success')
    return redirect(url_for('calendar_events', semester_id=semester_id))

@app.route('/calendar/seed-1-2569', methods=['POST'])
@login_required
@role_required('teacher','admin')
def calendar_seed_1_2569():
    # admin สร้างเป็นปฏิทินกลาง ทุกครูเห็น / teacher สร้างเฉพาะตัวเอง
    teacher_id = None if current_user.role == 'admin' else current_user.id
    created = seed_academic_calendar_1_2569(teacher_id=teacher_id)
    flash(f'นำเข้าปฏิทินการศึกษา 1/2569 แล้ว เพิ่มใหม่ {created} รายการ', 'success')
    return redirect(url_for('calendar_events'))

@app.route('/calendar/<int:event_id>/edit', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def calendar_edit(event_id):
    event = CalendarEvent.query.get_or_404(event_id)
    if current_user.role!='admin' and event.teacher_id not in (current_user.id, None): return deny_redirect('calendar_events')
    if request.method == 'POST':
        event.event_date=datetime.strptime(request.form.get('event_date',''),'%Y-%m-%d').date(); event.title=request.form.get('title',''); event.event_type=request.form.get('event_type','กิจกรรมโรงเรียน'); event.note=request.form.get('note','')
        db.session.commit(); flash('แก้ไขปฏิทินแล้ว', 'success'); return redirect(url_for('calendar_events'))
    semesters = Semester.query.order_by(Semester.start_date.desc()).all()
    selected_semester = get_active_semester()
    q = CalendarEvent.query if current_user.role=='admin' else CalendarEvent.query.filter((CalendarEvent.teacher_id==current_user.id) | (CalendarEvent.teacher_id==None))
    if selected_semester:
        q = q.filter(CalendarEvent.event_date>=selected_semester.start_date, CalendarEvent.event_date<=selected_semester.end_date)
    events = q.order_by(CalendarEvent.event_date.desc()).all()
    summary = working_day_summary(teacher_id=current_user.id if current_user.role=='teacher' else None, semester=selected_semester)
    images = CalendarImage.query.filter_by(semester_id=selected_semester.id).order_by(CalendarImage.uploaded_at.desc()).all() if selected_semester else []
    cal_months, upcoming, today_position = build_calendar_dashboard(teacher_id=current_user.id if current_user.role=='teacher' else None, semester=selected_semester)
    return render_template('calendar.html', events=events, event=event, summary=summary, semesters=semesters, selected_semester=selected_semester, images=images, cal_months=cal_months, upcoming=upcoming, today_position=today_position, active_semester=selected_semester)

@app.route('/calendar/<int:event_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def calendar_delete(event_id):
    event=CalendarEvent.query.get_or_404(event_id)
    if current_user.role!='admin' and event.teacher_id != current_user.id: return deny_redirect('calendar_events')
    db.session.delete(event); db.session.commit(); flash('ลบรายการปฏิทินแล้ว', 'success')
    return redirect(url_for('calendar_events'))




def ensure_default_classroom_activities(classroom_id):
    """สร้างรายการเช็กชื่อประจำห้องที่ควรมีทุกห้อง เช่น เข้าแถวหน้าเสาธง"""
    default_title = 'เข้าแถวหน้าเสาธง'
    row = ClassroomActivity.query.filter_by(classroom_id=classroom_id, activity_type='assembly', title=default_title).first()
    if not row:
        row = ClassroomActivity(
            classroom_id=classroom_id,
            title=default_title,
            activity_type='assembly',
            target_scope='classroom',
            event_date=None,
            note='เช็กชื่อเข้าแถวทุกวันที่เปิดเรียน',
            is_active=True,
        )
        db.session.add(row)
        db.session.commit()
    return row



def get_activity_scope(activity):
    """กำหนดกลุ่มนักเรียนสำหรับกิจกรรม: ห้องเดียว / รวมทั้งโรงเรียน / ลูกเสือ ม.1-ม.3"""
    explicit = getattr(activity, 'target_scope', None) or 'classroom'
    title = (activity.title or '').strip()
    if explicit in ('all', 'scout_m1_m3', 'classroom'):
        return explicit
    # fallback เผื่อฐานข้อมูลเก่ายังไม่มีค่า target_scope ที่ถูกต้อง
    if 'ลูกเสือ' in title or 'เนตรนารี' in title:
        return 'scout_m1_m3'
    if 'รวม' in title or 'ทั้งหมด' in title:
        return 'all'
    return 'classroom'

def classroom_grade_no(classroom_name):
    """ดึงเลขระดับชั้นจากชื่อห้อง เช่น ม.1/1 -> 1, ม.3/2 -> 3"""
    import re
    text = classroom_name or ''
    m = re.search(r'ม\.?\s*(\d+)', text)
    return int(m.group(1)) if m else None

def activity_student_links(activity, base_room):
    """คืนรายชื่อนักเรียนที่ต้องเช็กชื่อกิจกรรม พร้อมห้องเรียนของนักเรียนแต่ละคน"""
    scope = get_activity_scope(activity)
    q = ClassroomStudent.query.join(Classroom, ClassroomStudent.classroom_id == Classroom.id).join(User, ClassroomStudent.student_id == User.id)
    if scope == 'all':
        q = q.filter(Classroom.is_active == True)
    elif scope == 'scout_m1_m3':
        q = q.filter(Classroom.is_active == True)
    else:
        q = q.filter(ClassroomStudent.classroom_id == base_room.id)
    links = q.order_by(Classroom.name.asc(), User.username.asc(), User.full_name.asc()).all()
    if scope == 'scout_m1_m3':
        links = [l for l in links if classroom_grade_no(l.classroom.name) in (1, 2, 3)]
    # กันนักเรียนซ้ำ ถ้ามีชื่อเดียวกันอยู่หลายห้อง ให้เอาห้องแรกตามลำดับชื่อห้อง
    seen = set()
    out = []
    for l in links:
        if l.student_id in seen:
            continue
        seen.add(l.student_id)
        out.append(l)
    return out

def activity_scope_label(activity):
    scope = get_activity_scope(activity)
    if scope == 'all':
        return 'รวมทั้งหมดทุกห้องเรียน'
    if scope == 'scout_m1_m3':
        return 'ลูกเสือ/เนตรนารี ม.1-ม.3'
    return 'เฉพาะห้องเรียนนี้'

def attendance_scope_for_subject_room(subject, room):
    text = f"{subject.name if subject else ''} {room.name if room else ''}"
    if 'ลูกเสือ' in text or 'เนตรนารี' in text:
        return 'scout_m1_m3'
    if 'รวม' in text or 'ทั้งหมด' in text:
        return 'all'
    return 'classroom'

def attendance_student_links(subject, room):
    """รายชื่อนักเรียนสำหรับเช็กชื่อจากตารางเรียน: ปกติ=ห้องนั้น, กิจกรรมรวม=ทุกห้อง, ลูกเสือ=ม.1-ม.3"""
    scope = attendance_scope_for_subject_room(subject, room)
    q = ClassroomStudent.query.join(Classroom, ClassroomStudent.classroom_id == Classroom.id).join(User, ClassroomStudent.student_id == User.id)
    if scope == 'classroom':
        q = q.filter(ClassroomStudent.classroom_id == room.id)
    else:
        q = q.filter(Classroom.is_active == True)
    links = q.order_by(Classroom.name.asc(), User.username.asc(), User.full_name.asc()).all()
    if scope == 'scout_m1_m3':
        links = [l for l in links if classroom_grade_no(l.classroom.name) in (1, 2, 3)]
    seen = set()
    out = []
    for l in links:
        if l.student_id in seen:
            continue
        seen.add(l.student_id)
        out.append(l)
    return out

def attendance_scope_label(subject, room):
    scope = attendance_scope_for_subject_room(subject, room)
    if scope == 'all':
        return 'กิจกรรมรวม: นักเรียนทั้งหมดทุกห้องเรียน'
    if scope == 'scout_m1_m3':
        return 'กิจกรรมลูกเสือ/เนตรนารี: นักเรียน ม.1-ม.3'
    return f'ห้อง {room.name}'

def build_activity_attendance_report(activity_id, classroom_id, links):
    # สำหรับกิจกรรมรวม ให้ใช้ activity_id เป็นหลัก และใช้ classroom_id ของนักเรียนจริงในแต่ละแถว
    all_rows = ActivityAttendance.query.filter_by(activity_id=activity_id).order_by(ActivityAttendance.date.asc()).all()
    att_dates = sorted({a.date for a in all_rows})
    date_headers = [
        {'date': d, 'iso': d.isoformat(), 'short': thai_date_short(d), 'label': thai_date_label(d)}
        for d in att_dates
    ]
    att_data = {}
    for a in all_rows:
        att_data.setdefault(a.student_id, {})[a.date.isoformat()] = normalize_attendance_status(a.status)

    summary = []
    for link in links:
        student_map = att_data.get(link.student_id, {})
        row = {'student': link.student, 'มา': 0, 'สาย': 0, 'ขาด': 0, 'โดดเรียน': 0, 'ลาป่วย': 0, 'ไปกิจกรรม': 0}
        for st in student_map.values():
            st = normalize_attendance_status(st)
            if st in row:
                row[st] += 1
        total_checked = len(att_dates)
        row['checked'] = total_checked
        attended = row['มา'] + row['สาย'] + row['ไปกิจกรรม']
        row['percent'] = round((attended / total_checked * 100), 1) if total_checked else 100
        summary.append(row)
    return date_headers, att_data, summary

def build_daily_attendance_overview(report_date=None):
    report_date = report_date or local_today()
    allowed_room_ids = teacher_classroom_ids()
    q = Classroom.query.filter(Classroom.is_active == True, Classroom.id.in_(allowed_room_ids or [-1]))
    classrooms = q.order_by(Classroom.name.asc()).all()
    rows = []
    totals = {'students': 0, 'present': 0, 'absent': 0, 'checked_rooms': 0}
    for room in classrooms:
        student_links = ClassroomStudent.query.filter_by(classroom_id=room.id).join(User, ClassroomStudent.student_id == User.id).order_by(User.username.asc(), User.full_name.asc()).all()
        student_ids = [x.student_id for x in student_links]
        student_count = len(student_ids)
        today_att = Attendance.query.filter_by(classroom_id=room.id, date=report_date).filter(Attendance.student_id.in_(student_ids or [-1])).all()
        latest_by_student = {}
        for a in today_att:
            latest_by_student[a.student_id] = normalize_attendance_status(a.status)
        present = sum(1 for st in latest_by_student.values() if st in ['มา', 'สาย', 'ไปกิจกรรม'])
        absent = sum(1 for st in latest_by_student.values() if st in ['ขาด', 'โดดเรียน', 'ลาป่วย'])
        checked = bool(latest_by_student)
        percent = round((present / student_count) * 100, 2) if student_count else 100
        rows.append({
            'room': room, 'student_count': student_count, 'present': present, 'absent': absent,
            'unchecked': max(0, student_count - len(latest_by_student)), 'percent': percent, 'checked': checked
        })
        totals['students'] += student_count
        totals['present'] += present
        totals['absent'] += absent
        if checked:
            totals['checked_rooms'] += 1
    totals['percent'] = round((totals['present'] / totals['students']) * 100, 2) if totals['students'] else 100
    totals['unchecked_rooms'] = max(0, len(classrooms) - totals['checked_rooms'])
    return rows, totals

@app.route('/attendance-dashboard')
@login_required
@role_required('teacher','admin')
def attendance_dashboard():
    report_date = datetime.strptime(request.args.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    rows, totals = build_daily_attendance_overview(report_date)
    period_info = current_period_info(report_date, current_user.id if current_user.role == 'teacher' else None)
    return render_template('attendance_dashboard.html', rows=rows, totals=totals, report_date=report_date, report_day_label=thai_date_label(report_date), today_date=local_today().isoformat(), period_info=period_info)

@app.route('/records')
@login_required
@role_required('teacher','admin')
def records_center():
    subject_ids = teacher_subject_ids()
    room_ids = teacher_classroom_ids()
    pairs = SubjectClassroom.query.filter(
        SubjectClassroom.subject_id.in_(subject_ids or [-1]),
        SubjectClassroom.classroom_id.in_(room_ids or [-1])
    ).join(Subject, SubjectClassroom.subject_id == Subject.id).join(Classroom, SubjectClassroom.classroom_id == Classroom.id).order_by(Subject.name.asc(), Classroom.name.asc()).all()
    if current_user.role == 'admin':
        classrooms_for_daily = Classroom.query.filter_by(is_active=True).order_by(Classroom.name.asc()).all()
    else:
        classrooms_for_daily = Classroom.query.filter(Classroom.id.in_(room_ids or [-1]), Classroom.is_active == True).order_by(Classroom.name.asc()).all()
    range_type, start_date, end_date, label = attendance_range_from_request()
    _, subject_summaries = build_subject_attendance_summaries(start_date, end_date)
    return render_template(
        'records_center.html',
        pairs=pairs, subject_summaries=subject_summaries, classrooms_for_daily=classrooms_for_daily,
        range_type=range_type, start_date=start_date, end_date=end_date, label=label, today_date=local_today().isoformat()
    )

@app.route('/records/subject-summary/print')
@login_required
@role_required('teacher','admin')
def subject_summary_print():
    range_type, start_date, end_date, label = attendance_range_from_request()
    room_rows, subject_rows = build_subject_attendance_summaries(start_date, end_date)
    return render_template('subject_summary_print.html', room_rows=room_rows, subject_rows=subject_rows, label=label, start_date=start_date, end_date=end_date, setting=get_school_setting(), signer=current_user)

@app.route('/records/subject-summary/export')
@login_required
@role_required('teacher','admin')
def subject_summary_export():
    range_type, start_date, end_date, label = attendance_range_from_request()
    room_rows, subject_rows = build_subject_attendance_summaries(start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = 'summary_by_subject'
    ws.append([f'ตารางสรุปการเช็กชื่อเป็นรายวิชา - {label}'])
    ws.append(['รายวิชา', 'ครู', 'ห้องที่เกี่ยวข้อง', 'จำนวนนักเรียนรวม', 'จำนวนคาบ/ครั้งที่เช็ก', 'หน่วยเช็กชื่อรวม', 'มา', 'สาย', 'ขาด', 'โดดเรียน', 'ลาป่วย', 'ไปกิจกรรม', 'มาเรียน (%)'])
    for r in subject_rows:
        ws.append([r['subject'].name, r['teacher'], r['room_text'], r['student_count'], r['checked_slots'], r['total_units'], r['totals']['มา'], r['totals']['สาย'], r['totals']['ขาด'], r['totals']['โดดเรียน'], r['totals']['ลาป่วย'], r['totals']['ไปกิจกรรม'], r['percent']])
    ws2 = wb.create_sheet('summary_by_room')
    ws2.append([f'ตารางสรุปการเช็กชื่อแยกรายวิชา/ห้อง - {label}'])
    ws2.append(['รายวิชา', 'ห้อง/ขอบเขต', 'ครู', 'จำนวนนักเรียน', 'จำนวนคาบ/ครั้งที่เช็ก', 'หน่วยเช็กชื่อรวม', 'มา', 'สาย', 'ขาด', 'โดดเรียน', 'ลาป่วย', 'ไปกิจกรรม', 'มาเรียน (%)'])
    for r in room_rows:
        ws2.append([r['subject'].name, r['scope_label'], r['teacher'], r['student_count'], r['checked_slots'], r['total_units'], r['totals']['มา'], r['totals']['สาย'], r['totals']['ขาด'], r['totals']['โดดเรียน'], r['totals']['ลาป่วย'], r['totals']['ไปกิจกรรม'], r['percent']])
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)
    filename = f"subject_attendance_summary_{range_type}.xlsx"
    path = os.path.join(UPLOAD_DIR, filename)
    wb.save(path)
    return send_file(path, as_attachment=True, download_name=filename)


def schedule_period_block(schedule_id=None, subject_id=None, classroom_id=None, period_no=None):
    """คืนรายการคาบที่ต้องเช็กพร้อมกัน ถ้าคาบติดกัน วิชาเดียว ห้องเดียว ครูเดียว วันเดียวกัน จะเช็กครั้งเดียวแต่บันทึกทุกคาบ"""
    schedule_row = TeachingSchedule.query.get(schedule_id) if schedule_id else None
    if not schedule_row:
        return [int(period_no)] if period_no else []
    rows = TeachingSchedule.query.filter_by(
        teacher_id=schedule_row.teacher_id,
        weekday=schedule_row.weekday,
        subject_id=schedule_row.subject_id,
        classroom_id=schedule_row.classroom_id,
    ).order_by(TeachingSchedule.period_no.asc()).all()
    periods = sorted({r.period_no for r in rows})
    selected = schedule_row.period_no
    block = {selected}
    cur = selected - 1
    while cur in periods:
        block.add(cur); cur -= 1
    cur = selected + 1
    while cur in periods:
        block.add(cur); cur += 1
    return sorted(block)

def make_attendance_slot_headers(subject_id, classroom_id, start_date=None, end_date=None):
    q = Attendance.query.filter_by(subject_id=subject_id, classroom_id=classroom_id)
    if start_date:
        q = q.filter(Attendance.date >= start_date)
    if end_date:
        q = q.filter(Attendance.date <= end_date)
    rows = q.order_by(Attendance.date.asc(), Attendance.period_no.asc()).all()
    seen = []
    used = set()
    for a in rows:
        pno = a.period_no or 0
        key = (a.date, pno)
        if key not in used:
            used.add(key)
            seen.append({
                'date': a.date,
                'period_no': pno,
                'key': f"{a.date.isoformat()}_{pno}",
                'iso': a.date.isoformat(),
                'short': thai_date_short(a.date) + (f" คาบ {pno}" if pno else ''),
                'label': thai_date_label(a.date) + (f" คาบที่ {pno}" if pno else ''),
            })
    return seen, rows

def build_attendance_summary(subject_id, classroom_id, links, start_date=None, end_date=None):
    slot_headers, all_attendance = make_attendance_slot_headers(subject_id, classroom_id, start_date, end_date)
    att_data = {}
    for a in all_attendance:
        key = f"{a.date.isoformat()}_{a.period_no or 0}"
        att_data.setdefault(a.student_id, {})[key] = normalize_attendance_status(a.status)
    summary = []
    for link in links:
        student_map = att_data.get(link.student_id, {})
        row = {'student': link.student, 'มา': 0, 'สาย': 0, 'ขาด': 0, 'โดดเรียน': 0, 'ลาป่วย': 0, 'ไปกิจกรรม': 0}
        for h in slot_headers:
            st = normalize_attendance_status(student_map.get(h['key'], ''))
            if st in row:
                row[st] += 1
        total_checked = len(slot_headers)
        # นับสายเป็นมาเรียนในเปอร์เซ็นต์การมาเรียน แต่ยังแยกจำนวนสายไว้เหมือนเดิม
        attended = row['มา'] + row['สาย'] + row['ไปกิจกรรม']
        row['checked'] = total_checked
        row['percent'] = round((attended / total_checked * 100), 1) if total_checked else 100
        summary.append(row)
    return slot_headers, att_data, summary


def build_subject_attendance_summaries(start_date=None, end_date=None):
    """สรุปภาพรวมการเช็กชื่อ แยกเป็นรายวิชา/ห้องเรียน ใช้ในศูนย์บันทึก รายงานพิมพ์ และ Excel"""
    subject_ids = teacher_subject_ids()
    room_ids = teacher_classroom_ids()
    pairs = SubjectClassroom.query.filter(
        SubjectClassroom.subject_id.in_(subject_ids or [-1]),
        SubjectClassroom.classroom_id.in_(room_ids or [-1])
    ).join(Subject, SubjectClassroom.subject_id == Subject.id).join(Classroom, SubjectClassroom.classroom_id == Classroom.id).order_by(Subject.name.asc(), Classroom.name.asc()).all()

    rows = []
    grouped = {}
    for p in pairs:
        subject = p.subject
        room = p.classroom
        links = attendance_student_links(subject, room)
        date_headers, att_data, summary = build_attendance_summary(subject.id, room.id, links, start_date, end_date)
        totals = {'มา': 0, 'สาย': 0, 'ขาด': 0, 'โดดเรียน': 0, 'ลาป่วย': 0, 'ไปกิจกรรม': 0}
        for srow in summary:
            for key in totals:
                totals[key] += int(srow.get(key, 0) or 0)
        checked_slots = len(date_headers)
        student_count = len(links)
        total_units = checked_slots * student_count
        attended = totals['มา'] + totals['สาย'] + totals['ไปกิจกรรม']
        percent = round((attended / total_units * 100), 1) if total_units else 100
        row = {
            'subject': subject,
            'room': room,
            'teacher': subject.teacher.full_name if getattr(subject, 'teacher', None) else '-',
            'scope_label': attendance_scope_label(subject, room),
            'checked_slots': checked_slots,
            'student_count': student_count,
            'total_units': total_units,
            'totals': totals,
            'percent': percent,
        }
        rows.append(row)
        g = grouped.setdefault(subject.id, {
            'subject': subject,
            'teacher': row['teacher'],
            'rooms': [],
            'checked_slots': 0,
            'student_count': 0,
            'total_units': 0,
            'totals': {'มา': 0, 'สาย': 0, 'ขาด': 0, 'โดดเรียน': 0, 'ลาป่วย': 0, 'ไปกิจกรรม': 0},
            'percent': 100,
        })
        g['rooms'].append(room.name)
        g['checked_slots'] += checked_slots
        g['student_count'] += student_count
        g['total_units'] += total_units
        for key in g['totals']:
            g['totals'][key] += totals[key]

    subject_rows = []
    for g in grouped.values():
        attended = g['totals']['มา'] + g['totals']['สาย'] + g['totals']['ไปกิจกรรม']
        g['percent'] = round((attended / g['total_units'] * 100), 1) if g['total_units'] else 100
        g['room_text'] = ', '.join(sorted(set(g['rooms'])))
        subject_rows.append(g)
    subject_rows.sort(key=lambda x: x['subject'].name)
    return rows, subject_rows

def attendance_range_from_request():
    today = local_today()
    active_sem = Semester.query.filter_by(is_active=True).first()
    range_type = request.args.get('range', 'term')
    if range_type == 'month':
        year = request.args.get('year', type=int) or today.year
        month = request.args.get('month', type=int) or today.month
        start_date = date(year, month, 1)
        end_date = date(year, month, py_calendar.monthrange(year, month)[1])
        label = f"รายเดือน {THAI_MONTHS[month]} {year + 543}"
    else:
        start_date = active_sem.start_date if active_sem else date(today.year, 5, 1)
        end_date = active_sem.end_date if active_sem else today
        label = f"รายเทอม {active_sem.name if active_sem else ''}".strip()
    return range_type, start_date, end_date, label

def classroom_day_schedule_rows(classroom_id, report_date):
    """คาบ 1-8 ของห้องในวันที่เลือก ดึงวิชาจากตารางสอนตามวันในสัปดาห์"""
    weekday = report_date.weekday()
    rows = TeachingSchedule.query.filter_by(classroom_id=classroom_id, weekday=weekday).order_by(TeachingSchedule.period_no.asc()).all()
    by_period = {}
    for r in rows:
        by_period.setdefault(r.period_no, []).append(r)
    return by_period

def build_classroom_day_attendance_report(classroom, report_date):
    links = ClassroomStudent.query.filter_by(classroom_id=classroom.id).join(User, ClassroomStudent.student_id == User.id).order_by(User.username.asc(), User.full_name.asc()).all()
    schedule_by_period = classroom_day_schedule_rows(classroom.id, report_date)
    attendance_rows = Attendance.query.filter_by(classroom_id=classroom.id, date=report_date).filter(Attendance.period_no.in_(list(range(1, 9)))).all()
    attendance_map = {}
    for a in attendance_rows:
        # ถ้ามีหลายวิชาในคาบเดียวกัน ให้เอาข้อมูลล่าสุดที่บันทึกไว้
        attendance_map[(a.student_id, a.period_no or 0)] = normalize_attendance_status(a.status)
    period_headers = []
    for pno in range(1, 9):
        scheds = schedule_by_period.get(pno, [])
        period_headers.append({
            'period_no': pno,
            'schedules': scheds,
            'subject_text': ' / '.join([r.subject.name for r in scheds]) if scheds else 'ว่าง',
            'time_text': f"{scheds[0].start_time}-{scheds[0].end_time}" if scheds else '',
            'has_class': bool(scheds),
            'edit_schedule': scheds[0] if scheds else None,
        })
    summary = []
    for link in links:
        totals = {'มา': 0, 'สาย': 0, 'ขาด': 0, 'โดดเรียน': 0, 'ลาป่วย': 0, 'ไปกิจกรรม': 0}
        checked = 0
        for h in period_headers:
            if not h['has_class']:
                continue
            st = attendance_map.get((link.student_id, h['period_no']), '')
            if st:
                checked += 1
                if st in totals:
                    totals[st] += 1
        attended = totals['มา'] + totals['สาย'] + totals['ไปกิจกรรม']
        percent = round((attended / checked * 100), 1) if checked else 100
        summary.append({'student': link.student, 'checked': checked, 'totals': totals, 'percent': percent})
    return links, period_headers, attendance_map, summary

@app.route('/records/classroom-day')
@login_required
@role_required('teacher','admin')
def classroom_day_report():
    classroom_id = request.args.get('classroom_id', type=int)
    if not classroom_id:
        flash('กรุณาเลือกห้องเรียนก่อนดูสรุปรายวัน', 'danger')
        return redirect(url_for('records_center'))
    room = Classroom.query.get_or_404(classroom_id)
    if current_user.role != 'admin' and not owns_classroom(room):
        return deny_redirect('records_center')
    report_date = datetime.strptime(request.args.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    links, period_headers, attendance_map, summary = build_classroom_day_attendance_report(room, report_date)
    return render_template(
        'classroom_day_report.html',
        room=room, report_date=report_date, report_day_label=thai_date_label(report_date),
        links=links, period_headers=period_headers, attendance_map=attendance_map, summary=summary,
        status_symbols=ATTENDANCE_SYMBOLS, statuses=ATTENDANCE_STATUSES
    )

@app.route('/records/classroom-day/export')
@login_required
@role_required('teacher','admin')
def classroom_day_report_export():
    classroom_id = request.args.get('classroom_id', type=int)
    room = Classroom.query.get_or_404(classroom_id)
    if current_user.role != 'admin' and not owns_classroom(room):
        return deny_redirect('records_center')
    report_date = datetime.strptime(request.args.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    links, period_headers, attendance_map, summary = build_classroom_day_attendance_report(room, report_date)
    wb = Workbook(); ws = wb.active; ws.title = 'classroom_day'
    ws.append([f'สรุปเช็กชื่อรายวัน ห้อง {room.name} - {thai_date_label(report_date)}'])
    ws.append(['ที่', 'รหัส/username', 'ชื่อ-สกุล'] + [f"คาบ {h['period_no']} {h['subject_text']}" for h in period_headers] + ['มา','สาย','ขาด','โดดเรียน','ลาป่วย','ไปกิจกรรม','มาเรียน (%)'])
    for idx, link in enumerate(links, start=1):
        srow = summary[idx-1]
        row = [idx, link.student.username, link.student.full_name]
        for h in period_headers:
            st = attendance_map.get((link.student_id, h['period_no']), '') if h['has_class'] else ''
            row.append(ATTENDANCE_SYMBOLS.get(st, st) if st else ('ว่าง' if not h['has_class'] else ''))
        row += [srow['totals']['มา'], srow['totals']['สาย'], srow['totals']['ขาด'], srow['totals']['โดดเรียน'], srow['totals']['ลาป่วย'], srow['totals']['ไปกิจกรรม'], srow['percent']]
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)
    filename = f"classroom_day_attendance_{room.name}_{report_date.isoformat()}.xlsx".replace('/', '-')
    path = os.path.join(UPLOAD_DIR, filename); wb.save(path)
    return send_file(path, as_attachment=True, download_name=filename)

@app.route('/attendance/<int:subject_id>/<int:classroom_id>', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def attendance(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    substitute_mode = request.args.get('substitute') == '1' or request.form.get('substitute') == '1'
    sub_teacher_id = request.args.get('sub_teacher_id', type=int) or request.form.get('sub_teacher_id', type=int)
    if not substitute_mode and (not owns_subject(subject) or not owns_classroom(room)):
        return deny_redirect('records_center')
    links = attendance_student_links(subject, room)
    attendance_scope = attendance_scope_for_subject_room(subject, room)
    show_classroom_column = attendance_scope in ('all', 'scout_m1_m3')
    att_date = datetime.strptime(request.args.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    period_no = request.args.get('period_no', type=int)
    from_schedule_id = request.args.get('schedule_id', type=int)
    period_block = schedule_period_block(from_schedule_id, subject.id, room.id, period_no)
    if request.method == 'POST':
        att_date = datetime.strptime(request.form.get('date',''), '%Y-%m-%d').date()
        period_no = request.form.get('period_no', type=int) or period_no
        from_schedule_id = request.form.get('schedule_id', type=int) or from_schedule_id
        form_periods = [int(x) for x in request.form.getlist('periods') if str(x).isdigit()]
        period_block = form_periods or schedule_period_block(from_schedule_id, subject.id, room.id, period_no) or ([period_no] if period_no else [0])
        bulk_status = normalize_attendance_status(request.form.get('bulk_status')) if request.form.get('bulk_status') else None
        for link in links:
            st = bulk_status if bulk_status else normalize_attendance_status(request.form.get(f's_{link.student_id}', 'มา'))
            for pno in period_block:
                found = Attendance.query.filter_by(subject_id=subject.id, classroom_id=room.id, student_id=link.student_id, date=att_date, period_no=pno).order_by(Attendance.id.asc()).all()
                row = found[0] if found else None
                if not row:
                    row = Attendance(subject_id=subject.id, classroom_id=room.id, student_id=link.student_id, date=att_date, period_no=pno)
                    db.session.add(row)
                for extra in found[1:]:
                    db.session.delete(extra)
                row.status = st
                row.schedule_id = from_schedule_id
                row.checked_by_id = current_user.id
                row.substitute_for_teacher_id = sub_teacher_id if substitute_mode else None
        db.session.commit(); flash('บันทึกเช็กชื่อเรียบร้อยแล้ว ระบบนับตามจำนวนคาบที่เลือก', 'success')
        return redirect(url_for('attendance', subject_id=subject.id, classroom_id=room.id, date=att_date.isoformat(), period_no=period_no, schedule_id=from_schedule_id, substitute='1' if substitute_mode else None, sub_teacher_id=sub_teacher_id))

    if not period_block:
        period_block = [period_no] if period_no else []
    old = {}
    old_q = Attendance.query.filter_by(subject_id=subject.id, classroom_id=room.id, date=att_date)
    if period_block:
        old_q = old_q.filter(Attendance.period_no.in_(period_block))
    for a in old_q.order_by(Attendance.period_no.asc()).all():
        old.setdefault(a.student_id, normalize_attendance_status(a.status))

    date_headers, att_data, summary = build_attendance_summary(subject.id, room.id, links)
    selected_period_text = ', '.join(str(x) for x in period_block if x) if period_block else ''
    score_q = ClassworkScoreItem.query.filter_by(subject_id=subject.id, classroom_id=room.id, date=att_date)
    if period_block:
        score_q = score_q.filter(ClassworkScoreItem.period_no.in_(period_block))
    classwork_items = score_q.order_by(ClassworkScoreItem.period_no.asc(), ClassworkScoreItem.id.asc()).all()
    classwork_scores = {}
    if classwork_items:
        rows = ClassworkScore.query.filter(ClassworkScore.item_id.in_([x.id for x in classwork_items])).all()
        for row in rows:
            classwork_scores.setdefault(row.item_id, {})[row.student_id] = row
    return render_template(
        'attendance.html',
        subject=subject, room=room, links=links, old=old, att_date=att_date,
        selected_day_label=thai_date_label(att_date), selected_period_text=selected_period_text,
        summary=summary, statuses=ATTENDANCE_STATUSES, period_no=period_no, schedule_id=from_schedule_id,
        period_block=period_block, substitute_mode=substitute_mode, sub_teacher_id=sub_teacher_id,
        date_headers=date_headers, att_data=att_data, status_symbols=ATTENDANCE_SYMBOLS,
        classwork_items=classwork_items, classwork_scores=classwork_scores,
        attendance_scope=attendance_scope, attendance_scope_label=attendance_scope_label(subject, room),
        show_classroom_column=show_classroom_column,
        classrooms_all=Classroom.query.filter(db.or_(Classroom.is_active == True, Classroom.is_active.is_(None))).order_by(Classroom.name).all()
    )


@app.route('/attendance/<int:subject_id>/<int:classroom_id>/class-score', methods=['POST'])
@login_required
@role_required('teacher','admin')
def attendance_class_score(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    substitute_mode = request.form.get('substitute') == '1'
    sub_teacher_id = request.form.get('sub_teacher_id', type=int)
    if not substitute_mode and (not owns_subject(subject) or not owns_classroom(room)):
        return deny_redirect('records_center')
    score_date = datetime.strptime(request.form.get('date', local_today().isoformat()), '%Y-%m-%d').date()
    period_no = request.form.get('period_no', type=int)
    schedule_id = request.form.get('schedule_id', type=int)
    item_id = request.form.get('item_id', type=int)
    title = (request.form.get('title') or '').strip() or f'คะแนนในคาบ {period_no or ""}'.strip()
    try:
        max_score = float(request.form.get('max_score') or 10)
    except Exception:
        max_score = 10
    max_score = max(0.1, max_score)

    if item_id:
        item = ClassworkScoreItem.query.filter_by(id=item_id, subject_id=subject.id, classroom_id=room.id).first_or_404()
        item.title = title
        item.max_score = max_score
        item.period_no = period_no
        item.schedule_id = schedule_id
    else:
        lesson_id = None
        if schedule_id:
            sched = TeachingSchedule.query.get(schedule_id)
            lesson_id = sched.lesson_id if sched else None
        item = ClassworkScoreItem(
            subject_id=subject.id,
            classroom_id=room.id,
            schedule_id=schedule_id,
            lesson_id=lesson_id,
            date=score_date,
            period_no=period_no,
            title=title,
            max_score=max_score,
            created_by_id=current_user.id,
        )
        db.session.add(item)
        db.session.flush()

    links = attendance_student_links(subject, room)
    for link in links:
        raw = request.form.get(f'score_{link.student_id}', '').strip()
        note = request.form.get(f'note_{link.student_id}', '').strip()
        row = ClassworkScore.query.filter_by(item_id=item.id, student_id=link.student_id).first()
        if raw == '' and not note:
            # ว่างไว้ = ยังไม่ให้คะแนน ไม่ลบคะแนนเดิมเพื่อกันพลาด
            continue
        try:
            score = float(raw or 0)
        except Exception:
            score = 0
        score = max(0, min(score, max_score))
        if not row:
            row = ClassworkScore(item_id=item.id, student_id=link.student_id)
            db.session.add(row)
        row.score = score
        row.note = note
    db.session.commit()
    flash('บันทึกคะแนนในคาบแล้ว คะแนนนี้จะถูกรวมในสมุดคะแนนตามน้ำหนัก “คะแนนในคาบ”', 'success')
    return redirect(url_for('attendance', subject_id=subject.id, classroom_id=room.id, date=score_date.isoformat(), period_no=period_no, schedule_id=schedule_id, substitute='1' if substitute_mode else None, sub_teacher_id=sub_teacher_id))

@app.route('/attendance/<int:subject_id>/<int:classroom_id>/class-score/<int:item_id>/delete', methods=['POST'])
@login_required
@role_required('teacher','admin')
def attendance_class_score_delete(subject_id, classroom_id, item_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    if not owns_subject(subject) or not owns_classroom(room):
        return deny_redirect('records_center')
    item = ClassworkScoreItem.query.filter_by(id=item_id, subject_id=subject.id, classroom_id=room.id).first_or_404()
    back_date = item.date
    period_no = item.period_no
    schedule_id = item.schedule_id
    ClassworkScore.query.filter_by(item_id=item.id).delete(synchronize_session=False)
    db.session.delete(item)
    db.session.commit()
    flash('ลบหัวข้อคะแนนในคาบแล้ว', 'success')
    return redirect(url_for('attendance', subject_id=subject.id, classroom_id=room.id, date=back_date.isoformat(), period_no=period_no, schedule_id=schedule_id))

@app.route('/attendance/<int:subject_id>/<int:classroom_id>/delete-day', methods=['POST'])
@login_required
@role_required('teacher','admin')
def attendance_delete_day(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    substitute_mode = request.form.get('substitute') == '1'
    if not substitute_mode and (not owns_subject(subject) or not owns_classroom(room)):
        return deny_redirect('records_center')
    att_date = datetime.strptime(request.form.get('date',''), '%Y-%m-%d').date()
    periods = [int(x) for x in request.form.getlist('periods') if str(x).isdigit()]
    q = Attendance.query.filter_by(subject_id=subject.id, classroom_id=room.id, date=att_date)
    if periods:
        q = q.filter(Attendance.period_no.in_(periods))
    q.delete(synchronize_session=False)
    db.session.commit(); flash('ลบเช็กชื่อของวันที่/คาบที่เลือกแล้ว', 'success')
    return redirect(url_for('attendance', subject_id=subject.id, classroom_id=room.id, date=att_date.isoformat()))

@app.route('/attendance/<int:subject_id>/<int:classroom_id>/export')
@login_required
@role_required('teacher','admin')
def export_attendance(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    substitute_mode = request.args.get('substitute') == '1'
    if not substitute_mode and (not owns_subject(subject) or not owns_classroom(room)):
        return deny_redirect('records_center')
    range_type, start_date, end_date, label = attendance_range_from_request()
    links = attendance_student_links(subject, room)
    date_headers, att_data, summary = build_attendance_summary(subject.id, room.id, links, start_date, end_date)
    wb=Workbook(); ws=wb.active; ws.title='attendance'
    show_classroom_column = attendance_scope_for_subject_room(subject, room) in ('all', 'scout_m1_m3')
    base_header = ['เลขที่','รหัส/username','ชื่อ-สกุล']
    if show_classroom_column:
        base_header.append('ชั้น/ห้อง')
    header = base_header + [h['label'] for h in date_headers] + ['มา','สาย','ขาด','โดดเรียน','ลาป่วย','ไปกิจกรรม','มาเรียน (%)']
    ws.append([f'{subject.name} / {room.name} / {label}'])
    ws.append(header)
    for idx, link in enumerate(links, start=1):
        srow = summary[idx-1]
        row = [idx, link.student.username, link.student.full_name]
        if show_classroom_column:
            row.append(link.classroom.name if link.classroom else '')
        for h in date_headers:
            st = att_data.get(link.student_id, {}).get(h['key'], '')
            row.append(ATTENDANCE_SYMBOLS.get(st, st) if st else '')
        row += [srow['มา'], srow['สาย'], srow['ขาด'], srow['โดดเรียน'], srow['ลาป่วย'], srow['ไปกิจกรรม'], srow['percent']]
        ws.append(row)
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 35)
    filename = f"attendance_{room.name}_{subject.name}_{range_type}.xlsx".replace('/', '-')
    path=os.path.join(UPLOAD_DIR, filename); wb.save(path)
    return send_file(path, as_attachment=True, download_name=filename)

@app.route('/attendance/<int:subject_id>/<int:classroom_id>/print')
@login_required
@role_required('teacher','admin')
def attendance_print(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    substitute_mode = request.args.get('substitute') == '1'
    if not substitute_mode and (not owns_subject(subject) or not owns_classroom(room)):
        return deny_redirect('records_center')
    range_type, start_date, end_date, label = attendance_range_from_request()
    links = attendance_student_links(subject, room)
    date_headers, att_data, summary = build_attendance_summary(subject.id, room.id, links, start_date, end_date)
    return render_template('attendance_print.html', subject=subject, room=room, links=links, date_headers=date_headers, att_data=att_data, summary=summary, status_symbols=ATTENDANCE_SYMBOLS, label=label, start_date=start_date, end_date=end_date, show_classroom_column=(attendance_scope_for_subject_room(subject, room) in ('all', 'scout_m1_m3')), attendance_scope_label=attendance_scope_label(subject, room), setting=get_school_setting(), signer=current_user)

@app.route('/substitute')
@login_required
@role_required('teacher','admin')
def substitute_schedule():
    teachers = active_teachers_query().order_by(User.full_name).all()
    teacher_id = request.args.get('teacher_id', type=int)
    selected_teacher = User.query.get(teacher_id) if teacher_id else None
    day_names = ['จันทร์','อังคาร','พุธ','พฤหัสบดี','ศุกร์','เสาร์','อาทิตย์']
    schedules = []
    if selected_teacher:
        schedules = TeachingSchedule.query.filter_by(teacher_id=selected_teacher.id).order_by(TeachingSchedule.weekday, TeachingSchedule.period_no).all()
    return render_template('substitute.html', teachers=teachers, selected_teacher=selected_teacher, schedules=schedules, day_names=day_names, today_date=local_today().isoformat())

@app.route('/grades/<int:subject_id>/<int:classroom_id>', methods=['GET','POST'])
@login_required
@role_required('teacher','admin')
def grades(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    if not owns_subject(subject) or not owns_classroom(room): return deny_redirect('records_center')
    links = ClassroomStudent.query.filter_by(classroom_id=room.id).join(User, ClassroomStudent.student_id == User.id).order_by(User.username.asc(), User.full_name.asc()).all()
    setting = get_grade_setting(subject.id)
    if request.method == 'POST':
        setting.worksheet_weight = float(request.form.get('worksheet_weight', setting.worksheet_weight))
        setting.quiz_weight = float(request.form.get('quiz_weight', setting.quiz_weight))
        setting.attendance_weight = float(request.form.get('attendance_weight', setting.attendance_weight))
        setting.classwork_weight = float(request.form.get('classwork_weight', getattr(setting, 'classwork_weight', 0) or 0))
        setting.midterm_weight = float(request.form.get('midterm_weight', setting.midterm_weight))
        setting.final_weight = float(request.form.get('final_weight', setting.final_weight))
        for link in links:
            manual = get_manual_score(subject.id, link.student_id)
            manual.midterm = float(request.form.get(f'midterm_{link.student_id}', manual.midterm or 0) or 0)
            manual.final = float(request.form.get(f'final_{link.student_id}', manual.final or 0) or 0)
            manual.behavior = float(request.form.get(f'behavior_{link.student_id}', manual.behavior or 0) or 0)
        db.session.commit(); flash('บันทึกสมุดคะแนนแล้ว', 'success')
        return redirect(url_for('grades', subject_id=subject.id, classroom_id=room.id))
    rows=[]
    for link in links:
        rows.append(calculate_grade_row(subject, room, link.student))
    total_weight = setting.worksheet_weight + setting.quiz_weight + setting.attendance_weight + (getattr(setting, 'classwork_weight', 0) or 0) + setting.midterm_weight + setting.final_weight
    return render_template('grades.html', subject=subject, room=room, rows=rows, setting=setting, total_weight=total_weight)

@app.route('/export_grades/<int:subject_id>/<int:classroom_id>')
@login_required
@role_required('teacher','admin')
def export_grades(subject_id, classroom_id):
    subject = Subject.query.get_or_404(subject_id); room = Classroom.query.get_or_404(classroom_id)
    if not owns_subject(subject) or not owns_classroom(room): return deny_redirect('records_center')
    wb=Workbook(); ws=wb.active; ws.title='grades'
    ws.append(['ชื่อ','แบบทดสอบเฉลี่ย','เรียนจบ/งานทั้งหมด','คะแนนในคาบ','มา','ขาด','โดดเรียน','ลา','มาสาย','กิจกรรม','กลางภาค','ปลายภาค','จิตพิสัย','คะแนนรวม','เกรด'])
    links = ClassroomStudent.query.filter_by(classroom_id=room.id).join(User, ClassroomStudent.student_id == User.id).order_by(User.username.asc(), User.full_name.asc()).all()
    for link in links:
        r=calculate_grade_row(subject, room, link.student)
        ws.append([r['student'].full_name, r['quiz_avg'], f"{r['complete']}/{r['total_assignments']}", f"{r['classwork_raw']}/{r['classwork_max']} ({r['classwork_percent']}%)", r['present'], r['absent'], r['skipped'], r['leave'], r['late'], r['activity'], r['manual'].midterm, r['manual'].final, r['manual'].behavior, r['total'], r['grade']])
    path=os.path.join(UPLOAD_DIR, f'grades_{subject_id}_{classroom_id}.xlsx'); wb.save(path)
    return send_file(path, as_attachment=True)

@app.cli.command('init-db')
def init_db_cmd():
    init_db(); print('initialized')


def seed_prem_schedule_for_teacher(t):
    """สร้างรายวิชา ห้องเรียน และตารางสอนตามภาพตัวอย่างที่ผู้ใช้ให้ไว้"""
    period_times = {
        1: ('08:40','09:30'),
        2: ('09:31','10:20'),
        3: ('10:31','11:20'),
        4: ('11:21','12:10'),
        5: ('13:00','13:50'),
        6: ('13:51','14:40'),
        7: ('14:41','15:30'),
        8: ('15:41','16:00'),
    }

    def get_subject(name):
        sub = Subject.query.filter_by(name=name).first()
        if not sub:
            sub = Subject(name=name, teacher_id=None, credit=1.0, total_periods=40)
            db.session.add(sub); db.session.flush()
            db.session.add(GradeSetting(subject_id=sub.id))
        if not TeacherSubject.query.filter_by(teacher_id=t.id, subject_id=sub.id).first():
            db.session.add(TeacherSubject(teacher_id=t.id, subject_id=sub.id))
        return sub

    def get_room(name):
        room = Classroom.query.filter_by(name=name).first()
        if not room:
            room = Classroom(name=name, teacher_id=None)
            db.session.add(room); db.session.flush()
        if not TeacherClassroom.query.filter_by(teacher_id=t.id, classroom_id=room.id).first():
            db.session.add(TeacherClassroom(teacher_id=t.id, classroom_id=room.id))
        return room

    def link_subject_room(sub, room):
        link_subject_classroom(sub.id, room.id)

    def add_schedule(weekday, period_no, subject_name, room_name, location='', topic=''):
        sub = get_subject(subject_name)
        room = get_room(room_name)
        link_subject_room(sub, room)
        st, en = period_times[period_no]
        exists = TeachingSchedule.query.filter_by(
            teacher_id=t.id, weekday=weekday, period_no=period_no,
            subject_id=sub.id, classroom_id=room.id
        ).first()
        if not exists:
            db.session.add(TeachingSchedule(
                teacher_id=t.id,
                subject_id=sub.id,
                classroom_id=room.id,
                weekday=weekday,
                period_no=period_no,
                start_time=st,
                end_time=en,
                room_name=location or room_name,
                topic=topic or subject_name
            ))

    # จันทร์
    add_schedule(0, 1, 'ว21104 วิทยาการคำนวณ 1', 'ม.1/2-3', 'ห้อง ม.1/2-3')
    add_schedule(0, 3, 'ว22104 วิทยาการคำนวณ 2', 'ม.2/2-3', 'ห้อง ม.2/2-3')
    add_schedule(0, 4, 'ว22104 วิทยาการคำนวณ 2', 'ม.2/1', 'ห้อง ม.2/1')
    add_schedule(0, 5, 'ว21104 วิทยาการคำนวณ 1', 'ม.1/1', 'ห้อง ม.1/1')
    add_schedule(0, 8, 'บำเพ็ญประโยชน์ PLC', 'PLC', 'PLC', 'บำเพ็ญประโยชน์ / PLC')

    # อังคาร
    add_schedule(1, 1, 'ว23104 วิทยาการคำนวณ 3', 'ม.3/2', 'ห้อง ม.3/2')
    add_schedule(1, 3, 'พ21202 เปตอง 2', 'ม.1/3', 'ห้อง ม.1/3')
    add_schedule(1, 4, 'พ21202 เปตอง 2', 'ม.1/3', 'ห้อง ม.1/3')
    add_schedule(1, 8, 'บำเพ็ญประโยชน์ PLC', 'PLC', 'PLC', 'บำเพ็ญประโยชน์ / PLC')

    # พุธ
    add_schedule(2, 3, 'พ23206 คอมพิวเตอร์ 6', 'ม.3/1', 'ห้อง ม.3/1')
    add_schedule(2, 4, 'พ23206 คอมพิวเตอร์ 6', 'ม.3/1', 'ห้อง ม.3/1')
    add_schedule(2, 6, 'ว31103 วิทยาการคำนวณ', 'ม.4/1', 'ห้อง ม.4/1')
    add_schedule(2, 7, 'ลูกเสือเนตรนารี', 'กิจกรรมรวม', 'กิจกรรมรวม', 'ลูกเสือเนตรนารี')
    add_schedule(2, 8, 'บำเพ็ญประโยชน์ PLC', 'PLC', 'PLC', 'บำเพ็ญประโยชน์ / PLC')

    # พฤหัสบดี
    add_schedule(3, 2, 'ว32104 วิทยาการคำนวณ 2', 'ม.5/1-3', 'ห้อง ม.5/1-3')
    add_schedule(3, 3, 'พ23206 คอมพิวเตอร์ 6', 'ม.3/2', 'ห้อง ม.3/2')
    add_schedule(3, 4, 'พ23206 คอมพิวเตอร์ 6', 'ม.3/2', 'ห้อง ม.3/2')
    add_schedule(3, 7, 'ชุมนุม/คณะสี', 'กิจกรรมรวม', 'กิจกรรมรวม', 'ชุมนุม/คณะสี')
    add_schedule(3, 8, 'บำเพ็ญประโยชน์ PLC', 'PLC', 'PLC', 'บำเพ็ญประโยชน์ / PLC')

    # ศุกร์
    add_schedule(4, 3, 'พ22202 เปตอง 4', 'ม.2/3', 'ห้อง ม.2/3')
    add_schedule(4, 4, 'พ22202 เปตอง 4', 'ม.2/3', 'ห้อง ม.2/3')
    add_schedule(4, 5, 'ว23104 วิทยาการคำนวณ 3', 'ม.3/1', 'ห้อง ม.3/1')
    add_schedule(4, 6, 'ว31103 วิทยาการคำนวณ', 'ม.4/2-3', 'ห้อง ม.4/2-3')
    add_schedule(4, 7, 'อบรมคุณธรรม', 'กิจกรรมรวม', 'กิจกรรมรวม', 'อบรมคุณธรรม')
    add_schedule(4, 8, 'บำเพ็ญประโยชน์ PLC', 'PLC', 'PLC', 'บำเพ็ญประโยชน์ / PLC')

def seed():
    if Semester.query.count() == 0:
        db.session.add(Semester(name='ภาคเรียนที่ 1/2569', academic_year='2569', term_no='1', start_date=date(2026,5,11), end_date=date(2026,10,1), is_active=True))
        db.session.commit()
    if not User.query.filter_by(username='admin').first():
        admin=User(username='admin', full_name='ผู้ดูแลระบบ', role='admin'); admin.set_password('1234'); db.session.add(admin)
    if not User.query.filter_by(username='dekchairukna').first():
        t=User(username='dekchairukna', full_name='นายพศิน พิมพ์คำไหล', role='teacher'); t.set_password('1225'); db.session.add(t); db.session.flush()
        room=Classroom(name='ม.1/1', teacher_id=None); db.session.add(room); db.session.flush(); db.session.add(TeacherClassroom(teacher_id=t.id, classroom_id=room.id))
        sub=Subject(name='วิทยาการคำนวณ', teacher_id=None, credit=1.0, total_periods=40); db.session.add(sub); db.session.flush(); db.session.add(TeacherSubject(teacher_id=t.id, subject_id=sub.id)); db.session.add(GradeSetting(subject_id=sub.id))
        u=Unit(subject_id=sub.id, title='หน่วยที่ 1 อัลกอริทึม', indicators='ว 4.2 เข้าใจและใช้อัลกอริทึม', required_periods=3); db.session.add(u); db.session.flush()
        l=Lesson(unit_id=u.id, title='อัลกอริทึมเบื้องต้น', objective='อธิบายความหมายของอัลกอริทึมได้', content='อัลกอริทึมคือขั้นตอนการแก้ปัญหาอย่างเป็นลำดับ'); db.session.add(l); db.session.flush()
        ws=Worksheet(lesson_id=l.id, title='ใบงานอัลกอริทึม', worksheet_type='text', description='ตอบคำถามตามข้อ'); db.session.add(ws); db.session.flush()
        db.session.add(WorksheetQuestion(worksheet_id=ws.id, number=1, question_text='อัลกอริทึมคืออะไร'))
        qz=Quiz(lesson_id=l.id, title='แบบทดสอบท้ายบท', pass_percent=60); db.session.add(qz); db.session.flush()
        db.session.add(QuizQuestion(quiz_id=qz.id, question_text='อัลกอริทึมหมายถึงอะไร', question_type='abcd', choices='ก. ขั้นตอนแก้ปัญหา\nข. เครื่องพิมพ์\nค. จอภาพ\nง. เมาส์', correct_answer='ก'))
    if not User.query.filter_by(username='teacher').first():
        sample=User(username='teacher', full_name='ครูตัวอย่าง', role='teacher'); sample.set_password('1234'); db.session.add(sample); db.session.flush()
    if not User.query.filter_by(username='student').first():
        s=User(username='student', full_name='นักเรียนตัวอย่าง', role='student'); s.set_password('1234'); db.session.add(s); db.session.flush()
        room=Classroom.query.filter_by(name='ม.1/1').first()
        if room: db.session.add(ClassroomStudent(classroom_id=room.id, student_id=s.id))
    t = User.query.filter_by(username='dekchairukna').first()
    # ไม่เติมตารางตัวอย่างซ้ำ ถ้ามีตารางจริงในฐานข้อมูลแล้ว
    if t and TeachingSchedule.query.filter_by(teacher_id=t.id).count() == 0:
        seed_prem_schedule_for_teacher(t)
    # ปฏิทินกลาง 1/2569 ตามภาพตัวอย่าง
    seed_academic_calendar_1_2569(teacher_id=None)
    db.session.commit()


def ensure_schema_columns():
    """อัปเกรด schema โดยไม่ล้างข้อมูลเดิม รองรับทั้ง SQLite และ PostgreSQL"""
    dialect = db.engine.dialect.name

    with db.engine.connect() as conn:
        if dialect == 'sqlite':
            def table_exists(table):
                return bool(conn.exec_driver_sql(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)
                ).fetchone())

            def cols(table):
                if not table_exists(table):
                    return set()
                return {row[1] for row in conn.exec_driver_sql(f"PRAGMA table_info({table})").fetchall()}

            def add_column(table, column, ddl_sqlite, ddl_postgres=None):
                if table_exists(table) and column not in cols(table):
                    conn.exec_driver_sql(f"ALTER TABLE {table} ADD COLUMN {column} {ddl_sqlite}")

            add_column('subject', 'is_active', 'BOOLEAN DEFAULT 1')
            add_column('user', 'position', "VARCHAR(80) DEFAULT 'ครู'")
            add_column('user', 'is_active', 'BOOLEAN DEFAULT 1')
            for col, ddl in [
                ('student_no', "VARCHAR(30) DEFAULT ''"), ('prefix', "VARCHAR(30) DEFAULT ''"),
                ('first_name', "VARCHAR(120) DEFAULT ''"), ('last_name', "VARCHAR(120) DEFAULT ''"),
                ('gender', "VARCHAR(20) DEFAULT ''"), ('nationality', "VARCHAR(80) DEFAULT ''"),
                ('ethnicity', "VARCHAR(80) DEFAULT ''"), ('religion', "VARCHAR(80) DEFAULT ''"),
                ('blood_type', "VARCHAR(10) DEFAULT ''"), ('phone', "VARCHAR(50) DEFAULT ''"),
                ('address', 'TEXT DEFAULT \'\''), ('guardian_name', "VARCHAR(255) DEFAULT ''"),
                ('guardian_phone', "VARCHAR(50) DEFAULT ''"),
            ]:
                add_column('user', col, ddl)
            add_column('classroom', 'is_active', 'BOOLEAN DEFAULT 1')
            add_column('worksheet', 'file_path', "VARCHAR(500) DEFAULT ''")
            add_column('worksheet', 'original_file_name', "VARCHAR(255) DEFAULT ''")
            add_column('worksheet_answer', 'file_path', "VARCHAR(500) DEFAULT ''")
            add_column('worksheet_answer', 'original_file_name', "VARCHAR(255) DEFAULT ''")
            add_column('assignment', 'assignment_type', "VARCHAR(30) DEFAULT 'special'")
            add_column('attendance', 'period_no', 'INTEGER')
            add_column('attendance', 'schedule_id', 'INTEGER')
            add_column('attendance', 'checked_by_id', 'INTEGER')
            add_column('attendance', 'substitute_for_teacher_id', 'INTEGER')
            add_column('grade_setting', 'classwork_weight', 'FLOAT DEFAULT 10')
            add_column('classroom_activity', 'target_scope', "VARCHAR(30) DEFAULT 'classroom'")
            conn.commit()
            return

        # PostgreSQL / อื่น ๆ: ใช้ SQLAlchemy inspector แทน PRAGMA
        inspector = sa_inspect(db.engine)

        def table_exists(table):
            return inspector.has_table(table)

        def cols(table):
            if not table_exists(table):
                return set()
            return {c['name'] for c in inspector.get_columns(table)}

        def q(table):
            # user เป็น reserved/พิเศษใน PostgreSQL ต้อง quote เสมอ
            return f'"{table}"'

        def add_column(table, column, ddl_sqlite, ddl_postgres):
            if table_exists(table) and column not in cols(table):
                conn.exec_driver_sql(f'ALTER TABLE {q(table)} ADD COLUMN {column} {ddl_postgres}')

        add_column('subject', 'is_active', 'BOOLEAN DEFAULT 1', 'BOOLEAN DEFAULT TRUE')
        add_column('user', 'position', "VARCHAR(80) DEFAULT 'ครู'", "VARCHAR(80) DEFAULT 'ครู'")
        add_column('user', 'is_active', 'BOOLEAN DEFAULT 1', 'BOOLEAN DEFAULT TRUE')
        for col, ddl in [
            ('student_no', "VARCHAR(30) DEFAULT ''"), ('prefix', "VARCHAR(30) DEFAULT ''"),
            ('first_name', "VARCHAR(120) DEFAULT ''"), ('last_name', "VARCHAR(120) DEFAULT ''"),
            ('gender', "VARCHAR(20) DEFAULT ''"), ('nationality', "VARCHAR(80) DEFAULT ''"),
            ('ethnicity', "VARCHAR(80) DEFAULT ''"), ('religion', "VARCHAR(80) DEFAULT ''"),
            ('blood_type', "VARCHAR(10) DEFAULT ''"), ('phone', "VARCHAR(50) DEFAULT ''"),
            ('address', "TEXT DEFAULT ''"), ('guardian_name', "VARCHAR(255) DEFAULT ''"),
            ('guardian_phone', "VARCHAR(50) DEFAULT ''"),
        ]:
            add_column('user', col, ddl, ddl)
        add_column('classroom', 'is_active', 'BOOLEAN DEFAULT 1', 'BOOLEAN DEFAULT TRUE')
        add_column('worksheet', 'file_path', "VARCHAR(500) DEFAULT ''", "VARCHAR(500) DEFAULT ''")
        add_column('worksheet', 'original_file_name', "VARCHAR(255) DEFAULT ''", "VARCHAR(255) DEFAULT ''")
        add_column('worksheet_answer', 'file_path', "VARCHAR(500) DEFAULT ''", "VARCHAR(500) DEFAULT ''")
        add_column('worksheet_answer', 'original_file_name', "VARCHAR(255) DEFAULT ''", "VARCHAR(255) DEFAULT ''")
        add_column('assignment', 'assignment_type', "VARCHAR(30) DEFAULT 'special'", "VARCHAR(30) DEFAULT 'special'")
        add_column('attendance', 'period_no', 'INTEGER', 'INTEGER')
        add_column('attendance', 'schedule_id', 'INTEGER', 'INTEGER')
        add_column('attendance', 'checked_by_id', 'INTEGER', 'INTEGER')
        add_column('attendance', 'substitute_for_teacher_id', 'INTEGER', 'INTEGER')
        add_column('grade_setting', 'classwork_weight', 'FLOAT DEFAULT 10', 'DOUBLE PRECISION DEFAULT 10')
        add_column('classroom_activity', 'target_scope', "VARCHAR(30) DEFAULT 'classroom'", "VARCHAR(30) DEFAULT 'classroom'")
        conn.commit()

def sync_schedule_teacher_links():
    """ซ่อมลิงก์มอบหมายครูจากตารางสอนจริง

    ปัญหาที่เจอบ่อย: นำเข้า teaching_schedule แล้วมี teacher_id/subject_id/classroom_id ครบ
    แต่ teacher_subject, teacher_classroom หรือ subject_classroom ไม่มีแถวตามมา
    ทำให้ครูเข้าระบบแล้วหน้า “รายวิชา” หรือ dropdown ห้อง/วิชาไม่ขึ้น
    """
    changed = False
    for row in TeachingSchedule.query.all():
        if not TeacherSubject.query.filter_by(teacher_id=row.teacher_id, subject_id=row.subject_id).first():
            db.session.add(TeacherSubject(teacher_id=row.teacher_id, subject_id=row.subject_id))
            changed = True
        if not TeacherClassroom.query.filter_by(teacher_id=row.teacher_id, classroom_id=row.classroom_id).first():
            db.session.add(TeacherClassroom(teacher_id=row.teacher_id, classroom_id=row.classroom_id))
            changed = True
        if not SubjectClassroom.query.filter_by(subject_id=row.subject_id, classroom_id=row.classroom_id).first():
            db.session.add(SubjectClassroom(subject_id=row.subject_id, classroom_id=row.classroom_id))
            changed = True
        if row.subject and not row.subject.teacher_id:
            row.subject.teacher_id = row.teacher_id
            changed = True
    if changed:
        db.session.commit()

def init_db():
    db.create_all(); ensure_schema_columns(); seed(); sync_schedule_teacher_links()

# ให้ Railway/Gunicorn อัปเกรด schema และซ่อมลิงก์ตารางสอนอัตโนมัติเมื่อแอปเริ่มทำงาน
# ปิดได้ด้วย RUN_STARTUP_INIT=0 ถ้าต้องการจัดการฐานข้อมูลเอง
if os.environ.get('RUN_STARTUP_INIT', '1') == '1':
    with app.app_context():
        init_db()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 8000)))
